import pandas as pd
import os
import io
import sys
import json
import re
import urllib.parse
import time
import hmac
import hashlib
import base64
from datetime import datetime, date

if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

# --- GESTIÓN DE ALMACENAMIENTO PERSISTENTE (RAILWAY VOLUMES & LOCAL) ---
import shutil

DIR_PERSISTENTE = '/app/data' if os.path.isdir('/app/data') else (
    'data' if os.path.isdir('data') else ('/app/data' if os.path.exists('/app') else 'data')
)

try:
    if DIR_PERSISTENTE and not os.path.exists(DIR_PERSISTENTE):
        os.makedirs(DIR_PERSISTENTE, exist_ok=True)
except Exception:
    pass

def ruta_persistente(nombre_archivo: str) -> str:
    """
    Retorna la ruta del archivo dentro del volumen persistente.
    Si el archivo no existe aún en el volumen persistente pero sí en la raíz del proyecto,
    lo inicializa automáticamente copiándolo al volumen para garantizar persistencia continua.
    """
    if not nombre_archivo:
        return nombre_archivo
    
    if DIR_PERSISTENTE and DIR_PERSISTENTE in str(nombre_archivo):
        return nombre_archivo

    nombre_base = os.path.basename(str(nombre_archivo))
    if DIR_PERSISTENTE and os.path.isdir(DIR_PERSISTENTE):
        path_dest = os.path.join(DIR_PERSISTENTE, nombre_base)
        if not os.path.exists(path_dest):
            if os.path.exists(nombre_archivo):
                try:
                    shutil.copy2(nombre_archivo, path_dest)
                except Exception:
                    pass
            elif os.path.exists(nombre_base):
                try:
                    shutil.copy2(nombre_base, path_dest)
                except Exception:
                    pass
        return path_dest
    return nombre_archivo


def safe_print(*args, **kwargs):
    try:
        print(*args, **kwargs)
    except Exception:
        try:
            clean_args = [str(a).encode('ascii', 'ignore').decode('ascii') for a in args]
            print(*clean_args, **kwargs)
        except Exception:
            pass

def limpiar_serie_numerica(series):
    if series is None or series.empty:
        return series
    def _limpiar(val):
        if pd.isna(val):
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        try:
            s = str(val).replace('$', '').replace(',', '').replace(' ', '').strip()
            return float(s)
        except:
            return 0.0
    return series.apply(_limpiar)

def normalizar_columnas(df):
    """
    Limpia y estandariza los nombres de las columnas del DataFrame de Excel,
    corregiendo la codificación de caracteres especiales (ej. 'Cdigo' -> 'Código').
    """
    columnas_limpias = []
    for c in df.columns:
        # Reemplazar el carácter especial de reemplazo \ufffd por 'ó'
        s = str(c).replace('\ufffd', 'ó')
        
        # Mapeos específicos de respaldo
        if 'digo' in s and 'consultora' in s:
            s = 'Código de consultora'
        elif 'digo' in s and 'grupo' in s:
            s = 'Código de grupo'
        elif 'Facturac' in s and '_anterior' not in s:
            if 'Objetivo' in s:
                s = 'Objetivo Facturación'
            elif 'Real' in s:
                s = 'Real Facturación'
            elif 'Cumplimiento' in s:
                s = 'Cumplimiento Facturación'
        columnas_limpias.append(s.strip())
        
    df.columns = columnas_limpias

    # Limpiar automáticamente columnas numéricas conocidas para eliminar el '$' de los textos
    cols_numericas = [
        'Objetivo Facturación', 'Real Facturación', 'Objetivo Activas', 'Real Activas',
        'Cumplimiento Facturación', 'Cumplimiento Activas', 'Saldo', 'Disponibles',
        'Inicios', 'Reinicios', 'Recuperos', 'Inactiva 1', 'Inactiva 2', 'Inactiva 3'
    ]
    for col in cols_numericas:
        if col in df.columns:
            df[col] = limpiar_serie_numerica(df[col])

    return df

def calcular_metas_ciclo(origen='Base para el como vamos.xlsx'):
    """
    Calcula las metas de crecimiento (+1, +3, +5, +7, +9) basándose en las activas reales
    y calcula el 'Avance % Facturación' comparando con la hoja 'Como vamos anterior'.
    """
    if isinstance(origen, str):
        if not os.path.exists(origen):
            print(f"Error: No se encontró el archivo '{origen}'.")
            return None
        print("Iniciando lectura de la base de datos...")
        df = pd.read_excel(origen, sheet_name="Base para el como vamos")
        df = normalizar_columnas(df)
        
        # Eliminar filas 'None', 'nan' o totalmente nulas que dañen la presentación de tablas
        col_nom_ref = 'Nombre de consultora' if 'Nombre de consultora' in df.columns else df.columns[0]
        if col_nom_ref in df.columns:
            mask_valida_df = df[col_nom_ref].notna() & (~df[col_nom_ref].astype(str).str.strip().str.lower().isin(['none', 'nan', '', 'null', '0']))
            df = df[mask_valida_df]
        
        # Verificar si existe la hoja 'Como vamos anterior' para calcular el Avance %
        try:
            xl = pd.ExcelFile(origen)
            hojas_lower = {s.lower().strip(): s for s in xl.sheet_names}
            if 'como vamos anterior' in hojas_lower:
                nombre_hoja_anterior = hojas_lower['como vamos anterior']
                df_prev = pd.read_excel(origen, sheet_name=nombre_hoja_anterior)
                df_prev = normalizar_columnas(df_prev)
                
                col_id = 'Código de consultora' if 'Código de consultora' in df_prev.columns else df_prev.columns[0]
                cols_prev = [c for c in [col_id, 'Cumplimiento Facturación', 'Real Facturación', 'Inactiva 1', 'Inactiva 2', 'Inactiva 3', 'Inactivas 1', 'Inactivas 2', 'Inactivas 3'] if c in df_prev.columns]
                
                if col_id in cols_prev:
                    df_prev_sub = df_prev[cols_prev].drop_duplicates(subset=[col_id])
                    renames_prev = {c_p: f"{c_p}_anterior" for c_p in cols_prev if c_p != col_id}
                    df_prev_sub = df_prev_sub.rename(columns=renames_prev)
                    df = pd.merge(df, df_prev_sub, on=col_id, how='left')
        except Exception as e:
            print(f"Nota: No se pudo cargar la hoja 'Como vamos anterior': {e}")

    elif isinstance(origen, pd.DataFrame):
        df = normalizar_columnas(origen.copy())
    else:
        print("Error: El origen debe ser una ruta de archivo o un DataFrame de pandas.")
        return None
    
    # Nombre de la columna de activas de referencia
    col_activas = 'Real Activas' if 'Real Activas' in df.columns else 'Objetivo Activas'
    
    if col_activas in df.columns:
        # Cálculo de Metas de Crecimiento sobre las Activas Reales
        df['Meta_Crecer_1plus_150k'] = df[col_activas] + 1
        df['Meta_Crecer_3plus_200k'] = df[col_activas] + 3
        df['Meta_Crecer_5plus_300k'] = df[col_activas] + 5
        df['Meta_Crecer_7plus_500k'] = df[col_activas] + 7
        df['Meta_Crecer_9plus_750k'] = df[col_activas] + 9
    else:
        print(f"Advertencia: No se encontró la columna '{col_activas}'. Usando valor 0 como base.")
        df['Meta_Crecer_1plus_150k'] = 1
        df['Meta_Crecer_3plus_200k'] = 3
        df['Meta_Crecer_5plus_300k'] = 5
        df['Meta_Crecer_7plus_500k'] = 7
        df['Meta_Crecer_9plus_750k'] = 9

    # Cálculo del Avance % Facturación (Cumplimiento Actual - Cumplimiento Anterior)
    if 'Cumplimiento Facturación_anterior' in df.columns and 'Cumplimiento Facturación' in df.columns:
        c_actual = pd.to_numeric(df['Cumplimiento Facturación'], errors='coerce')
        c_anterior = pd.to_numeric(df['Cumplimiento Facturación_anterior'], errors='coerce')
        df['Avance % Facturación'] = c_actual - c_anterior
        
    if 'Real Facturación_anterior' in df.columns and 'Real Facturación' in df.columns:
        r_actual = pd.to_numeric(df['Real Facturación'], errors='coerce')
        r_anterior = pd.to_numeric(df['Real Facturación_anterior'], errors='coerce')
        df['Avance Facturación COP'] = r_actual - r_anterior

    # Cálculo de la columna 'Falta para el 100%', 'Falta para el 110%' y 'Cumplimiento Facturación' (en escala 0-100%)
    if 'Objetivo Facturación' in df.columns and 'Real Facturación' in df.columns:
        obj_f = pd.to_numeric(df['Objetivo Facturación'], errors='coerce')
        real_f = pd.to_numeric(df['Real Facturación'], errors='coerce')
        df['Falta para el 100%'] = obj_f - real_f
        df['Brecha Meta 100%'] = df['Falta para el 100%']
        df['Falta para el 110%'] = (obj_f * 1.10) - real_f
        df['Cumplimiento Facturación'] = (real_f / obj_f * 100.0).where(obj_f > 0, 0.0)

    # Cálculo de Productividad (Real Facturación / Real Activas)
    if 'Real Facturación' in df.columns and 'Real Activas' in df.columns:
        real_f = pd.to_numeric(df['Real Facturación'], errors='coerce')
        real_a = pd.to_numeric(df['Real Activas'], errors='coerce')
        df['Productividad'] = (real_f / real_a).fillna(0.0)

    # Cálculo de Cumplimiento Activas (en escala 0-100%)
    if 'Objetivo Activas' in df.columns and 'Real Activas' in df.columns:
        obj_a = pd.to_numeric(df['Objetivo Activas'], errors='coerce')
        real_a = pd.to_numeric(df['Real Activas'], errors='coerce')
        df['Cumplimiento Activas'] = (real_a / obj_a * 100.0).where(obj_a > 0, 0.0)
    elif 'Cumplimiento Activas' in df.columns:
        c_a = pd.to_numeric(df['Cumplimiento Activas'], errors='coerce')
        df['Cumplimiento Activas'] = c_a.apply(lambda v: v * 100.0 if pd.notna(v) and 0 < v <= 2.5 else (v if pd.notna(v) else 0.0))

    # Cálculo dinámico de Activas Frecuentes y % Actividad Frecuente
    if 'Real Activas' in df.columns:
        r_act = pd.to_numeric(df['Real Activas'], errors='coerce').fillna(0.0)
        r_rec = pd.to_numeric(df['Recuperos'], errors='coerce').fillna(0.0) if 'Recuperos' in df.columns else 0.0
        r_ini = pd.to_numeric(df['Inicios'], errors='coerce').fillna(0.0) if 'Inicios' in df.columns else 0.0
        r_rei = pd.to_numeric(df['Reinicios'], errors='coerce').fillna(0.0) if 'Reinicios' in df.columns else 0.0
        df['Activas Frecuentes'] = (r_act - r_rec - r_ini - r_rei).clip(lower=0)
        if 'Disponibles' in df.columns:
            disp_s = pd.to_numeric(df['Disponibles'], errors='coerce').fillna(0.0)
            df['%Actividad Frecuente'] = (df['Activas Frecuentes'] / disp_s.replace(0, pd.NA) * 100.0).fillna(0.0)

    # Cálculo dinámico de Ganancia Estimada según Matriz y Potencializador de Saldo
    df = calcular_ganancia_estimada_df(df)

    # Integrar Metas de Objetivos Arte (Desafíos LNN: Inicios + Reinicios y Recuperos)
    try:
        mapa_arte = cargar_objetivos_arte()
        mapa_grp = mapa_arte.get('por_grupo', {})
        mapa_nom = mapa_arte.get('por_nombre', {})
        
        col_cv_grp = next((c for c in df.columns if 'grupo' in str(c).lower()), None)
        col_cv_nom = next((c for c in df.columns if 'nombre' in str(c).lower() and 'consultora' in str(c).lower()), None)
        
        def _obtener_meta_ini(row):
            g = str(row.get(col_cv_grp, '')).strip().split('.')[0] if col_cv_grp else ''
            nom = str(row.get(col_cv_nom, '')).strip().lower() if col_cv_nom else ''
            target = mapa_grp.get(g) or mapa_nom.get(nom)
            if target:
                return target.get('meta_inicios_reinicios', 0)
            return 0

        def _obtener_meta_rec(row):
            g = str(row.get(col_cv_grp, '')).strip().split('.')[0] if col_cv_grp else ''
            nom = str(row.get(col_cv_nom, '')).strip().lower() if col_cv_nom else ''
            target = mapa_grp.get(g) or mapa_nom.get(nom)
            if target:
                return target.get('meta_recuperos', 0)
            return 0

        def _obtener_meta_disp(row):
            g = str(row.get(col_cv_grp, '')).strip().split('.')[0] if col_cv_grp else ''
            nom = str(row.get(col_cv_nom, '')).strip().lower() if col_cv_nom else ''
            target = mapa_grp.get(g) or mapa_nom.get(nom)
            if target:
                return target.get('disponibles_proyectadas', 0) or target.get('disponibles_esperadas', 0)
            return 0

        def _obtener_desafio_act(row):
            g = str(row.get(col_cv_grp, '')).strip().split('.')[0] if col_cv_grp else ''
            nom = str(row.get(col_cv_nom, '')).strip().lower() if col_cv_nom else ''
            target = mapa_grp.get(g) or mapa_nom.get(nom)
            if target:
                return target.get('desafio_activas', 0)
            return 0

        df['Meta Inicios + Reinicios'] = df.apply(_obtener_meta_ini, axis=1)
        df['Meta Recuperos'] = df.apply(_obtener_meta_rec, axis=1)
        df['Meta Disponibles Esperadas'] = df.apply(_obtener_meta_disp, axis=1)
        df['Desafío Activas Arte'] = df.apply(_obtener_desafio_act, axis=1)
    except Exception as e_arte:
        safe_print(f"Nota al integrar Objetivos Arte en calcular_metas_ciclo: {e_arte}")

    # Si se ejecuta directamente desde archivo, exportamos
    if isinstance(origen, str):
        archivo_salida = 'Resultado_Metas_Procesadas.xlsx'
        try:
            df.to_excel(archivo_salida, index=False)
            safe_print(f"[OK] Resultados guardados exitosamente en: '{archivo_salida}'!")
        except Exception as e:
            safe_print(f"Advertencia al guardar archivo de salida: {e}")

    return df

# --- MÓDULO DE GESTIÓN DE OBJETIVOS ARTE (DESAFÍOS LNN: INICIOS, REINICIOS, RECUPEROS) ---

RUTA_OBJETIVOS_ARTE_JSON = ruta_persistente('objetivos_arte.json')

def cargar_objetivos_arte():
    """
    Carga el diccionario mapeado de Objetivos Arte desde objetivos_arte.json o lo genera desde Objetivos Arte.xlsx si existe.
    """
    if os.path.exists(RUTA_OBJETIVOS_ARTE_JSON):
        try:
            with open(RUTA_OBJETIVOS_ARTE_JSON, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if data and isinstance(data, dict):
                    return data
        except Exception as e:
            safe_print(f"Nota al cargar {RUTA_OBJETIVOS_ARTE_JSON}: {e}")

    # Fallback automático: procesar Objetivos Arte.xlsx si existe localmente
    if os.path.exists('Objetivos Arte.xlsx'):
        res = procesar_archivo_objetivos_arte('Objetivos Arte.xlsx')
        if res.get('exito'):
            return res.get('data', {})
            
    return {}

def guardar_objetivos_arte(dict_data):
    """
    Guarda el diccionario mapeado de Objetivos Arte en objetivos_arte.json.
    """
    try:
        with open(RUTA_OBJETIVOS_ARTE_JSON, 'w', encoding='utf-8') as f:
            json.dump(dict_data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        safe_print(f"Error al guardar {RUTA_OBJETIVOS_ARTE_JSON}: {e}")
        return False

def procesar_archivo_objetivos_arte(origen_arte, ruta_guardar_excel='Objetivos Arte.xlsx'):
    """
    Lee y procesa el archivo Objetivos Arte (.xlsx, .xls), extrayendo las metas de:
    - 'Inicios + reinicios'
    - 'Recuperos'
    de la hoja 'Desafíos LNN'.
    Guarda el mapeo persistente en 'objetivos_arte.json' y actualiza el archivo 'Objetivos Arte.xlsx'.
    """
    try:
        if hasattr(origen_arte, 'read'):
            try:
                origen_arte.seek(0)
            except Exception:
                pass
            with open(ruta_guardar_excel, 'wb') as f_out:
                f_out.write(origen_arte.read())
            try:
                origen_arte.seek(0)
            except Exception:
                pass
            ruta_leer = ruta_guardar_excel
        elif isinstance(origen_arte, str):
            ruta_leer = origen_arte
        else:
            return {'exito': False, 'error': "Formato de origen no soportado."}

        xl = pd.ExcelFile(ruta_leer)
        
        # Buscar hoja Desafíos LNN (insensible a mayúsculas/tildes)
        hoja_des = next((s for s in xl.sheet_names if 'desaf' in s.lower() and 'ln' in s.lower()), None)
        if not hoja_des:
            hoja_des = next((s for s in xl.sheet_names if 'desaf' in s.lower()), xl.sheet_names[0])

        df_raw = xl.parse(hoja_des)
        
        # Detectar cabeceras desplazadas si vienen con 'Unnamed'
        if any('unnamed' in str(c).lower() for c in df_raw.columns[:5]):
            for r_idx in range(min(5, len(df_raw))):
                row_vals = [str(x).lower() for x in df_raw.iloc[r_idx].values if pd.notna(x)]
                if any('sector' in v or 'grupo' in v or 'lider' in v or 'líder' in v for v in row_vals):
                    df_raw.columns = [str(col_name).strip() for col_name in df_raw.iloc[r_idx]]
                    df_raw = df_raw.iloc[r_idx + 1:].reset_index(drop=True)
                    break

        col_grp = next((c for c in df_raw.columns if any(k in str(c).lower() for k in ['cód. grupo', 'cod grupo', 'cód grupo', 'grupo', 'codigo grupo'])), None)
        
        # Priorizar Nombre de Líder exacto evitando 'Cód. Líder'
        col_lider = next((c for c in df_raw.columns if 'nombre' in str(c).lower() and ('lider' in str(c).lower() or 'líder' in str(c).lower())), None)
        if not col_lider:
            col_lider = next((c for c in df_raw.columns if ('lider' in str(c).lower() or 'líder' in str(c).lower()) and 'cód' not in str(c).lower() and 'cod' not in str(c).lower()), None)
            
        col_sec = next((c for c in df_raw.columns if str(c).lower() in ['sector', 'nombre setor', 'nombre sector']), None)
        
        col_ini_meta = next((c for c in df_raw.columns if 'inicios + reinicios' in str(c).lower() or ('inicio' in str(c).lower() and 'meta' in str(c).lower())), None)
        if not col_ini_meta:
            col_ini_meta = next((c for c in df_raw.columns if 'inicios' in str(c).lower() and 'reinicio' in str(c).lower()), None)
        
        col_rec_meta = next((c for c in df_raw.columns if 'recupero' in str(c).lower()), None)
        
        col_disp_esp = next((c for c in df_raw.columns if 'disponibles esperadas' in str(c).lower() and '202612' in str(c).lower()), None)
        if not col_disp_esp:
            col_disp_esp = next((c for c in df_raw.columns if 'disponibles esperadas' in str(c).lower()), None)
            
        col_disp_proy = next((c for c in df_raw.columns if 'disponibles proyectadas' in str(c).lower()), None)
        col_desafio_act = next((c for c in df_raw.columns if 'desafío de activas' in str(c).lower() or 'desafio de activas' in str(c).lower() or 'desafio activas' in str(c).lower() or 'desafío activas' in str(c).lower()), None)
        col_desafio_fact = next((c for c in df_raw.columns if 'desafío facturación' in str(c).lower() or 'desafio facturacion' in str(c).lower() or 'desafio facturación' in str(c).lower()), None)
        col_saldo_meta = next((c for c in df_raw.columns if str(c).strip().lower() == 'saldo'), None)
        col_i1 = next((c for c in df_raw.columns if str(c).strip().lower() == 'inactiva 1'), None)
        col_i2 = next((c for c in df_raw.columns if str(c).strip().lower() == 'inactiva 2'), None)
        col_i3 = next((c for c in df_raw.columns if str(c).strip().lower() == 'inactiva 3'), None)

        if not col_grp and not col_lider:
            return {'exito': False, 'error': "No se identificó la columna de Grupo o Líder en la hoja 'Desafíos LNN'."}

        mapa_por_grupo = {}
        mapa_por_nombre = {}
        sectores_encontrados = set()

        for _, row in df_raw.iterrows():
            g_raw = str(row.get(col_grp, '')).strip().split('.')[0] if col_grp else ''
            nom = str(row.get(col_lider, '')).strip() if col_lider else ''
            sec = str(row.get(col_sec, '')).strip() if col_sec else ''
            if sec:
                sectores_encontrados.add(sec)

            val_ini = int(round(limpiar_numero(row.get(col_ini_meta, 0), 0))) if col_ini_meta else 0
            val_rec = int(round(limpiar_numero(row.get(col_rec_meta, 0), 0))) if col_rec_meta else 0
            val_disp_esp = int(round(limpiar_numero(row.get(col_disp_esp, 0), 0))) if col_disp_esp else 0
            val_disp_proy = int(round(limpiar_numero(row.get(col_disp_proy, 0), 0))) if col_disp_proy else 0
            val_desafio_act = int(round(limpiar_numero(row.get(col_desafio_act, 0), 0))) if col_desafio_act else 0
            val_desafio_fact = float(limpiar_numero(row.get(col_desafio_fact, 0), 0.0)) if col_desafio_fact else 0.0
            val_saldo_meta = int(round(limpiar_numero(row.get(col_saldo_meta, 2), 2))) if col_saldo_meta else 2
            val_i1 = int(round(limpiar_numero(row.get(col_i1, 0), 0))) if col_i1 else 0
            val_i2 = int(round(limpiar_numero(row.get(col_i2, 0), 0))) if col_i2 else 0
            val_i3 = int(round(limpiar_numero(row.get(col_i3, 0), 0))) if col_i3 else 0

            nom_limpio = nom
            if ' - ' in nom_limpio:
                nom_limpio = nom_limpio.split(' - ', 1)[1].strip()

            data_lider = {
                'meta_inicios_reinicios': val_ini,
                'meta_recuperos': val_rec,
                'disponibles_esperadas': val_disp_esp if val_disp_esp > 0 else val_disp_proy,
                'disponibles_proyectadas': val_disp_proy if val_disp_proy > 0 else val_disp_esp,
                'desafio_activas': val_desafio_act,
                'desafio_facturacion': val_desafio_fact,
                'saldo_meta': val_saldo_meta,
                'inactiva_1_meta': val_i1,
                'inactiva_2_meta': val_i2,
                'inactiva_3_meta': val_i3,
                'nombre_lider': nom_limpio,
                'grupo': g_raw,
                'sector': sec
            }

            if g_raw and g_raw not in ['-', 'nan', '']:
                mapa_por_grupo[g_raw] = data_lider
            if nom_limpio and nom_limpio not in ['-', 'nan', '']:
                mapa_por_nombre[nom_limpio.lower()] = data_lider

        dict_final = {
            'por_grupo': mapa_por_grupo,
            'por_nombre': mapa_por_nombre
        }

        guardar_objetivos_arte(dict_final)
        extraer_catalogo_sectores_desde_arte(ruta_leer)

        return {
            'exito': True,
            'total_mapeados': len(mapa_por_grupo),
            'sectores': sorted(list(sectores_encontrados)),
            'data': dict_final
        }
    except Exception as e:
        return {'exito': False, 'error': f"Error al procesar Objetivos Arte: {e}"}

# --- REGLAS DE GANANCIA ESTIMADA SEGÚN MATRIZ Y POTENCIALIZADOR DE SALDO ---

MATRIZ_GANANCIA = [
    # <94.99% Fact | 95-99.99% Fact | 100-109.99% Fact | >=110% Fact
    [0.0150, 0.0160, 0.0170, 0.0180],  # <95% Activas
    [0.0160, 0.0275, 0.0400, 0.0450],  # 95% a 97.49% Activas
    [0.0170, 0.0300, 0.0425, 0.0525],  # 97.5% a 99.99% Activas
    [0.0180, 0.0325, 0.0500, 0.0575],  # 100% a 102.45% Activas
    [0.0190, 0.0340, 0.0525, 0.0600],  # 102.5% a 104.99% Activas
    [0.0200, 0.0355, 0.0550, 0.0625],  # 105% a 109.99% Activas
    [0.0210, 0.0370, 0.0575, 0.0650],  # >=110% Activas
]

ETIQUETAS_ACTIVAS = [
    "Menor a 95%", "95% a 97,49%", "97,5% a 99,99%",
    "100% a 102,45%", "102,5% a 104,99%", "105% a 109,99%", "110% a más"
]

ETIQUETAS_FACTURACION = [
    "Menor a 94,99%", "95% a 99,99%", "100% a 109,99%", "110% a más"
]

def obtener_indice_activas(cump_activas):
    if cump_activas < 0.95:
        return 0
    elif cump_activas < 0.975:
        return 1
    elif cump_activas < 1.00:
        return 2
    elif cump_activas < 1.025:
        return 3
    elif cump_activas < 1.05:
        return 4
    elif cump_activas < 1.10:
        return 5
    else:
        return 6

def obtener_indice_facturacion(cump_fact):
    if cump_fact < 0.95:
        return 0
    elif cump_fact < 1.00:
        return 1
    elif cump_fact < 1.10:
        return 2
    else:
        return 3

def calcular_matriz_ganancia(cump_activas, cump_fact, inicios=6):
    idx_act = obtener_indice_activas(cump_activas)
    idx_fact = obtener_indice_facturacion(cump_fact)
    pct_base = MATRIZ_GANANCIA[idx_act][idx_fact]
    
    # Penalización por Inicios < 6 (-0.5%)
    penalizacion_inicios = 0.0
    if inicios < 6:
        penalizacion_inicios = 0.005
        pct_base = max(0.0, pct_base - penalizacion_inicios)
        
    return pct_base, idx_act, idx_fact

def limpiar_numero(val, default=0.0):
    if pd.isna(val):
        return default
    if isinstance(val, (int, float)):
        return float(val)
    try:
        s = str(val).replace('$', '').replace(',', '').replace(' ', '').strip()
        return float(s)
    except:
        return default

def obtener_potencializador_saldo(saldo):
    s = limpiar_numero(saldo, 0.0)

    if s <= -4:
        return -0.30
    elif s <= -2:
        return -0.25
    elif s == -1:
        return -0.20
    elif s == 0:
        return -0.15
    elif s == 1:
        return -0.05
    elif s == 2:
        return 0.00
    elif s == 3:
        return 0.05
    elif s == 4:
        return 0.10
    elif s == 5:
        return 0.15
    elif s <= 7:
        return 0.20
    elif s <= 9:
        return 0.25
    else:
        return 0.30

def calcular_ganancia_estimada_df(df):
    ganancias_total = []
    ganancias_matriz = []
    pct_matriz_list = []
    potencializador_list = []
    potencializador_cop_list = []

    for _, row in df.iterrows():
        real_fact = limpiar_numero(row.get('Real Facturación', 0.0))
        obj_fact = limpiar_numero(row.get('Objetivo Facturación', 0.0))
        cump_fact = (real_fact / obj_fact) if obj_fact > 0 else 0.0
        
        real_act = limpiar_numero(row.get('Real Activas', 0.0))
        obj_act = limpiar_numero(row.get('Objetivo Activas', 0.0))
        cump_act = (real_act / obj_act) if obj_act > 0 else 0.0
        
        inicios = limpiar_numero(row.get('Inicios', 4.0), 4.0)
        saldo = limpiar_numero(row.get('Saldo', 0.0))
        
        pct_matriz, _, _ = calcular_matriz_ganancia(cump_act, cump_fact, inicios)
        ganancia_matriz_cop = real_fact * pct_matriz
        
        pct_potencializador = obtener_potencializador_saldo(saldo)
        potencializador_cop = ganancia_matriz_cop * pct_potencializador
        
        ganancia_total = ganancia_matriz_cop + potencializador_cop
        
        ganancias_total.append(ganancia_total)
        ganancias_matriz.append(ganancia_matriz_cop)
        pct_matriz_list.append(pct_matriz)
        potencializador_list.append(pct_potencializador)
        potencializador_cop_list.append(potencializador_cop)

    df['Ganancia_Matriz_Pct'] = pct_matriz_list
    df['Ganancia_Matriz_COP'] = ganancias_matriz
    df['Potencializador_Pct'] = potencializador_list
    df['Potencializador_COP'] = potencializador_cop_list
    df['Ganancia estimada'] = ganancias_total
    
    return df

def calcular_bono_lider_mentora(activas_reales, activas_desafio, saldo_real):
    """
    Calcula el Bono Líder Mentora oficial según la matriz de la hoja 'Bono Líder mentora'
    y la tabla Datos!M20:O24.
    Requiere dos habilitadores obligatorios:
    1. % Cumplimiento Activas >= 95%
    2. Saldo Real >= 2
    Escalas de Pago:
    - Menos de 40: $0 COP
    - 40 a 59: $300.000 COP
    - 60 a 79: $400.000 COP
    - 80 a 99: $500.000 COP
    - 100+: $600.000 COP
    """
    try:
        act_r = float(limpiar_numero(activas_reales, 0.0))
        act_d = float(limpiar_numero(activas_desafio, 0.0))
        sal_r = float(limpiar_numero(saldo_real, 0.0))
    except Exception:
        act_r, act_d, sal_r = 0.0, 0.0, 0.0

    pct_act = (act_r / act_d) if act_d > 0 else 0.0
    cumple_act = pct_act >= 0.95
    cumple_sal = sal_r >= 2.0
    cumple_ambos = cumple_act and cumple_sal

    bono_cop = 0.0
    if act_r < 40:
        rango_texto = "Menos de 40 Activas"
        bono_base = 0.0
    elif act_r < 60:
        rango_texto = "40 a 59 Activas"
        bono_base = 300000.0
    elif act_r < 80:
        rango_texto = "60 a 79 Activas"
        bono_base = 400000.0
    elif act_r < 100:
        rango_texto = "80 a 99 Activas"
        bono_base = 500000.0
    else:
        rango_texto = "100 o más Activas"
        bono_base = 600000.0

    if cumple_ambos:
        bono_cop = bono_base
        mensaje = f"🎉 ¡Felicitaciones! Cumples ambos habilitadores. Tu bono estimado es de ${bono_cop:,.0f} COP."
    else:
        faltantes = []
        if not cumple_act:
            falt_act = max(0, int(act_d * 0.95 - act_r) + 1)
            faltantes.append(f"alcanzar el 95% de activas (te faltan {falt_act} activas)")
        if not cumple_sal:
            falt_sal = int(2 - sal_r)
            faltantes.append(f"lograr un saldo de al menos 2 (te faltan {falt_sal} de saldo)")
        mensaje = f"⚠️ Para habilitar el bono de ${bono_base:,.0f} COP requieres: " + " y ".join(faltantes) + "."

    return {
        'pct_alcanzado_activas': pct_act,
        'cumple_activas': cumple_act,
        'cumple_saldo': cumple_sal,
        'cumple_ambos': cumple_ambos,
        'rango_activas': rango_texto,
        'bono_base_escala': bono_base,
        'bono_cop': bono_cop,
        'mensaje_estado': mensaje
    }

def calcular_puntos_convencion_ciclo(saldo_real):
    """
    Calcula los Puntos a Convención por ciclo según la tabla oficial Datos!D47:F52:
    - Saldo < 2: 0 Pts
    - Saldo = 2: 40 Pts
    - Saldo = 3: 60 Pts
    - Saldo = 4: 80 Pts
    - Saldo = 5: 100 Pts
    - Saldo >= 6: 120 Pts
    """
    try:
        s = int(limpiar_numero(saldo_real, 0))
    except Exception:
        s = 0

    if s < 2:
        return 0, "Menor a 2"
    elif s == 2:
        return 40, "2"
    elif s == 3:
        return 60, "3"
    elif s == 4:
        return 80, "4"
    elif s == 5:
        return 100, "5"
    else:
        return 120, "6 o más"

def obtener_diagnostico_retencion_grupo(grupo=None, sector=None):
    """
    Obtiene el conteo real de consultoras por estado comercial (Activa, Inactiva 1 a 6)
    desde la base relacional consultoras_tableau en SQLite.
    """
    conn = obtener_conexion_db()
    cursor = conn.cursor()
    
    query = "SELECT sit_comercial, COUNT(*) FROM consultoras_tableau"
    where = []
    params = []
    if grupo and str(grupo).strip():
        where.append("(grupo = ? OR grupo LIKE ?)")
        params.extend([str(grupo).strip(), f"%{str(grupo).strip()}%"])
    if sector and str(sector).strip():
        where.append("(cod_sector = ? OR sector LIKE ?)")
        params.extend([str(sector).strip(), f"%{str(sector).strip()}%"])
        
    if where:
        query += " WHERE " + " AND ".join(where)
    query += " GROUP BY sit_comercial"
    
    counts = {
        'Activa': 0,
        'Inactiva 1': 0,
        'Inactiva 2': 0,
        'Inactiva 3': 0,
        'Inactiva 4': 0,
        'Inactiva 5': 0,
        'Inactiva 6': 0
    }
    
    try:
        cursor.execute(query, params)
        for r in cursor.fetchall():
            k = str(r[0] or '').strip()
            if k in counts:
                counts[k] = int(r[1])
            elif 'activa' in k.lower() and '1' not in k and '2' not in k:
                counts['Activa'] += int(r[1])
    except Exception:
        pass
    finally:
        conn.close()
        
    disp = counts['Activa'] + counts['Inactiva 1'] + counts['Inactiva 2'] + counts['Inactiva 3']
    pct_act = (counts['Activa'] / disp * 100.0) if disp > 0 else 0.0
    
    return {
        'conteos': counts,
        'disponibles': disp,
        'pct_actividad': pct_act
    }


def generar_analisis_como_vamos(df):
    """
    Realiza un análisis diagnóstico completo y automatizado sobre las métricas clave de
    'Cómo Vamos': Facturación, Saldos, Embudo de Disponibles y Metas de Crecimiento.
    """
    if df is None or df.empty:
        return {}

    df = normalizar_columnas(df.copy())
    
    # 1. Facturación
    obj_fact = float(pd.to_numeric(df['Objetivo Facturación'], errors='coerce').sum()) if 'Objetivo Facturación' in df.columns else 0.0
    real_fact = float(pd.to_numeric(df['Real Facturación'], errors='coerce').sum()) if 'Real Facturación' in df.columns else 0.0
    cump_fact = (real_fact / obj_fact * 100.0) if obj_fact > 0 else 0.0
    meta_110 = obj_fact * 1.10
    brecha_100 = obj_fact - real_fact
    brecha_110 = max(0.0, meta_110 - real_fact)

    # Brecha 100% detallada por líderes
    s_brecha = pd.to_numeric(df['Brecha Meta 100%'], errors='coerce') if 'Brecha Meta 100%' in df.columns else pd.Series(dtype=float)
    faltante_100_lideres = float(s_brecha[s_brecha > 0].sum()) if not s_brecha.empty else 0.0
    superavit_100_lideres = float(s_brecha[s_brecha < 0].abs().sum()) if not s_brecha.empty else 0.0
    
    # Avance % Facturación vs corte anterior
    s_avance_pct = pd.to_numeric(df['Avance % Facturación'], errors='coerce').dropna() if 'Avance % Facturación' in df.columns else pd.Series(dtype=float)
    avance_pct_prom = float(s_avance_pct.mean() * 100.0) if not s_avance_pct.empty else 0.0
    
    s_avance_cop = pd.to_numeric(df['Avance Facturación COP'], errors='coerce').dropna() if 'Avance Facturación COP' in df.columns else pd.Series(dtype=float)
    avance_cop_total = float(s_avance_cop.sum()) if not s_avance_cop.empty else 0.0

    real_activas = float(pd.to_numeric(df['Real Activas'], errors='coerce').sum()) if 'Real Activas' in df.columns else 0.0
    obj_activas = float(pd.to_numeric(df['Objetivo Activas'], errors='coerce').sum()) if 'Objetivo Activas' in df.columns else 0.0
    cump_activas = (real_activas / obj_activas * 100.0) if obj_activas > 0 else 0.0
    productividad_prom = (real_fact / real_activas) if real_activas > 0 else 0.0

    # 2. Saldos
    s_saldo = pd.to_numeric(df['Saldo'], errors='coerce').fillna(0) if 'Saldo' in df.columns else pd.Series([0])
    total_saldo = float(s_saldo.sum())
    lideres_con_saldo = int((s_saldo > 0).sum())
    max_saldo = float(s_saldo.max())
    saldo_prom = float(s_saldo.mean())

    # 3. Embudo de Disponibles
    total_disponibles = float(pd.to_numeric(df['Disponibles'], errors='coerce').sum()) if 'Disponibles' in df.columns else 0.0
    tasa_conversion = (real_activas / total_disponibles * 100.0) if total_disponibles > 0 else 0.0
    pendientes_pedido = max(0.0, total_disponibles - real_activas)
    
    inicios = float(pd.to_numeric(df['Inicios'], errors='coerce').sum()) if 'Inicios' in df.columns else 0.0
    reinicios = float(pd.to_numeric(df['Reinicios'], errors='coerce').sum()) if 'Reinicios' in df.columns else 0.0
    recuperos = float(pd.to_numeric(df['Recuperos'], errors='coerce').sum()) if 'Recuperos' in df.columns else 0.0
    
    inactivas = {}
    for i in range(1, 7):
        col_inact = f'Inactiva {i}'
        inactivas[f'Inactiva_{i}'] = float(pd.to_numeric(df[col_inact], errors='coerce').sum()) if col_inact in df.columns else 0.0
    total_inactivas = sum(inactivas.values())

    # 4. Metas de Crecimiento
    metas = {
        'meta_1plus': float(pd.to_numeric(df['Meta_Crecer_1plus_150k'], errors='coerce').sum()) if 'Meta_Crecer_1plus_150k' in df.columns else 0.0,
        'meta_3plus': float(pd.to_numeric(df['Meta_Crecer_3plus_200k'], errors='coerce').sum()) if 'Meta_Crecer_3plus_200k' in df.columns else 0.0,
        'meta_5plus': float(pd.to_numeric(df['Meta_Crecer_5plus_300k'], errors='coerce').sum()) if 'Meta_Crecer_5plus_300k' in df.columns else 0.0,
        'meta_7plus': float(pd.to_numeric(df['Meta_Crecer_7plus_500k'], errors='coerce').sum()) if 'Meta_Crecer_7plus_500k' in df.columns else 0.0,
        'meta_9plus': float(pd.to_numeric(df['Meta_Crecer_9plus_750k'], errors='coerce').sum()) if 'Meta_Crecer_9plus_750k' in df.columns else 0.0,
    }

    # Estructura del diagnóstico
    analisis = {
        'facturacion': {
            'objetivo': obj_fact,
            'real': real_fact,
            'cumplimiento_pct': cump_fact,
            'meta_110': meta_110,
            'brecha_100': brecha_100,
            'brecha_110': brecha_110,
            'productividad_promedio': productividad_prom,
            'avance_pct_promedio': avance_pct_prom,
            'avance_cop_total': avance_cop_total,
            'faltante_100_lideres': faltante_100_lideres,
            'superavit_100_lideres': superavit_100_lideres
        },
        'activas': {
            'objetivo': obj_activas,
            'real': real_activas,
            'cumplimiento_pct': cump_activas
        },
        'saldos': {
            'total_saldo': total_saldo,
            'lideres_afectados': lideres_con_saldo,
            'max_saldo': max_saldo,
            'saldo_promedio': saldo_prom
        },
        'disponibles': {
            'total_disponibles': total_disponibles,
            'tasa_conversion_pct': tasa_conversion,
            'pendientes_pedido': pendientes_pedido,
            'inicios': inicios,
            'reinicios': reinicios,
            'recuperos': recuperos,
            'inactivas_detalle': inactivas,
            'total_inactivas': total_inactivas,
            'retencion': {
                'inactiva_1': {
                    'total': inactivas.get('Inactiva_1', 0.0),
                    'max_fuga_pct': 12.0,
                    'min_retencion_pct': 88.0,
                    'max_fuga_cant': inactivas.get('Inactiva_1', 0.0) * 0.12,
                    'min_retencion_cant': inactivas.get('Inactiva_1', 0.0) * 0.88,
                },
                'inactiva_2': {
                    'total': inactivas.get('Inactiva_2', 0.0),
                    'max_fuga_pct': 8.0,
                    'min_retencion_pct': 92.0,
                    'max_fuga_cant': inactivas.get('Inactiva_2', 0.0) * 0.08,
                    'min_retencion_cant': inactivas.get('Inactiva_2', 0.0) * 0.92,
                },
                'inactiva_3': {
                    'total': inactivas.get('Inactiva_3', 0.0),
                    'max_fuga_pct': 6.0,
                    'min_retencion_pct': 94.0,
                    'max_fuga_cant': inactivas.get('Inactiva_3', 0.0) * 0.06,
                    'min_retencion_cant': inactivas.get('Inactiva_3', 0.0) * 0.94,
                }
            }
        },
        'metas': metas
    }

def rotar_y_guardar_nuevo_ciclo(nuevo_excel_origen, ruta_destino='Base para el como vamos.xlsx'):
    """
    Toma el archivo de datos del ciclo actual para un sector (o varios),
    actualiza la hoja 'Base para el como vamos' y 'Como vamos anterior' en 'Base para el como vamos.xlsx',
    preservando de manera multi-tenant los datos de otros sectores para que ninguna gerente sobreescriba a otra.
    """
    df_exist_act = None
    df_exist_ant = None
    if os.path.exists(ruta_destino):
        try:
            xl_existente = pd.ExcelFile(ruta_destino)
            hojas_lower = {s.lower().strip(): s for s in xl_existente.sheet_names}
            if 'base para el como vamos' in hojas_lower:
                df_exist_act = pd.read_excel(ruta_destino, sheet_name=hojas_lower['base para el como vamos'])
            if 'como vamos anterior' in hojas_lower:
                df_exist_ant = pd.read_excel(ruta_destino, sheet_name=hojas_lower['como vamos anterior'])
        except Exception as e:
            print(f"Advertencia al leer ciclo anterior de '{ruta_destino}': {e}")

    # Leer el nuevo archivo entregado
    if hasattr(nuevo_excel_origen, 'seek'):
        try:
            nuevo_excel_origen.seek(0)
        except Exception:
            pass

    xl_nuevo = pd.ExcelFile(nuevo_excel_origen)
    hoja_nuevo = xl_nuevo.sheet_names[0]
    for s in xl_nuevo.sheet_names:
        s_low = s.lower()
        if 'base para el como vamos' in s_low:
            hoja_nuevo = s
            break
        elif 'como vamos' in s_low or 'metas' in s_low:
            hoja_nuevo = s
            break

    if hasattr(nuevo_excel_origen, 'seek'):
        try:
            nuevo_excel_origen.seek(0)
        except Exception:
            pass

    df_nuevo = pd.read_excel(nuevo_excel_origen, sheet_name=hoja_nuevo)

    # Identificar columna y valores de sector del nuevo archivo
    def _obtener_col_sector(df_in):
        for c in df_in.columns:
            c_low = str(c).lower().replace('ó', 'o')
            if ('cod' in c_low or 'cd' in c_low or 'cód' in c_low) and ('sector' in c_low or 'setor' in c_low):
                return c
        for c in df_in.columns:
            c_low = str(c).lower().replace('ó', 'o')
            if 'sector' in c_low or 'setor' in c_low:
                return c
        return None

    col_sec_nuevo = _obtener_col_sector(df_nuevo)
    secs_nuevos = set()
    if col_sec_nuevo:
        secs_nuevos = set(df_nuevo[col_sec_nuevo].dropna().astype(str).str.strip().str.replace('.0', '', regex=False).unique())

    # Separar datos de otros sectores vs este sector en el archivo actual
    df_act_otros = pd.DataFrame()
    df_act_este_sector = pd.DataFrame()
    if df_exist_act is not None and not df_exist_act.empty:
        col_sec_exist = _obtener_col_sector(df_exist_act)
        if col_sec_exist and secs_nuevos:
            vals_exist = df_exist_act[col_sec_exist].astype(str).str.strip().str.replace('.0', '', regex=False)
            mask_este = vals_exist.isin(secs_nuevos)
            df_act_este_sector = df_exist_act[mask_este]
            df_act_otros = df_exist_act[~mask_este]
        else:
            df_act_este_sector = df_exist_act

    # Separar datos de otros sectores en el archivo anterior
    df_ant_otros = pd.DataFrame()
    if df_exist_ant is not None and not df_exist_ant.empty:
        col_sec_ant = _obtener_col_sector(df_exist_ant)
        if col_sec_ant and secs_nuevos:
            vals_ant = df_exist_ant[col_sec_ant].astype(str).str.strip().str.replace('.0', '', regex=False)
            df_ant_otros = df_exist_ant[~vals_ant.isin(secs_nuevos)]

    # Construir dataframes consolidados multi-tenant
    df_final_actual = pd.concat([df_act_otros, df_nuevo], ignore_index=True) if not df_act_otros.empty else df_nuevo
    
    componentes_ant = []
    if not df_ant_otros.empty:
        componentes_ant.append(df_ant_otros)
    if not df_act_este_sector.empty:
        componentes_ant.append(df_act_este_sector)
    df_final_anterior = pd.concat(componentes_ant, ignore_index=True) if componentes_ant else None

    # Guardar ambas hojas en el archivo destino
    try:
        with pd.ExcelWriter(ruta_destino, engine='openpyxl') as writer:
            df_final_actual.to_excel(writer, sheet_name='Base para el como vamos', index=False)
            if df_final_anterior is not None and not df_final_anterior.empty:
                df_final_anterior.to_excel(writer, sheet_name='Como vamos anterior', index=False)
    except PermissionError:
        raise PermissionError(f"El archivo '{ruta_destino}' está abierto en Microsoft Excel. Por favor ciérralo en Excel y vuelve a presionar el botón.")
            
    safe_print(f"[OK] ¡Ciclo rotado exitosamente en '{ruta_destino}' preservando datos multi-sector!")
    return calcular_metas_ciclo(ruta_destino)

# --- FUNCIONES DE FORMATO CONDICIONAL Y EXPORTACIÓN A COLORES ---

def color_cumplimiento(val):
    """
    Retorna estilo CSS para formatear celdas de porcentaje según semáforo:
    <95%: Rojo (#FEE2E2), 95-99.99%: Amarillo (#FEF9C3), 100-109.99%: Verde (#DCFCE7), >=110%: Morado (#F3E8FF).
    """
    try:
        if pd.isna(val):
            return ""
        s = str(val)
        for char in ['%', '+', '➡️', '⬆️', '⬇️', '$', ' ']:
            s = s.replace(char, '')
        if not s:
            return ""
        num = float(s)
            
        if num < 95.0:
            return 'background-color: #FEE2E2; color: #991B1B; font-weight: bold;'
        elif num < 100.0:
            return 'background-color: #FEF9C3; color: #854D0E; font-weight: bold;'
        elif num < 110.0:
            return 'background-color: #DCFCE7; color: #166534; font-weight: bold;'
        else:
            return 'background-color: #F3E8FF; color: #6B21A8; font-weight: bold;'
    except:
        return ""

def color_avance(val):
    """
    Retorna estilo CSS para Avance %:
    > 0%: Verde (#DCFCE7), < 0%: Rojo (#FEE2E2), 0%: Neutro.
    """
    try:
        if pd.isna(val):
            return ""
        s = str(val)
        for char in ['%', '+', '➡️', '⬆️', '⬇️', '$', ' ']:
            s = s.replace(char, '')
        if not s:
            return ""
        num = float(s)
        if num > 0:
            return 'background-color: #DCFCE7; color: #166534; font-weight: bold;'
        elif num < 0:
            return 'background-color: #FEE2E2; color: #991B1B; font-weight: bold;'
        else:
            return ''
    except:
        return ""

def color_saldo(val):
    """
    Retorna estilo CSS para formatear celdas de saldo:
    Saldo negativo (< 0): Rojo (#FEE2E2), Saldo >= 0: Verde (#DCFCE7).
    """
    try:
        if pd.isna(val):
            return ""
        s = str(val).replace('$', '').replace(',', '').replace(' ', '').strip()
        num = float(s)
        if num < 0:
            return 'background-color: #FEE2E2; color: #991B1B; font-weight: bold;'
        else:
            return 'background-color: #DCFCE7; color: #166534; font-weight: bold;'
    except:
        return ""

def exportar_excel_con_colores(df_dict, buffer_salida=None):
    """
    Genera un archivo Excel profesional (.xlsx) aplicando colores reales en celdas (openpyxl PatternFill).
    df_dict: dict de {sheet_name: dataframe}
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Eliminar hoja por defecto

    # Estilos profesionales
    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    fill_green = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    font_green = Font(name="Calibri", size=10, bold=True, color="166534")
    
    fill_yellow = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    font_yellow = Font(name="Calibri", size=10, bold=True, color="854D0E")
    
    fill_red = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    font_red = Font(name="Calibri", size=10, bold=True, color="991B1B")
    
    fill_purple = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")
    font_purple = Font(name="Calibri", size=10, bold=True, color="6B21A8")

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    for sheet_name, df in df_dict.items():
        ws = wb.create_sheet(title=str(sheet_name)[:31])
        
        headers = list(df.columns)
        ws.append(headers)
        
        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for _, row in df.iterrows():
            row_vals = [row[c] for c in headers]
            ws.append(row_vals)
            r_idx = ws.max_row
            
            for col_idx, col_name in enumerate(headers, 1):
                cell = ws.cell(row=r_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
                
                val = row[col_name]
                col_lower = str(col_name).lower()
                
                if 'cumplimiento' in col_lower or 'avance' in col_lower:
                    try:
                        clean_str = str(val).replace('%', '').replace('+', '').strip()
                        num = float(clean_str)
                        if 0 < abs(num) <= 2.5:
                            num = num * 100.0
                        
                        if num < 90.0:
                            cell.fill = fill_red
                            cell.font = font_red
                        elif num < 100.0:
                            cell.fill = fill_yellow
                            cell.font = font_yellow
                        elif num < 110.0:
                            cell.fill = fill_green
                            cell.font = font_green
                        else:
                            cell.fill = fill_purple
                            cell.font = font_purple
                    except (ValueError, TypeError):
                        pass
                elif 'saldo' in col_lower:
                    try:
                        clean_s = str(val).replace('$', '').replace(',', '').strip()
                        num = float(clean_s)
                        if num < 0:
                            cell.fill = fill_red
                            cell.font = font_red
                        else:
                            cell.fill = fill_green
                            cell.font = font_green
                    except (ValueError, TypeError):
                        pass

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    if buffer_salida is not None:
        wb.save(buffer_salida)
        return buffer_salida
    else:
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

COLUMNAS_ORDEN_TABLEAU = [
    'Código CB',
    'Asesora / Consultora',
    'DocumentoGPP',
    'Nivel / Color',
    'Sit. Comercial',
    'Pts Acum',
    'Pts Mant',
    'Pts Asc',
    'Deuda Total',
    'Deuda Mora',
    'Credito Total',
    'Credito Disponible',
    'Pts Natura',
    'Pts AVON',
    'Ped. Pendientes',
    'Celular',
    'Líder / Grupo',
    'Notas / Comentarios Líder'
]

def limpiar_y_ordenar_columnas_tableau(df_raw, mapa_lideres=None, es_lider=False):
    """
    Estandariza cualquier DataFrame de Tableau para que conserve exactamente el mismo orden
    y cantidad de columnas limpias de la tabla 'Base de Datos' / 'Base Maestra Gestionable',
    eliminando información repetida, columnas duplicadas o campos técnicos.
    Para el perfil de Líderes, omite 'Líder / Grupo' y ubica DocumentoGPP tras Asesora y Celular tras Ped. Pendientes.
    """
    if df_raw is None or df_raw.empty:
        cols_base = [c for c in COLUMNAS_ORDEN_TABLEAU if not (es_lider and c == 'Líder / Grupo')]
        return pd.DataFrame(columns=cols_base)

    df = df_raw.copy()

    if mapa_lideres is None:
        mapa_lideres = obtener_mapa_lideres()

    # 1. Enriquecer columna 'Líder / Grupo'
    if not es_lider:
        if 'Líder / Grupo' not in df.columns:
            col_grp = next((c for c in ['Grupo', 'grupo', 'codigo_grupo', 'Código de grupo'] if c in df.columns), None)
            if col_grp:
                df['Líder / Grupo'] = df[col_grp].apply(
                    lambda g: f"Grupo {g} — {mapa_lideres[str(g).strip()]}" if str(g).strip() in mapa_lideres else f"Grupo {g}"
                )
            else:
                df['Líder / Grupo'] = ''

    # 2. Diccionario de normalización a nombres de cabecera canónicos
    rename_dict = {}
    
    # Código CB
    if 'Código CB' not in df.columns:
        c_cb = next((c for c in ['Codigo CB', 'codigo_cb', 'Código de consultora', 'Cd Consultora'] if c in df.columns), None)
        if c_cb: rename_dict[c_cb] = 'Código CB'

    # Asesora / Consultora
    if 'Asesora / Consultora' not in df.columns:
        c_nom = next((c for c in ['Nombre', 'nombre', 'Nombre de consultora', 'Consultora'] if c in df.columns), None)
        if c_nom: rename_dict[c_nom] = 'Asesora / Consultora'

    # DocumentoGPP
    if 'DocumentoGPP' not in df.columns:
        c_doc = next((c for c in ['DocumentoGPP', 'documentogpp', 'Documento GPP', 'Documento', 'Cedula', 'Cédula'] if c in df.columns), None)
        if c_doc: rename_dict[c_doc] = 'DocumentoGPP'

    # Nivel / Color
    if 'Nivel / Color' not in df.columns:
        c_col = next((c for c in ['Color', 'color', 'Nivel', 'nivel'] if c in df.columns), None)
        if c_col: rename_dict[c_col] = 'Nivel / Color'

    # Sit. Comercial
    if 'Sit. Comercial' not in df.columns:
        c_sit = next((c for c in ['sit_comercial', 'Sit Comercial', 'Situación', 'situacion', 'Situacion'] if c in df.columns), None)
        if c_sit: rename_dict[c_sit] = 'Sit. Comercial'

    # Puntos
    if 'Pts Acum' not in df.columns:
        c_pts_ac = next((c for c in ['pts_acum', 'Pts Acumulados'] if c in df.columns), None)
        if c_pts_ac: rename_dict[c_pts_ac] = 'Pts Acum'

    if 'Pts Mant' not in df.columns:
        c_pts_mt = next((c for c in ['pts_mant', 'Pts Para Mantener'] if c in df.columns), None)
        if c_pts_mt: rename_dict[c_pts_mt] = 'Pts Mant'

    if 'Pts Asc' not in df.columns:
        c_pts_as = next((c for c in ['pts_asc', 'Pts para Ascender', 'Pts para Ascender '] if c in df.columns), None)
        if c_pts_as: rename_dict[c_pts_as] = 'Pts Asc'

    # Deuda y Crédito
    if 'Deuda Total' not in df.columns:
        c_dt = next((c for c in ['deuda_total'] if c in df.columns), None)
        if c_dt: rename_dict[c_dt] = 'Deuda Total'

    if 'Deuda Mora' not in df.columns:
        c_dm = next((c for c in ['deuda_mora', 'Deuda Mora '] if c in df.columns), None)
        if c_dm: rename_dict[c_dm] = 'Deuda Mora'

    if 'Credito Total' not in df.columns:
        c_ct = next((c for c in ['credito_total', 'Crédito Total'] if c in df.columns), None)
        if c_ct: rename_dict[c_ct] = 'Credito Total'

    if 'Credito Disponible' not in df.columns:
        c_cd = next((c for c in ['credito_disponible', 'Crédito Disponible'] if c in df.columns), None)
        if c_cd: rename_dict[c_cd] = 'Credito Disponible'

    if 'Pts Natura' not in df.columns:
        c_pn = next((c for c in ['pts_natura'] if c in df.columns), None)
        if c_pn: rename_dict[c_pn] = 'Pts Natura'

    if 'Pts AVON' not in df.columns:
        c_pa = next((c for c in ['pts_avon'] if c in df.columns), None)
        if c_pa: rename_dict[c_pa] = 'Pts AVON'

    if 'Ped. Pendientes' not in df.columns:
        c_pp = next((c for c in ['pedidos_pendientes', 'Pedidos Pendientes'] if c in df.columns), None)
        if c_pp: rename_dict[c_pp] = 'Ped. Pendientes'

    # Celular
    if 'Celular' not in df.columns:
        c_cel = next((c for c in ['celular', 'Celular', 'telefono', 'Teléfono', 'Telefono', 'Telefono Celular'] if c in df.columns), None)
        if c_cel: rename_dict[c_cel] = 'Celular'

    if 'Notas / Comentarios Líder' not in df.columns:
        c_nl = next((c for c in ['Comentarios_Lider', 'notas_lider', 'Notas / Comentarios'] if c in df.columns), None)
        if c_nl: rename_dict[c_nl] = 'Notas / Comentarios Líder'

    if rename_dict:
        df = df.rename(columns=rename_dict)

    # Limpiar formato de texto para DocumentoGPP y Celular
    if 'DocumentoGPP' in df.columns:
        df['DocumentoGPP'] = df['DocumentoGPP'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip().replace({'nan': '', 'None': ''})

    if 'Celular' in df.columns:
        df['Celular'] = df['Celular'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip().replace({'nan': '', 'None': ''})

    # Eliminar duplicados de columnas
    df = df.loc[:, ~df.columns.duplicated()].copy()

    # Seleccionar orden de columnas objetivo
    cols_objetivo = [c for c in COLUMNAS_ORDEN_TABLEAU if not (es_lider and c == 'Líder / Grupo')]
    cols_existentes = [c for c in cols_objetivo if c in df.columns]
    df_resultado = df[cols_existentes].copy()

    return df_resultado

def obtener_base_tableau_completa_original(grupo=None, sector=None, cbs_filtrados=None):
    """
    Retorna el DataFrame completo de Tableau con las 61+ columnas originales en su orden exacto,
    incluyendo los comentarios persistentes de la líder y aplicando los filtros y ordenamiento correspondientes.
    """
    if os.path.exists('Base de Datos.xlsx'):
        try:
            df = pd.read_excel('Base de Datos.xlsx')
            
            # Detectar y promover cabecera real si viene desplazada
            if any('unnamed' in str(c).lower() for c in df.columns[:5]):
                for r_idx in range(min(10, len(df))):
                    row_vals = [str(x).lower() for x in df.iloc[r_idx].values if pd.notna(x)]
                    if any('codigo' in x or 'código' in x or 'cb' in x for x in row_vals):
                        df.columns = [str(col_name).strip() for col_name in df.iloc[r_idx]]
                        df = df.iloc[r_idx + 1:].reset_index(drop=True)
                        break

            # Limpiar codificaciones comunes en encabezados
            df = df.rename(columns=lambda c: str(c).replace('Situacin', 'Situación').replace('Mes Cumpleaos', 'Mes Cumpleaños').replace('Direccin', 'Dirección'))
            
            col_cb = next((c for c in ['Codigo CB', 'Código CB', 'codigo_cb'] if c in df.columns), df.columns[0])
            comentarios = cargar_comentarios_lideres()
            
            df['cb_clean_tmp'] = df[col_cb].apply(limpiar_codigo_cb_estandar)
            df['Notas / Comentarios Líder'] = df['cb_clean_tmp'].map(lambda k: comentarios.get(k, ''))
            
            # 1. Si se pasan CBs filtrados directamente de la consulta activa en pantalla
            if cbs_filtrados is not None:
                cbs_clean_list = [limpiar_codigo_cb_estandar(x) for x in cbs_filtrados if limpiar_codigo_cb_estandar(x) != '']
                cbs_set = set(cbs_clean_list)
                df = df[df['cb_clean_tmp'].isin(cbs_set)].copy()
                
                # Mantener estrictamente el orden de los filtros de la pantalla
                order_map = {cb: idx for idx, cb in enumerate(cbs_clean_list)}
                df['__sort_rank__'] = df['cb_clean_tmp'].map(order_map)
                df = df.sort_values(by='__sort_rank__').drop(columns=['cb_clean_tmp', '__sort_rank__'], errors='ignore')
                return df

            # 2. Si se proporciona grupo o sector
            if grupo and str(grupo).strip() and str(grupo).strip() != 'Todas las Líderes (Consolidado Zona)':
                g_str = str(grupo).strip()
                mask_grp = pd.Series(False, index=df.index)
                for c_grp in ['Grupo', 'grupo', 'codigo_grupo', 'Código de grupo', 'Cód. Grupo']:
                    if c_grp in df.columns:
                        mask_grp = mask_grp | df[c_grp].astype(str).str.strip().str.split('.').str[0].str.contains(g_str, case=False, na=False)
                if mask_grp.any():
                    df = df[mask_grp].copy()

            if sector and str(sector).strip() and str(sector).strip() != 'Todos':
                s_str = str(sector).strip()
                mask_sec = pd.Series(False, index=df.index)
                for c_sec in ['Cod. Sector', 'cod_sector', 'Codigo Sector', 'Código Sector']:
                    if c_sec in df.columns:
                        mask_sec = mask_sec | df[c_sec].astype(str).str.strip().str.split('.').str[0].str.contains(s_str, case=False, na=False)
                for c_sec in ['Sector ', 'Sector', 'sector', 'Nombre Setor', 'Nombre Sector']:
                    if c_sec in df.columns:
                        mask_sec = mask_sec | df[c_sec].astype(str).str.strip().str.contains(s_str, case=False, na=False)
                if mask_sec.any():
                    df = df[mask_sec].copy()
                    
            df = df.drop(columns=['cb_clean_tmp'], errors='ignore')
            return df
        except Exception as e:
            safe_print(f"Nota al leer Base de Datos.xlsx para exportación: {e}")

    # Fallback a SQLite
    df_sql = consultar_tableau_sql(grupo=grupo, sector=sector)
    if df_sql is not None and not df_sql.empty:
        col_cb_sql = next((c for c in ['Código CB', 'Codigo CB', 'codigo_cb'] if c in df_sql.columns), df_sql.columns[0])
        df_sql['cb_clean_tmp'] = df_sql[col_cb_sql].apply(limpiar_codigo_cb_estandar)
        if cbs_filtrados is not None:
            cbs_clean_list = [limpiar_codigo_cb_estandar(x) for x in cbs_filtrados if limpiar_codigo_cb_estandar(x) != '']
            cbs_set = set(cbs_clean_list)
            df_sql = df_sql[df_sql['cb_clean_tmp'].isin(cbs_set)].copy()
            order_map = {cb: idx for idx, cb in enumerate(cbs_clean_list)}
            df_sql['__sort_rank__'] = df_sql['cb_clean_tmp'].map(order_map)
            df_sql = df_sql.sort_values(by='__sort_rank__').drop(columns=['cb_clean_tmp', '__sort_rank__'], errors='ignore')
        else:
            df_sql = df_sql.drop(columns=['cb_clean_tmp'], errors='ignore')
    return df_sql

def exportar_tableau_excel_con_colores(df, nombre_hoja="Base_Consultoras", buffer_salida=None):
    """
    Genera un archivo Excel profesional (.xlsx) con los colores exactos de la consulta,
    manteniendo estrictamente el orden y todas las columnas originales de la base de datos.
    Optimizado para alta velocidad (bulk append y pre-indexación de columnas).
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df = df.loc[:, ~df.columns.duplicated()].copy()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(nombre_hoja)[:31]

    # Paleta de colores para niveles
    fill_bronce = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    font_bronce = Font(name="Calibri", size=10, bold=True, color="92400E")
    
    fill_plata = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    font_plata = Font(name="Calibri", size=10, bold=True, color="334155")
    
    fill_oro = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
    font_oro = Font(name="Calibri", size=10, bold=True, color="854D0E")
    
    fill_platino = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
    font_platino = Font(name="Calibri", size=10, bold=True, color="0369A1")
    
    fill_zafiro = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    font_zafiro = Font(name="Calibri", size=10, bold=True, color="1E40AF")
    
    fill_diamante = PatternFill(start_color="F3E8FF", end_color="F3E8FF", fill_type="solid")
    font_diamante = Font(name="Calibri", size=10, bold=True, color="6B21A8")

    # Paleta para situación comercial
    fill_activa = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    font_activa = Font(name="Calibri", size=10, bold=True, color="166534")
    
    fill_inactiva_leve = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
    font_inactiva_leve = Font(name="Calibri", size=10, bold=True, color="854D0E")
    
    fill_inactiva_critica = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    font_inactiva_critica = Font(name="Calibri", size=10, bold=True, color="991B1B")

    fill_mora_media = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")
    font_mora_media = Font(name="Calibri", size=10, bold=True, color="9A3412")

    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    headers = list(df.columns)
    ws.append(headers)

    col_nivel_indices = []
    col_sit_indices = []
    col_mora_indices = []
    col_money_indices = []
    col_int_indices = []

    for col_idx, col_name in enumerate(headers, 1):
        c_low = str(col_name).lower()
        if c_low in ['color', 'nivel', 'nivel / color']:
            col_nivel_indices.append(col_idx)
        elif any(k in c_low for k in ['sit. comercial', 'sit comercial', 'situacion', 'situación']):
            col_sit_indices.append(col_idx)
        elif 'deuda mora' in c_low or c_low.startswith('mora'):
            col_mora_indices.append(col_idx)
        elif any(k in c_low for k in ['deuda', 'credito', 'crédito', 'fact']):
            col_money_indices.append(col_idx)
        elif any(k in c_low for k in ['pts', 'puntos', 'pedidos', 'ped.', 'ciclos']):
            col_int_indices.append(col_idx)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_vals in df.itertuples(index=False):
        ws.append(list(row_vals))

    num_filas = len(df)
    for r_idx in range(2, num_filas + 2):
        for col_idx in col_nivel_indices:
            cell = ws.cell(row=r_idx, column=col_idx)
            v_str = str(cell.value or '').strip().lower()
            if 'bronce' in v_str:
                cell.fill, cell.font = fill_bronce, font_bronce
            elif 'plata' in v_str:
                cell.fill, cell.font = fill_plata, font_plata
            elif 'oro' in v_str:
                cell.fill, cell.font = fill_oro, font_oro
            elif 'platino' in v_str:
                cell.fill, cell.font = fill_platino, font_platino
            elif 'zafiro' in v_str:
                cell.fill, cell.font = fill_zafiro, font_zafiro
            elif 'diamante' in v_str:
                cell.fill, cell.font = fill_diamante, font_diamante

        for col_idx in col_sit_indices:
            cell = ws.cell(row=r_idx, column=col_idx)
            v_str = str(cell.value or '').strip().lower()
            if 'activa' in v_str or 'inicio' in v_str or 'reinicio' in v_str or 'recupero' in v_str:
                cell.fill, cell.font = fill_activa, font_activa
            elif any(x in v_str for x in ['inactiva 1', 'inactiva 2', 'inactiva 3', 'inactiva1', 'inactiva2', 'inactiva3']):
                cell.fill, cell.font = fill_inactiva_leve, font_inactiva_leve
            elif any(x in v_str for x in ['inactiva 4', 'inactiva 5', 'inactiva 6', 'indisponible', 'inactiva4', 'inactiva5', 'inactiva6']):
                cell.fill, cell.font = fill_inactiva_critica, font_inactiva_critica
            elif 'fuga' in v_str:
                cell.fill, cell.font = fill_mora_media, font_mora_media

        for col_idx in col_mora_indices:
            cell = ws.cell(row=r_idx, column=col_idx)
            num = limpiar_numero(cell.value, 0.0)
            if num <= 0:
                cell.fill, cell.font = fill_activa, font_activa
            elif num <= 200000:
                cell.fill, cell.font = fill_inactiva_leve, font_inactiva_leve
            elif num <= 500000:
                cell.fill, cell.font = fill_mora_media, font_mora_media
            else:
                cell.fill, cell.font = fill_inactiva_critica, font_inactiva_critica
            cell.number_format = '$#,##0'

        for col_idx in col_money_indices:
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.number_format = '$#,##0'

        for col_idx in col_int_indices:
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.number_format = '#,##0'

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = 16

    if buffer_salida is not None:
        wb.save(buffer_salida)
        return buffer_salida
    else:
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

def exportar_tabla_ods(df_dict_or_df, buffer_salida=None):
    """
    Exporta un DataFrame o diccionario de DataFrames a formato OpenDocument Spreadsheet (.ods).
    """
    out = buffer_salida if buffer_salida is not None else io.BytesIO()
    with pd.ExcelWriter(out, engine='odf') as writer:
        if isinstance(df_dict_or_df, dict):
            for s_name, df_item in df_dict_or_df.items():
                df_item.to_excel(writer, sheet_name=str(s_name)[:31], index=False)
        else:
            df_dict_or_df.to_excel(writer, sheet_name='Datos', index=False)
    if buffer_salida is not None:
        return buffer_salida
    return out.getvalue()

def exportar_tabla_pdf(df, titulo="Reporte Ejecutivo - Panel Matices", subtitulo="", columnas=None, max_filas=1000, buffer_salida=None):
    """
    Genera un documento PDF profesional horizontal (Landscape) con tabla estilizada,
    encabezados corporativos y semáforo de colores en celdas de estado, nivel y mora.
    """
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    out = buffer_salida if buffer_salida is not None else io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=landscape(letter),
        leftMargin=20,
        rightMargin=20,
        topMargin=25,
        bottomMargin=25
    )

    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=15,
        leading=17,
        textColor=colors.HexColor('#1E293B'),
        spaceAfter=3
    )

    sub_style = ParagraphStyle(
        'SubStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=10
    )

    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7,
        leading=8,
        textColor=colors.HexColor('#1E293B')
    )

    cell_header_style = ParagraphStyle(
        'CellHeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7.5,
        leading=9,
        textColor=colors.white,
        alignment=1
    )

    story.append(Paragraph(f"<b>{titulo}</b>", title_style))
    if subtitulo:
        story.append(Paragraph(subtitulo, sub_style))

    # Seleccionar columnas prioritarias si no se pasan explícitas
    if columnas:
        cols_usar = [c for c in columnas if c in df.columns]
    else:
        cols_prioritarias = [
            'Código CB', 'Líder / Grupo', 'Asesora / Consultora', 'Nivel / Color', 'Sit. Comercial',
            'Pts Acum', 'Deuda Total', 'Deuda Mora', 'Credito Disponible', 'Ped. Pendientes', 'Notas / Comentarios Líder'
        ]
        cols_usar = [c for c in cols_prioritarias if c in df.columns]
        if not cols_usar:
            cols_usar = list(df.columns[:10])

    df_pdf = df[cols_usar].head(max_filas).copy()

    table_data = []
    header_row = [Paragraph(f"<b>{c}</b>", cell_header_style) for c in cols_usar]
    table_data.append(header_row)

    custom_table_styles = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('TOPPADDING', (0, 0), (-1, -1), 2.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
        ('LEFTPADDING', (0, 0), (-1, -1), 2.5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2.5),
    ]

    for r_idx, (_, row) in enumerate(df_pdf.iterrows(), start=1):
        row_cells = []
        for c_idx, col_name in enumerate(cols_usar):
            val = row[col_name]
            val_str = str(val) if pd.notna(val) else ""
            col_lower = str(col_name).lower()

            # Formatear números
            if any(k in col_lower for k in ['deuda', 'credito', 'crédito', 'fact']):
                num = limpiar_numero(val, 0.0)
                val_str = f"${num:,.0f}".replace(",", ".")
            elif any(k in col_lower for k in ['pts', 'puntos', 'pedidos', 'ped.']):
                num = int(limpiar_numero(val, 0))
                val_str = f"{num:,}".replace(",", ".")

            # Colorear celdas según estado / nivel / mora
            if 'color' in col_lower or 'nivel' in col_lower:
                v_low = str(val).lower()
                if 'bronce' in v_low:
                    custom_table_styles.append(('BACKGROUND', (c_idx, r_idx), (c_idx, r_idx), colors.HexColor('#FEF3C7')))
                elif 'plata' in v_low:
                    custom_table_styles.append(('BACKGROUND', (c_idx, r_idx), (c_idx, r_idx), colors.HexColor('#E2E8F0')))
                elif 'oro' in v_low:
                    custom_table_styles.append(('BACKGROUND', (c_idx, r_idx), (c_idx, r_idx), colors.HexColor('#FEF08A')))
                elif 'platino' in v_low:
                    custom_table_styles.append(('BACKGROUND', (c_idx, r_idx), (c_idx, r_idx), colors.HexColor('#E0F2FE')))
                elif 'zafiro' in v_low:
                    custom_table_styles.append(('BACKGROUND', (c_idx, r_idx), (c_idx, r_idx), colors.HexColor('#DBEAFE')))
                elif 'diamante' in v_low:
                    custom_table_styles.append(('BACKGROUND', (c_idx, r_idx), (c_idx, r_idx), colors.HexColor('#F3E8FF')))

            elif 'sit. comercial' in col_lower or 'sit comercial' in col_lower or 'situacion' in col_lower or 'situación' in col_lower:
                v_low = str(val).lower()
                if 'activa' in v_low or 'inicio' in v_low or 'reinicio' in v_low or 'recupero' in v_low:
                    custom_table_styles.append(('BACKGROUND', (c_idx, r_idx), (c_idx, r_idx), colors.HexColor('#DCFCE7')))
                elif any(x in v_low for x in ['inactiva 1', 'inactiva 2', 'inactiva 3']):
                    custom_table_styles.append(('BACKGROUND', (c_idx, r_idx), (c_idx, r_idx), colors.HexColor('#FEF9C3')))
                elif any(x in v_low for x in ['inactiva 4', 'inactiva 5', 'inactiva 6', 'indisponible']):
                    custom_table_styles.append(('BACKGROUND', (c_idx, r_idx), (c_idx, r_idx), colors.HexColor('#FEE2E2')))
                elif 'fuga' in v_low:
                    custom_table_styles.append(('BACKGROUND', (c_idx, r_idx), (c_idx, r_idx), colors.HexColor('#FFEDD5')))

            elif 'deuda mora' in col_lower or 'mora' in col_lower:
                num = limpiar_numero(val, 0.0)
                if num <= 0:
                    custom_table_styles.append(('BACKGROUND', (c_idx, r_idx), (c_idx, r_idx), colors.HexColor('#DCFCE7')))
                elif num <= 200000:
                    custom_table_styles.append(('BACKGROUND', (c_idx, r_idx), (c_idx, r_idx), colors.HexColor('#FEF9C3')))
                elif num <= 500000:
                    custom_table_styles.append(('BACKGROUND', (c_idx, r_idx), (c_idx, r_idx), colors.HexColor('#FFEDD5')))
                else:
                    custom_table_styles.append(('BACKGROUND', (c_idx, r_idx), (c_idx, r_idx), colors.HexColor('#FEE2E2')))

            row_cells.append(Paragraph(val_str[:30], cell_style))
        table_data.append(row_cells)

    disponible_width = 750
    col_w = disponible_width / max(1, len(cols_usar))
    t = Table(table_data, colWidths=[col_w] * len(cols_usar), repeatRows=1)
    t.setStyle(TableStyle(custom_table_styles))
    story.append(t)

    doc.build(story)
    if buffer_salida is not None:
        return buffer_salida
    return out.getvalue()

# --- MÓDULO INFORME TABLEAU MANAGER ("INFORME TABLEAU CAM") ---

RUTA_COMENTARIOS = ruta_persistente('comentarios_lideres.json')

def autocorregir_texto_espanol(texto):
    """
    Normaliza el texto de notas eliminando espacios innecesarios.
    """
    return str(texto).strip() if texto else ""

def cargar_comentarios_lideres():
    """
    Carga el diccionario de comentarios/notas por Código CB desde un JSON persistente.
    """
    if os.path.exists(RUTA_COMENTARIOS):
        try:
            with open(RUTA_COMENTARIOS, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def guardar_comentario_lider(codigo_cb, comentario):
    """
    Guarda o actualiza el comentario de una consultora por su Código CB en el JSON persistente y SQLite.
    """
    comentarios = cargar_comentarios_lideres()
    codigo_clean = limpiar_codigo_cb_estandar(codigo_cb)
    if not codigo_clean:
        return False
    nota_limpia = str(comentario).strip()
    if nota_limpia:
        comentarios[codigo_clean] = nota_limpia
    elif codigo_clean in comentarios:
        comentarios.pop(codigo_clean, None)

    try:
        with open(RUTA_COMENTARIOS, 'w', encoding='utf-8') as f:
            json.dump(comentarios, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error al guardar comentario JSON: {e}")

    try:
        conn = obtener_conexion_db()
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE consultoras_tableau SET notas_lider = ? WHERE codigo_cb = ? OR codigo_cb = ? OR codigo_cb = ?",
            (nota_limpia, codigo_clean, str(codigo_cb).strip(), f"{codigo_clean}.0")
        )
        conn.commit()
        conn.close()
    except Exception as e_sql:
        print(f"Error al actualizar SQLite: {e_sql}")

    return True

def guardar_todos_comentarios(dict_comentarios):
    """
    Guarda masivamente un diccionario de comentarios {codigo_cb: comentario}.
    Actualiza de forma directa y liviana tanto 'comentarios_lideres.json' como la base SQLite 'consultoras_tableau'.
    """
    if not dict_comentarios:
        return True
    comentarios = cargar_comentarios_lideres()
    dict_limpio = {}
    for cb, nota in dict_comentarios.items():
        cb_clean = limpiar_codigo_cb_estandar(cb)
        if not cb_clean:
            continue
        nota_str = str(nota).strip()
        dict_limpio[cb_clean] = nota_str
        if nota_str:
            comentarios[cb_clean] = nota_str
        elif cb_clean in comentarios and nota_str == "":
            comentarios.pop(cb_clean, None)
    
    try:
        with open(RUTA_COMENTARIOS, 'w', encoding='utf-8') as f:
            json.dump(comentarios, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error guardando comentarios masivos en JSON: {e}")

    try:
        conn = obtener_conexion_db()
        cursor = conn.cursor()
        for cb_clean, nota_str in dict_limpio.items():
            cursor.execute(
                "UPDATE consultoras_tableau SET notas_lider = ? WHERE codigo_cb = ? OR codigo_cb = ? OR codigo_cb = ?",
                (nota_str, cb_clean, str(cb_clean).strip(), f"{cb_clean}.0")
            )
        conn.commit()
        conn.close()
    except Exception as e_sql:
        print(f"Nota actualizando SQLite en guardar_todos_comentarios: {e_sql}")

    return True

def procesar_base_tableau_manager(origen='Base de Datos.xlsx'):
    """
    Carga y procesa la sábana de datos de Tableau (Base de Datos.xlsx).
    - Normaliza nombres de columnas y tipos numéricos.
    - Incluye todos los estados de actividad comercial (Activas, Inactivas 1 a 6).
    - Integra comentarios persistentes por Código CB.
    """
    df = None
    if isinstance(origen, str):
        if os.path.exists(origen):
            df = pd.read_excel(origen, sheet_name=0)
    elif isinstance(origen, pd.DataFrame):
        df = origen.copy()
    elif hasattr(origen, 'read'): # UploadedFile de Streamlit
        try:
            origen.seek(0)
        except Exception:
            pass
        df = pd.read_excel(origen, sheet_name=0)

    if df is None or df.empty:
        return None

    # Detectar si la primera fila son celdas vacías o 'Unnamed' y promover la fila de cabecera real
    if any('unnamed' in str(c).lower() for c in df.columns[:5]):
        for r_idx in range(min(5, len(df))):
            row_vals = [str(x) for x in df.iloc[r_idx].values]
            if any('codigo' in x.lower() or 'código' in x.lower() or 'asesora' in x.lower() for x in row_vals):
                df.columns = df.iloc[r_idx]
                df = df.iloc[r_idx + 1:].reset_index(drop=True)
                break

    # Normalización de caracteres en nombres de columnas
    cols_rename = {}
    for c in df.columns:
        clean_c = str(c).replace('\ufffd', 'ó').strip()
        cols_rename[c] = clean_c
    df = df.rename(columns=cols_rename)

    # Reemplazar posibles problemas de codificación comunes y estandarizar nombres de cabecera
    renombres_clave = {
        'Codigo CB': 'Codigo CB',
        'Código CB': 'Codigo CB',
        'Gerencia ': 'Gerencia',
        'Nombre Gerencia': 'Gerencia',
        'Sector ': 'Sector',
        'Nombre Setor': 'Sector',
        'Nombre Sector': 'Sector',
        'Situación': 'Situación',
        'Deuda Mora ': 'Deuda Mora',
        'Pts para Ascender': 'Pts Asc',
        'Pts para Ascender ': 'Pts Asc',
        'Pts Para Ascender': 'Pts Asc',
        'Pts Para Ascender ': 'Pts Asc',
        'Pts Para Mantener': 'Pts Mant',
        'Pts para Mantener': 'Pts Mant',
        'Pts Para Mantener ': 'Pts Mant',
        'Pts Acumulados': 'Pts Acum',
        'Pts Acumulado': 'Pts Acum'
    }
    df = df.rename(columns=renombres_clave)

    # Columnas numéricas a limpiar
    cols_num = [
        'Credito Total', 'Credito Disponible', 'Deuda Total', 'Deuda Mora',
        'Pts Natura', 'Pts AVON', 'Pts Total VD', 'Pts VOL', 'Pts Acum', 'Pts Mant', 'Pts Asc',
        'Ped. Pendientes', 'Ped. Mora', 'Fact. Total', 'Grupo', 'Ciclo', 'Cod. Gerencia', 'Cod. Sector', 'Ciclos Inactividad'
    ]
    for col in cols_num:
        if col in df.columns:
            df[col] = limpiar_serie_numerica(df[col])

    # Convertir flotantes enteros y columnas como Ciclo/Grupo a int64 sin decimales (.000000)
    for col in df.columns:
        if pd.api.types.is_float_dtype(df[col]):
            if 'ciclo' in str(col).lower() or 'grupo' in str(col).lower():
                df[col] = df[col].apply(lambda v: int(round(v * 1000)) if (pd.notna(v) and 10 < v < 1000 and (v % 1 != 0)) else (int(round(v)) if pd.notna(v) else 0))
            else:
                non_nulls = df[col].dropna()
                if not non_nulls.empty and (non_nulls.round(4) % 1 == 0).all():
                    df[col] = df[col].fillna(0).round().astype('int64')

    # Unir comentarios persistentes por Código CB
    comentarios = cargar_comentarios_lideres()
    if 'Codigo CB' in df.columns:
        df['Codigo_CB_key'] = df['Codigo CB'].astype(str).str.strip()
        df['Comentarios_Lider'] = df['Codigo_CB_key'].map(lambda k: comentarios.get(k, ''))
    else:
        df['Comentarios_Lider'] = ''

    # Sincronización consistente de estado Activa entre Situación y Sit. Comercial
    if 'Sit. Comercial' in df.columns or 'Situación' in df.columns:
        if 'Sit. Comercial' not in df.columns:
            df['Sit. Comercial'] = ''
        if 'Situación' not in df.columns:
            df['Situación'] = ''
            
        mask_activa_cycle = pd.Series(False, index=df.index)
        if 'Pts Total VD' in df.columns:
            mask_activa_cycle = mask_activa_cycle | (pd.to_numeric(df['Pts Total VD'], errors='coerce').fillna(0) > 0)
        if 'Fact. Total' in df.columns:
            mask_activa_cycle = mask_activa_cycle | (pd.to_numeric(df['Fact. Total'], errors='coerce').fillna(0) > 0)
        if 'Sit. Comercial' in df.columns:
            mask_activa_cycle = mask_activa_cycle | (df['Sit. Comercial'].astype(str).str.strip().str.lower() == 'activa')
        if 'Indicador' in df.columns:
            mask_activa_cycle = mask_activa_cycle | (df['Indicador'].astype(str).str.strip().str.lower() == 'activas')
        
        df.loc[mask_activa_cycle, 'Sit. Comercial'] = 'Activa'
        df.loc[mask_activa_cycle, 'Situación'] = 'Activa'

        # Para las no activas, sincronizar Situación macro según nivel de inactividad
        mask_no_activa = ~mask_activa_cycle
        sit_lower = df['Sit. Comercial'].astype(str).str.lower()
        mask_disp = mask_no_activa & sit_lower.apply(lambda s: any(k in s for k in ['inactiva 1', 'inactiva 2', 'inactiva 3', 'i1', 'i2', 'i3']))
        mask_indisp = mask_no_activa & sit_lower.apply(lambda s: any(k in s for k in ['inactiva 4', 'inactiva 5', 'inactiva 6', 'i4', 'i5', 'i6']))
        df.loc[mask_disp, 'Situación'] = 'Disponible'
        df.loc[mask_indisp, 'Situación'] = 'Indisponible'

    return df

def color_nivel(val):
    """
    Retorna estilo CSS para formatear la celda de Nivel/Color:
    Bronce, Plata, Oro, Platino, Zafiro, Diamante.
    """
    if pd.isna(val):
        return ""
    v = str(val).strip().lower()
    if 'bronce' in v:
        return 'background-color: #FEF3C7; color: #92400E; font-weight: bold;'
    elif 'plata' in v:
        return 'background-color: #F1F5F9; color: #475569; font-weight: bold;'
    elif 'oro' in v:
        return 'background-color: #FEF9C3; color: #854D0E; font-weight: bold;'
    elif 'platino' in v:
        return 'background-color: #E0F2FE; color: #0369A1; font-weight: bold;'
    elif 'zafiro' in v:
        return 'background-color: #DBEAFE; color: #1E40AF; font-weight: bold;'
    elif 'diamante' in v:
        return 'background-color: #F3E8FF; color: #6B21A8; font-weight: bold;'
    return ""

def color_situacion(val):
    """
    Retorna estilo CSS para formatear la celda de Situación Comercial:
    - Activa: Verde suave (#E6F4EA / texto #137333).
    - Inactiva 1: Amarillo (#FEF9C3 / texto #854D0E).
    - Inactiva 2/3: Naranja (#FFEDD5 / texto #C2410C).
    - Inactiva 4/5: Rojo (#FEE2E2 / texto #991B1B).
    """
    if pd.isna(val):
        return ""
    v = str(val).strip().lower()
    if 'activa' in v and 'inactiva' not in v:
        return 'background-color: #E6F4EA; color: #137333; font-weight: bold;'
    elif 'inactiva 1' in v:
        return 'background-color: #FEF9C3; color: #854D0E; font-weight: bold;'
    elif 'inactiva 2' in v or 'inactiva 3' in v:
        return 'background-color: #FFEDD5; color: #C2410C; font-weight: bold;'
    elif 'inactiva 4' in v or 'inactiva 5' in v or 'inactiva 6' in v:
        return 'background-color: #FEE2E2; color: #991B1B; font-weight: bold;'
    return ""

def color_deuda_mora(val):
    """
    Retorna estilo CSS para formatear celdas de Deuda Mora según semáforo:
    - $0: Sin mora (neutro).
    - $1 a $200.000: Amarillo (#FEF9C3 / texto #854D0E).
    - $200.001 a $500.000: Naranja (#FFEDD5 / texto #C2410C).
    - > $500.000: Rojo (#FEE2E2 / texto #991B1B).
    """
    try:
        if pd.isna(val):
            return ""
        s = str(val)
        for char in ['$', '+', '➡️', '⬆️', '⬇️', ',', ' ', '.']:
            s = s.replace(char, '')
        if not s:
            return ""
        num = float(s)
        if num <= 0:
            return ""
        elif num <= 200000:
            return 'background-color: #FEF9C3; color: #854D0E; font-weight: bold;'
        elif num <= 500000:
            return 'background-color: #FFEDD5; color: #C2410C; font-weight: bold;'
        else:
            return 'background-color: #FEE2E2; color: #991B1B; font-weight: bold;'
    except Exception:
        return ""

def limpiar_codigo_cb_estandar(v):
    """
    Limpia y estandariza cualquier código CB o documento:
    - Remueve NaN/None.
    - Remueve sufijo .0 de conversiones a float.
    - Quita espacios en blanco.
    - Remueve ceros a la izquierda si es un valor numérico para asegurar coincidencia consistente.
    """
    if pd.isna(v):
        return ""
    s = str(v).strip()
    if s.endswith('.0'):
        s = s[:-2].strip()
    if s.isdigit():
        s_num = s.lstrip('0')
        return s_num if s_num else "0"
    return s

def normalizar_estado_mi_grupo(val):
    """
    Normaliza el texto del campo 'ESTADO' independientemente de si viene en MAYÚSCULAS,
    minúsculas o combinación de ambos (ej: 'ACTIVA', 'Activa', 'INACTIVAS 1', 'Inactivas 1', etc.)
    """
    if pd.isna(val):
        return ""
    s = str(val).strip()
    s_lower = s.lower()
    
    # Inactiva 1 a 6 (ej. 'inactivas 1', 'INACTIVAS 1', 'inactiva 1', 'INACTIVA 1')
    for i in range(1, 7):
        if 'inactiv' in s_lower and str(i) in s_lower:
            return f"Inactiva {i}"
            
    if 'activa' in s_lower and 'inactiva' not in s_lower:
        return "Activa"
    elif 'cesada' in s_lower:
        return "Cesada"
    elif 'registrada' in s_lower:
        return "Registrada"
    elif 'intenci' in s_lower:
        return "Intención"
    return s

def leer_excel_tolerante(origen, sheet_idx=0):
    """
    Lector universal y ultra-tolerante de archivos Excel (.xlsx, .xls, .csv, HTML/BIFF corrupto).
    Garantiza la lectura de reportes exportados por sistemas corporativos (Natura, Avon, SAP, CRM)
    incluso si presentan advertencias de corrupción de encabezado BIFF ('Workbook corruption: seen[2] == 4').
    """
    if origen is None:
        return None
    if isinstance(origen, pd.DataFrame):
        return origen.copy()
    
    # 1. Si es un archivo CSV
    name = getattr(origen, 'name', '') or (origen if isinstance(origen, str) else '')
    if str(name).lower().endswith('.csv'):
        for enc in ['utf-8', 'utf-8-sig', 'latin1', 'iso-8859-1', 'cp1252']:
            try:
                if hasattr(origen, 'seek'):
                    origen.seek(0)
                df_csv = pd.read_csv(origen, encoding=enc, sep=None, engine='python')
                if df_csv is not None and not df_csv.empty:
                    return df_csv
            except Exception:
                continue

    # 2. Intentar pandas estándar
    try:
        if hasattr(origen, 'seek'):
            origen.seek(0)
        return pd.read_excel(origen, sheet_name=sheet_idx)
    except Exception:
        pass

    # 3. Intentar xlrd con ignore_workbook_corruption=True (Especial para reportes .xls de Natura/Avon)
    try:
        import xlrd
        wb = None
        if isinstance(origen, str):
            if os.path.exists(origen):
                wb = xlrd.open_workbook(origen, ignore_workbook_corruption=True)
        elif hasattr(origen, 'read'):
            origen.seek(0)
            content = origen.read()
            wb = xlrd.open_workbook(file_contents=content, ignore_workbook_corruption=True)
        elif isinstance(origen, (bytes, bytearray)):
            wb = xlrd.open_workbook(file_contents=origen, ignore_workbook_corruption=True)
            
        if wb is not None:
            sh = wb.sheet_by_index(sheet_idx if isinstance(sheet_idx, int) else 0)
            data = [sh.row_values(r) for r in range(sh.nrows)]
            if data:
                headers = [str(h).replace('\ufffd', 'ó').strip() for h in data[0]]
                df = pd.DataFrame(data[1:], columns=headers)
                return df
    except Exception:
        pass

    # 4. Intentar con engine='xlrd'
    try:
        if hasattr(origen, 'seek'):
            origen.seek(0)
        return pd.read_excel(origen, sheet_name=sheet_idx, engine='xlrd')
    except Exception:
        pass

    # 5. Intentar con engine='openpyxl'
    try:
        if hasattr(origen, 'seek'):
            origen.seek(0)
        return pd.read_excel(origen, sheet_name=sheet_idx, engine='openpyxl')
    except Exception:
        pass

    # 6. Intentar como tabla HTML (muchos sistemas corporativos exportan HTML con extensión .xls)
    try:
        if hasattr(origen, 'seek'):
            origen.seek(0)
        dfs = pd.read_html(origen)
        if dfs:
            return dfs[0]
    except Exception:
        pass

    return None

def actualizar_situacion_comercial_desde_mi_grupo(origen_mi_grupo='mi_grupo.xls', ruta_base='Base de Datos.xlsx'):
    """
    Lee la tabla 'mi_grupo.xls' (o .xlsx), extrae la columna 'ESTADO' / 'Sit. Comercial'
    (reconociendo combinaciones de MAYÚSCULAS y minúsculas), la normaliza y actualiza
    la columna 'Sit. Comercial' en 'Base de Datos.xlsx' vinculando por Código CB.
    """
    if not os.path.exists(ruta_base):
        return {'exito': False, 'error': f"No se encontró el archivo base '{ruta_base}'."}

    df_grupo = leer_excel_tolerante(origen_mi_grupo)

    if df_grupo is None or df_grupo.empty:
        return {'exito': False, 'error': "No se pudo leer el archivo 'mi_grupo'."}

    # 1. Identificar columna de Código CB en mi_grupo (resistente a mayúsculas/minúsculas)
    col_code_grupo = None
    for c in df_grupo.columns:
        c_up = str(c).strip().upper()
        if ('DIGO' in c_up or 'CODIGO' in c_up or 'CB' in c_up) and 'NOMBRE' not in c_up:
            col_code_grupo = c
            break

    if not col_code_grupo:
        return {'exito': False, 'error': "No se encontró la columna de Código en el archivo 'mi_grupo'."}

    # 2. Identificar columna de Estado / Situación en mi_grupo (resistente a mayúsculas/minúsculas/combinadas)
    col_estado_grupo = None
    # Coincidencia exacta con 'ESTADO' primero
    for c in df_grupo.columns:
        if str(c).strip().upper() == 'ESTADO':
            col_estado_grupo = c
            break
    # Coincidencia parcial si no hay exacta
    if not col_estado_grupo:
        for c in df_grupo.columns:
            c_up = str(c).strip().upper()
            if 'ESTADO' in c_up or 'SITUACION' in c_up or 'SITUACI' in c_up:
                col_estado_grupo = c
                break

    if not col_estado_grupo:
        return {'exito': False, 'error': "No se encontró la columna 'ESTADO' o 'Situación' en 'mi_grupo'."}

    # Crear mapeo {codigo_cb_clean: estado_normalizado}
    df_grupo['cb_clean'] = df_grupo[col_code_grupo].apply(limpiar_codigo_cb_estandar)
    df_grupo['estado_clean'] = df_grupo[col_estado_grupo].apply(normalizar_estado_mi_grupo)

    mapa_estados = dict(zip(df_grupo['cb_clean'], df_grupo['estado_clean']))

    # Cargar la base principal
    df_base = pd.read_excel(ruta_base, sheet_name=0)
    # Detectar y promover cabecera real si viene con 'Unnamed' (formato estándar Tableau)
    if any('unnamed' in str(c).lower() for c in df_base.columns[:5]):
        for r_idx in range(min(10, len(df_base))):
            row_vals = [str(x).lower() for x in df_base.iloc[r_idx].values if pd.notna(x)]
            if any('codigo' in x or 'código' in x or 'cb' in x or 'asesora' in x or 'consultora' in x for x in row_vals):
                df_base.columns = [str(col_name).strip() for col_name in df_base.iloc[r_idx]]
                df_base = df_base.iloc[r_idx + 1:].reset_index(drop=True)
                break

    df_base.columns = [str(c).replace('\ufffd', 'ó').strip() for c in df_base.columns]

    col_code_base = None
    for c in df_base.columns:
        c_str = str(c).strip().lower()
        if any(k in c_str for k in ['codigo cb', 'código cb', 'código de consultora', 'codigo de consultora', 'cd consultora']):
            col_code_base = c
            break
    if not col_code_base:
        col_code_base = df_base.columns[0]

    col_sit_comercial = next((c for c in df_base.columns if 'sit. comercial' in str(c).lower() or 'sit comercial' in str(c).lower()), None)
    col_situacion_macro = next((c for c in df_base.columns if str(c).strip().lower() in ['situación', 'situacion']), None)

    if not col_sit_comercial:
        col_sit_comercial = 'Sit. Comercial'
        df_base[col_sit_comercial] = ''
    if not col_situacion_macro:
        col_situacion_macro = 'Situación'
        df_base[col_situacion_macro] = ''

    df_base['cb_clean'] = df_base[col_code_base].apply(limpiar_codigo_cb_estandar)

    for c in [col_sit_comercial, col_situacion_macro]:
        if c and c in df_base.columns:
            df_base[c] = df_base[c].astype(object)

    coincidencias = 0
    cambios = 0
    detalles_cambios = []

    for idx in df_base.index:
        cb = df_base.at[idx, 'cb_clean']
        if cb in mapa_estados:
            coincidencias += 1
            nuevo_estado = mapa_estados[cb]
            estado_actual = str(df_base.at[idx, col_sit_comercial]).strip() if pd.notna(df_base.at[idx, col_sit_comercial]) else ''
            
            # REGLA DE PARCHE INTELIGENTE Y NO DESTRUCTIVO:
            # 1. Ignorar estados no canónicos del portal (Cesada, Registrada, Intención, etc.)
            if nuevo_estado in ['Cesada', 'Registrada', 'Intención'] or not nuevo_estado:
                continue

            # 2. Si en mi_grupo viene como ACTIVA y en la base figuraba inactiva -> PARCHE DE ACTIVACIÓN
            debe_parchar = False
            if nuevo_estado.lower() == 'activa' and estado_actual.lower() != 'activa':
                debe_parchar = True
                estado_final = 'Activa'
            
            # 3. Si ya es ACTIVA en la base de datos maestra (con ventas/puntos o estado previo), NUNCA se degrada
            elif estado_actual.lower() == 'activa':
                continue
            
            if debe_parchar:
                nombre = str(df_base.at[idx, 'Nombre'] if 'Nombre' in df_base.columns else (df_base.at[idx, 'Asesora / Consultora'] if 'Asesora / Consultora' in df_base.columns else cb))
                detalles_cambios.append({
                    'Código CB': cb,
                    'Asesora / Consultora': nombre,
                    'Estado Anterior': estado_actual,
                    'Nuevo Estado (mi_grupo)': estado_final
                })
                df_base.at[idx, col_sit_comercial] = estado_final
                if col_situacion_macro:
                    df_base.at[idx, col_situacion_macro] = 'Activa'
                cambios += 1

    # Sincronizar cambios directamente en SQLite
    if cambios > 0:
        try:
            conn_patch = obtener_conexion_db()
            c_patch = conn_patch.cursor()
            for d in detalles_cambios:
                c_patch.execute(
                    "UPDATE consultoras_tableau SET sit_comercial = 'Activa', situacion = 'Activa' WHERE codigo_cb = ?",
                    (str(d['Código CB']).strip(),)
                )
            conn_patch.commit()
            conn_patch.close()
        except Exception as e_db:
            safe_print(f"Nota actualizando SQLite directo: {e_db}")

    # Guardar en Base de Datos.xlsx
    excel_guardado = True
    msg_alerta_excel = ""
    try:
        df_base.to_excel(ruta_base, index=False)
        try:
            sincronizar_excel_tableau_a_sqlite(ruta_base)
        except Exception:
            pass
    except PermissionError:
        excel_guardado = False
        msg_alerta_excel = f" (Nota: '{ruta_base}' está abierto en Excel; los cambios se guardaron en la plataforma, pero para actualizar el archivo físico ciérralo en Excel)."
    except Exception as e:
        safe_print(f"Error al escribir Excel: {e}")

    return {
        'exito': True,
        'coincidencias': coincidencias,
        'cambios': cambios,
        'detalles': detalles_cambios,
        'aviso_excel': msg_alerta_excel
    }

def filtrar_consultoras_portal_especial(df_base, opcion_portal, ruta_mi_grupo='mi_grupo.xls'):
    """
    Filtra el DataFrame de consultoras para visualizar las consultoras Cesadas, Registradas
    o con Intención reportadas en el portal (mi_grupo.xls), cruzándolas por Código CB
    para que queden asociadas a su grupo de líder correspondiente.
    """
    if df_base is None or df_base.empty:
        return df_base

    if not os.path.exists(ruta_mi_grupo):
        return df_base

    df_mg = leer_excel_tolerante(ruta_mi_grupo)
    if df_mg is None or df_mg.empty:
        return df_base

    # Identificar columna de estado y código en mi_grupo
    col_estado = next((c for c in df_mg.columns if 'ESTADO' in str(c).upper() or 'SITUAC' in str(c).upper()), None)
    col_code = next((c for c in df_mg.columns if any(k in str(c).upper() for k in ['CODIGO', 'CÓDIGO', 'CB'])), None)

    if not col_estado or not col_code:
        return df_base

    estado_target = None
    op_lower = str(opcion_portal).lower()
    if 'cesada' in op_lower:
        estado_target = 'cesada'
    elif 'registrada' in op_lower:
        estado_target = 'registrada'
    elif 'intenci' in op_lower:
        estado_target = 'intención'

    if not estado_target:
        return df_base

    df_mg['cb_clean'] = df_mg[col_code].apply(limpiar_codigo_cb_estandar)
    df_mg_target = df_mg[df_mg[col_estado].astype(str).str.lower().str.contains(estado_target, na=False)].copy()

    # Identificar columna código en df_base
    col_base_cb = next((c for c in df_base.columns if any(k in str(c).lower() for k in ['codigo cb', 'código cb', 'codigo_cb', 'cd consultora'])), df_base.columns[0])
    df_base_work = df_base.copy()
    df_base_work['cb_clean'] = df_base_work[col_base_cb].apply(limpiar_codigo_cb_estandar)

    cbs_target = set(df_mg_target['cb_clean'])
    df_filtrado = df_base_work[df_base_work['cb_clean'].isin(cbs_target)].copy()

    # Asignar la situación comercial de portal para mostrar claramente el estado
    col_sit = next((c for c in df_filtrado.columns if any(k in str(c).lower() for k in ['sit. comercial', 'situacion', 'situación'])), None)
    if col_sit:
        mapa_est = dict(zip(df_mg_target['cb_clean'], df_mg_target[col_estado]))
        df_filtrado[col_sit] = df_filtrado['cb_clean'].map(mapa_est).fillna(f"{estado_target.title()} (Portal)")

    if 'cb_clean' in df_filtrado.columns:
        df_filtrado = df_filtrado.drop(columns=['cb_clean'])

    return df_filtrado

def actualizar_base_desde_activas(origen_activas, ruta_base='Base de Datos.xlsx'):
    """
    Cruce del archivo de 'activas' (.xlsx, .xls, .csv) con la base de datos de Tableau:
    - Identifica los Códigos CB del archivo de activas.
    - Actualiza únicamente el campo 'Indicador' con el valor 'Activas'.
    - Sincroniza 'Sit. Comercial' y 'Situación' a 'Activa'.
    - No modifica otros campos (Facturación, Puntos, Pedidos se preservan intactos).
    - Guarda en 'Base de Datos.xlsx' y sincroniza la tabla SQLite consultoras_tableau.
    """
    if not os.path.exists(ruta_base):
        return {'exito': False, 'error': f"No se encontró el archivo base '{ruta_base}'."}

    df_act = leer_excel_tolerante(origen_activas)

    if df_act is None or df_act.empty:
        return {'exito': False, 'error': "El archivo de activas está vacío o no se pudo interpretar."}

    # Si la primera fila es encabezado desplazado, buscar la fila con 'CODIGO' o 'DOCUMENTO' o 'NOMBRE'
    if not any(any(k in str(c).upper() for k in ['CODIGO', 'CB', 'DOC', 'ASESORA', 'CONSULTORA', 'NOMBRE', 'CEDULA', 'IDENT']) for c in df_act.columns):
        for r_idx in range(min(15, len(df_act))):
            row_vals = [str(x).upper() for x in df_act.iloc[r_idx].values if pd.notna(x)]
            if any('COD' in v or 'DOC' in v or 'CONSULTORA' in v or 'ASESORA' in v or 'CEDULA' in v for v in row_vals):
                df_act.columns = df_act.iloc[r_idx]
                df_act = df_act.iloc[r_idx + 1:].reset_index(drop=True)
                break

    # 2. Identificar columna de Código CB o Documento en el archivo de activas
    col_code_act = None
    for c in df_act.columns:
        c_up = str(c).strip().upper()
        if ('DIGO' in c_up or 'CODIGO' in c_up or 'CB' in c_up) and 'NOMBRE' not in c_up and 'GERENCIA' not in c_up and 'SECTOR' not in c_up and 'GRUPO' not in c_up:
            col_code_act = c
            break
    if not col_code_act:
        for c in df_act.columns:
            c_up = str(c).strip().upper()
            if any(k in c_up for k in ['DOC', 'CEDULA', 'IDENT', 'CONSULTORA', 'ASESORA', 'ID_CONSULTORA', 'ID_ASESORA']):
                if 'NOMBRE' not in c_up and 'GERENCIA' not in c_up and 'SECTOR' not in c_up and 'GRUPO' not in c_up:
                    col_code_act = c
                    break

    if not col_code_act:
        return {'exito': False, 'error': "No se identificó la columna de Código o Documento en el archivo de activas."}

    # Obtener el conjunto de Códigos CB del archivo de activas
    df_act['cb_clean'] = df_act[col_code_act].apply(limpiar_codigo_cb_estandar)
    cbs_activas = set(df_act['cb_clean'].dropna().loc[lambda s: s != ''])

    # 3. Cargar Base de Datos.xlsx principal
    df_base = pd.read_excel(ruta_base, sheet_name=0)
    # Detectar y promover cabecera real si viene con 'Unnamed' (formato estándar Tableau)
    if any('unnamed' in str(c).lower() for c in df_base.columns[:5]):
        for r_idx in range(min(10, len(df_base))):
            row_vals = [str(x).lower() for x in df_base.iloc[r_idx].values if pd.notna(x)]
            if any('codigo' in x or 'código' in x or 'cb' in x or 'asesora' in x or 'consultora' in x for x in row_vals):
                df_base.columns = [str(col_name).strip() for col_name in df_base.iloc[r_idx]]
                df_base = df_base.iloc[r_idx + 1:].reset_index(drop=True)
                break

    df_base.columns = [str(c).replace('\ufffd', 'ó').strip() for c in df_base.columns]

    col_code_base = None
    for c in df_base.columns:
        c_str = str(c).strip().lower()
        if any(k in c_str for k in ['codigo cb', 'código cb', 'código de consultora', 'codigo de consultora', 'cd consultora']):
            col_code_base = c
            break
    if not col_code_base:
        col_code_base = df_base.columns[0]

    # Identificar o crear columna 'Indicador'
    col_indicador = next((c for c in df_base.columns if str(c).strip().lower() in ['indicador', 'indicadores']), None)
    if not col_indicador:
        col_indicador = 'Indicador'
        df_base[col_indicador] = ''

    col_sit_comercial = next((c for c in df_base.columns if 'sit. comercial' in str(c).lower() or 'sit comercial' in str(c).lower()), None)
    col_situacion_macro = next((c for c in df_base.columns if str(c).strip().lower() in ['situación', 'situacion']), None)

    if not col_sit_comercial:
        col_sit_comercial = 'Sit. Comercial'
        df_base[col_sit_comercial] = ''
    if not col_situacion_macro:
        col_situacion_macro = 'Situación'
        df_base[col_situacion_macro] = ''

    df_base['cb_clean'] = df_base[col_code_base].apply(limpiar_codigo_cb_estandar)

    # Asegurar compatibilidad de tipos (PyArrow / Pandas 2+ / Python 3.14)
    for c in [col_indicador, col_sit_comercial, col_situacion_macro]:
        if c and c in df_base.columns:
            df_base[c] = df_base[c].astype(object)

    coincidencias = 0
    cambios_totales = 0
    detalles_cambios = []
    cbs_base_set = set(df_base['cb_clean'].dropna())

    for idx in df_base.index:
        cb = df_base.at[idx, 'cb_clean']
        if cb in cbs_activas:
            coincidencias += 1
            ind_actual = str(df_base.at[idx, col_indicador]).strip() if pd.notna(df_base.at[idx, col_indicador]) else ''
            
            # Actualizar Indicador a 'Activas'
            df_base.at[idx, col_indicador] = 'Activas'
            df_base.at[idx, col_sit_comercial] = 'Activa'
            if col_situacion_macro:
                df_base.at[idx, col_situacion_macro] = 'Activa'

            if ind_actual != 'Activas':
                cambios_totales += 1
                nombre = str(df_base.at[idx, 'Nombre'] if 'Nombre' in df_base.columns else (df_base.at[idx, 'Asesora / Consultora'] if 'Asesora / Consultora' in df_base.columns else cb))
                detalles_cambios.append({
                    'Código CB': cb,
                    'Asesora / Consultora': nombre,
                    'Indicador Anterior': ind_actual or 'N/D',
                    'Nuevo Indicador': 'Activas',
                    'Sit. Comercial': 'Activa'
                })

    df_base = df_base.drop(columns=['cb_clean'], errors='ignore')

    # Guardar en Base de Datos.xlsx y refrescar SQLite
    try:
        df_base.to_excel(ruta_base, index=False)
        try:
            sincronizar_excel_tableau_a_sqlite(ruta_base)
        except Exception as e_sql:
            safe_print(f"Advertencia al sincronizar SQLite tras cruce de activas: {e_sql}")

        no_encontradas = [cb for cb in cbs_activas if cb not in cbs_base_set]

        return {
            'exito': True,
            'total_activas_archivo': len(cbs_activas),
            'coincidencias': coincidencias,
            'cambios_totales': cambios_totales,
            'detalles': detalles_cambios,
            'no_encontradas_count': len(no_encontradas)
        }
    except Exception as e_save:
        return {'exito': False, 'error': f"Error al guardar '{ruta_base}': {e_save}"}

# --- MÓDULO DE AUTENTICACIÓN Y GESTIÓN DE USUARIOS POR ROL ---
RUTA_USUARIOS = ruta_persistente('usuarios.json')

def hashlib_sha256(texto):
    import hashlib
    return hashlib.sha256(str(texto).encode('utf-8')).hexdigest()

def reconciliar_usuarios_sectores(usuarios_dict, persistir=True):
    """
    Garantiza la consistencia absoluta e integridad referencial entre:
    1. El rol del usuario (gerente, lider, superadmin).
    2. Su código de sector (700000459 vs 700000466 u otros).
    3. Su nombre de sector oficial ('MATICES CLERY' vs 'EMOCIONES DOLLY').
    4. Su código de grupo si es Líder de Negocio.
    Corrige automáticamente cualquier asignación errónea heredada de cargas previas.
    """
    if not isinstance(usuarios_dict, dict):
        return usuarios_dict

    GRUPOS_MATICES = {
        '9640', '9334', '7841', '10168', '9678', '9291', '8425', '9948',
        '9717', '8481', '10255', '10223', '10260', '9400', '9175', '8048',
        '10261', '9581', '9718', '9376'
    }
    GRUPOS_EMOCIONES = {
        '177', '9337', '9924', '7817', '9529', '5060', '10293', '10215',
        '10066', '9891', '9349', '10294', '7815', '8055', '8928'
    }

    # Intentar obtener mapa relacional de la base de datos
    mapa_db = {}
    try:
        conn = obtener_conexion_db()
        cursor = conn.cursor()
        cursor.execute("SELECT grupo, cod_sector, sector FROM consultoras_tableau GROUP BY grupo")
        for r in cursor.fetchall():
            if r[0]:
                mapa_db[str(r[0]).strip()] = (str(r[1]).strip() if r[1] else '', str(r[2]).strip() if r[2] else '')
    except Exception:
        pass

    cambios = False

    for uname, udata in usuarios_dict.items():
        if not isinstance(udata, dict):
            continue
        rol = str(udata.get("rol", "")).strip().lower()
        grp = str(udata.get("codigo_grupo") or "").strip()
        sec = str(udata.get("codigo_sector") or "").strip()
        nom_sec = str(udata.get("nombre_sector") or "").strip()

        if rol == "superadmin":
            if udata.get("codigo_sector") is not None:
                udata["codigo_sector"] = None
                cambios = True
            if udata.get("nombre_sector") != "Gestión Corporativa Global":
                udata["nombre_sector"] = "Gestión Corporativa Global"
                cambios = True
            continue

        if rol == "gerente":
            if sec == "700000459" or "clery" in uname or uname == "gerente":
                if udata.get("codigo_sector") != "700000459":
                    udata["codigo_sector"] = "700000459"
                    cambios = True
                if udata.get("nombre_sector") != "MATICES CLERY":
                    udata["nombre_sector"] = "MATICES CLERY"
                    cambios = True
            elif sec == "700000466" or "dolly" in uname:
                if udata.get("codigo_sector") != "700000466":
                    udata["codigo_sector"] = "700000466"
                    cambios = True
                if udata.get("nombre_sector") != "EMOCIONES DOLLY":
                    udata["nombre_sector"] = "EMOCIONES DOLLY"
                    cambios = True
            elif sec:
                nom_auto = None
                historico_lookup = cargar_historico_sectores()
                if sec in historico_lookup and historico_lookup[sec].get("nombre_sector"):
                    nom_auto = historico_lookup[sec].get("nombre_sector")
                else:
                    cat_lookup = cargar_catalogo_sectores()
                    if sec in cat_lookup and cat_lookup[sec].get("nombre_sector"):
                        nom_auto = cat_lookup[sec].get("nombre_sector")
                if nom_auto and udata.get("nombre_sector") != nom_auto:
                    udata["nombre_sector"] = nom_auto
                    cambios = True
            continue

        if rol == "lider" and grp:
            nuevo_sec = None
            nuevo_nom = None

            if grp in GRUPOS_MATICES:
                nuevo_sec = "700000459"
                nuevo_nom = "MATICES CLERY"
            elif grp in GRUPOS_EMOCIONES:
                nuevo_sec = "700000466"
                nuevo_nom = "EMOCIONES DOLLY"
            elif grp in mapa_db:
                db_sec, db_nom = mapa_db[grp]
                if db_sec == "700000459" or "matices" in db_nom.lower():
                    nuevo_sec = "700000459"
                    nuevo_nom = "MATICES CLERY"
                elif db_sec == "700000466" or "emociones" in db_nom.lower():
                    nuevo_sec = "700000466"
                    nuevo_nom = "EMOCIONES DOLLY"
                elif db_sec:
                    nuevo_sec = db_sec
                    nuevo_nom = db_nom or f"Sector {db_sec}"

            if not nuevo_sec:
                cat_lookup = cargar_catalogo_sectores()
                for cat_sec_id, cat_sec_data in cat_lookup.items():
                    for lid_info in cat_sec_data.get('lideres', []):
                        if str(lid_info.get('codigo_grupo', '')).strip() == grp:
                            nuevo_sec = str(cat_sec_id).strip()
                            nuevo_nom = cat_sec_data.get('nombre_sector', f'Sector {cat_sec_id}')
                            break
                    if nuevo_sec:
                        break

            if nuevo_sec:
                if udata.get("codigo_sector") != nuevo_sec:
                    udata["codigo_sector"] = nuevo_sec
                    cambios = True
                if udata.get("nombre_sector") != nuevo_nom:
                    udata["nombre_sector"] = nuevo_nom
                    cambios = True
            else:
                # Coherencia canónica cuando no se conoce el grupo
                if sec == "700000459" and nom_sec != "MATICES CLERY":
                    udata["nombre_sector"] = "MATICES CLERY"
                    cambios = True
                elif sec == "700000466" and nom_sec != "EMOCIONES DOLLY":
                    udata["nombre_sector"] = "EMOCIONES DOLLY"
                    cambios = True
                elif sec:
                    historico_lookup = cargar_historico_sectores()
                    if sec in historico_lookup and historico_lookup[sec].get("nombre_sector"):
                        nom_h = historico_lookup[sec].get("nombre_sector")
                        if nom_sec != nom_h:
                            udata["nombre_sector"] = nom_h
                            cambios = True

    if cambios and persistir:
        # Guardado silencioso de corrección
        rutas_guardar = set(filter(None, [RUTA_USUARIOS, 'usuarios.json']))
        if DIR_PERSISTENTE and os.path.isdir(DIR_PERSISTENTE):
            rutas_guardar.add(os.path.join(DIR_PERSISTENTE, 'usuarios.json'))
        for r in rutas_guardar:
            try:
                p_dir = os.path.dirname(r)
                if p_dir:
                    os.makedirs(p_dir, exist_ok=True)
                with open(r, 'w', encoding='utf-8') as f:
                    json.dump(usuarios_dict, f, ensure_ascii=False, indent=2)
            except Exception:
                pass
        try:
            sincronizar_usuarios_a_sqlite()
        except Exception:
            pass

    return usuarios_dict

def cargar_usuarios():
    """
    Carga el diccionario de usuarios desde el almacenamiento persistente o local.
    Si no existe, inicializa con los usuarios predeterminados.
    Aplica reconciliación automática de consistencia de sectores.
    """
    rutas_a_probar = [RUTA_USUARIOS, 'usuarios.json', os.path.join('data', 'usuarios.json')]
    for r in rutas_a_probar:
        if r and os.path.exists(r):
            try:
                with open(r, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if data and isinstance(data, dict):
                        return reconciliar_usuarios_sectores(data)
            except Exception:
                pass

    # Usuarios predeterminados si el archivo no existe
    def_pass_super = hashlib_sha256("superadmin123")
    def_pass_admin = hashlib_sha256("admin123")
    def_pass_lider = hashlib_sha256("lider123")
    def_pass_asesor = hashlib_sha256("asesor123")

    usuarios_default = {
        "admin": {
            "nombre": "Super Administrador del Sistema",
            "password_hash": def_pass_super,
            "rol": "superadmin",
            "codigo_grupo": None,
            "debe_cambiar_password": False,
            "nombre_sector": "Gestión Corporativa Global"
        },
        "gerente": {
            "nombre": "Clery Cuellar",
            "password_hash": def_pass_admin,
            "rol": "gerente",
            "codigo_grupo": None,
            "codigo_sector": "700000459",
            "nombre_sector": "MATICES CLERY",
            "estado_suscripcion": "activo",
            "fecha_vencimiento": None,
            "debe_cambiar_password": False
        },
        "lider8425": {
            "nombre": "Luz Dary Chacon Gaitan",
            "password_hash": def_pass_lider,
            "rol": "lider",
            "codigo_grupo": "8425",
            "codigo_sector": "700000459",
            "nombre_sector": "MATICES CLERY",
            "estado_suscripcion": "activo",
            "fecha_vencimiento": None,
            "debe_cambiar_password": False
        },
        "lider7841": {
            "nombre": "Carmenza Roncancio Gachancipa",
            "password_hash": def_pass_lider,
            "rol": "lider",
            "codigo_grupo": "7841",
            "codigo_sector": "700000459",
            "nombre_sector": "MATICES CLERY",
            "estado_suscripcion": "activo",
            "fecha_vencimiento": None,
            "debe_cambiar_password": False
        },
        "asesor": {
            "nombre": "Usuario Consulta Facturación",
            "password_hash": def_pass_asesor,
            "rol": "asesor",
            "codigo_grupo": None,
            "debe_cambiar_password": False
        }
    }
    guardar_usuarios(usuarios_default)
    return usuarios_default

def guardar_usuarios(dict_usuarios):
    """
    Guarda el diccionario de usuarios en almacenamiento persistente y sincroniza con el archivo local.
    """
    exito = False
    rutas_guardar = set(filter(None, [RUTA_USUARIOS, 'usuarios.json']))
    if DIR_PERSISTENTE and os.path.isdir(DIR_PERSISTENTE):
        rutas_guardar.add(os.path.join(DIR_PERSISTENTE, 'usuarios.json'))

    for r in rutas_guardar:
        try:
            p_dir = os.path.dirname(r)
            if p_dir:
                os.makedirs(p_dir, exist_ok=True)
            with open(r, 'w', encoding='utf-8') as f:
                json.dump(dict_usuarios, f, ensure_ascii=False, indent=2)
            exito = True
        except Exception as e:
            safe_print(f"Nota al guardar usuarios en {r}: {e}")
    return exito

def refrescar_perfil_usuario_en_sesion(user_dict):
    """
    Retorna el perfil más fresco del usuario desde el almacenamiento persistente.
    Útil para actualizar permisos, rol, suscripción y contraseñas en vivo sin desloguear.
    """
    if not user_dict or not isinstance(user_dict, dict):
        return user_dict
    u_name = str(user_dict.get("username") or "").strip().lower()
    if not u_name:
        return user_dict
    usuarios = cargar_usuarios()
    if u_name in usuarios:
        fresco = usuarios[u_name].copy()
        fresco["username"] = u_name
        return fresco
    return user_dict

# --- MÓDULO DE SUSCRIPCIONES, PRUEBAS GRATIS (15 DÍAS) Y CONTROL ANTI-FRAUDE ---
RUTA_HISTORICO_SECTORES = ruta_persistente('sectores_historico.json')
RUTA_MARCA_AGUA_TIEMPO = ruta_persistente('marca_agua_sistema.json')
RUTA_AUDITORIA_JSON = ruta_persistente('auditoria_logs.json')

def cargar_historico_sectores():
    """
    Carga el historial de sectores registrados para el control de pruebas y suscripciones.
    """
    rutas_a_probar = [RUTA_HISTORICO_SECTORES, 'sectores_historico.json', os.path.join('data', 'sectores_historico.json')]
    for r in rutas_a_probar:
        if r and os.path.exists(r):
            try:
                with open(r, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if data and isinstance(data, dict):
                        return data
            except Exception:
                pass
    
    sectores_init = {
        "700000459": {
            "codigo_sector": "700000459",
            "nombre_sector": "SECTOR MATICES CLERY",
            "primera_prueba_fecha": "2026-01-01T00:00:00",
            "ha_consumido_prueba": True,
            "ha_pagado": True,
            "estado": "activo",
            "fecha_vencimiento": None,
            "correo_gerente": "gerente",
            "telefono_gerente": "3057939537",
            "nombre_gerente": "Clery Cuellar"
        },
        "700000466": {
            "codigo_sector": "700000466",
            "nombre_sector": "EMOCIONES DOLLY",
            "primera_prueba_fecha": "2026-01-01T00:00:00",
            "ha_consumido_prueba": True,
            "ha_pagado": True,
            "estado": "activo",
            "fecha_vencimiento": None,
            "correo_gerente": "dolly.parra@natura.net",
            "telefono_gerente": "3113201145",
            "nombre_gerente": "Dolly Parra"
        }
    }
    guardar_historico_sectores(sectores_init)
    return sectores_init

def guardar_historico_sectores(dict_sectores):
    """
    Persiste el histórico de sectores en almacenamiento persistente y archivo local.
    """
    exito = False
    rutas_guardar = set(filter(None, [RUTA_HISTORICO_SECTORES, 'sectores_historico.json']))
    if DIR_PERSISTENTE and os.path.isdir(DIR_PERSISTENTE):
        rutas_guardar.add(os.path.join(DIR_PERSISTENTE, 'sectores_historico.json'))

    for r in rutas_guardar:
        try:
            p_dir = os.path.dirname(r)
            if p_dir:
                os.makedirs(p_dir, exist_ok=True)
            with open(r, 'w', encoding='utf-8') as f:
                json.dump(dict_sectores, f, ensure_ascii=False, indent=2)
            exito = True
        except Exception as e:
            safe_print(f"Nota al guardar histórico de sectores en {r}: {e}")
    return exito

def obtener_nombre_sector_usuario(user_info):
    """
    Retorna el nombre del sector del usuario de forma dinámica y autoritativa.
    Prioridad:
    1. Si es superadmin -> 'Gestión Corporativa Global'
    2. Lookup canónico estricto por 'codigo_sector' (700000459 -> MATICES CLERY, 700000466 -> EMOCIONES DOLLY).
    3. Lookup en sectores_historico.json o catalogo_sectores.json mediante 'codigo_sector'.
    4. Campo 'nombre_sector' en el perfil del usuario (validando que no entre en conflicto con el código).
    5. Fallback -> 'Sector {codigo_sector}' o 'Liderazgo Empresarial'.
    """
    if not user_info or not isinstance(user_info, dict):
        return "Liderazgo Empresarial"
        
    rol = str(user_info.get("rol", "")).strip().lower()
    if rol == "superadmin":
        return "Gestión Corporativa Global"
        
    cod_sec = str(user_info.get("codigo_sector") or "").strip()
    
    # Verificación canónica estricta para evitar cruces
    if cod_sec == "700000459":
        return "MATICES CLERY"
    elif cod_sec == "700000466":
        return "EMOCIONES DOLLY"
        
    if cod_sec and cod_sec.lower() not in ["none", "nan", "null", ""]:
        historico = cargar_historico_sectores()
        if cod_sec in historico:
            nom_hist = historico[cod_sec].get("nombre_sector")
            if nom_hist and str(nom_hist).strip() and str(nom_hist).lower() not in ["none", "nan", "null", ""]:
                return str(nom_hist).strip()
        catalogo = cargar_catalogo_sectores()
        if cod_sec in catalogo:
            nom_cat = catalogo[cod_sec].get("nombre_sector")
            if nom_cat and str(nom_cat).strip():
                return str(nom_cat).strip()

    nom_sec = user_info.get("nombre_sector")
    if nom_sec and str(nom_sec).strip() and str(nom_sec).lower() not in ["none", "nan", "null", ""]:
        if cod_sec == "700000466" and "matices" in str(nom_sec).lower():
            return "EMOCIONES DOLLY"
        if cod_sec == "700000459" and "emociones" in str(nom_sec).lower():
            return "MATICES CLERY"
        return str(nom_sec).strip()
        
    if cod_sec and cod_sec.lower() not in ["none", "nan", "null", ""]:
        return f"Sector {cod_sec}"
        
    return "Liderazgo Empresarial"

def verificar_estado_suscripcion(user_info_o_sector):
    """
    Determina si un usuario o sector tiene acceso permitido al sistema de forma dinámica y tolerante a fallos.
    Retorna un diccionario con:
    - permitido: bool
    - estado: 'superadmin' | 'activo' | 'prueba' | 'vencido' | 'bloqueado'
    - dias_restantes: int
    - fecha_vencimiento_str: str
    - motivo: str
    """
    from datetime import datetime

    if isinstance(user_info_o_sector, dict):
        rol = user_info_o_sector.get("rol", "")
        if rol == "superadmin":
            return {
                "permitido": True,
                "estado": "superadmin",
                "dias_restantes": 9999,
                "fecha_vencimiento_str": "Permanente (Super Administrador)",
                "motivo": "Acceso Total Super Administrador"
            }
        
        # Obtener datos frescos del usuario
        u_name = str(user_info_o_sector.get("username") or "").strip().lower()
        if u_name:
            usuarios_frescos = cargar_usuarios()
            if u_name in usuarios_frescos:
                user_info_o_sector = usuarios_frescos[u_name]

        cod_sector = str(user_info_o_sector.get("codigo_sector") or "").strip()
        user_estado = user_info_o_sector.get("estado_suscripcion")
        user_vence = user_info_o_sector.get("fecha_vencimiento")
    else:
        cod_sector = str(user_info_o_sector or "").strip()
        user_estado = None
        user_vence = None

    if not cod_sector:
        return {
            "permitido": True,
            "estado": "activo",
            "dias_restantes": 9999,
            "fecha_vencimiento_str": "Activo",
            "motivo": "Cuenta activa"
        }

    historico = cargar_historico_sectores()
    sec_info = historico.get(cod_sector, {})
    
    estado = sec_info.get("estado") or user_estado or "activo"
    vence_iso = sec_info.get("fecha_vencimiento") if "fecha_vencimiento" in sec_info else user_vence
    ha_pagado = sec_info.get("ha_pagado", False)
    
    # 1. Si está bloqueado explícitamente por el Administrador
    if estado == "bloqueado":
        return {
            "permitido": False,
            "estado": "bloqueado",
            "dias_restantes": 0,
            "fecha_vencimiento_str": "Bloqueado",
            "motivo": "El acceso para este sector ha sido suspendido por el Administrador."
        }

    # 2. Si no tiene fecha de vencimiento o es un sector pagado / activo permanente
    if (not vence_iso) or (estado == "activo" and ha_pagado and not vence_iso):
        return {
            "permitido": True,
            "estado": "activo",
            "dias_restantes": 9999,
            "fecha_vencimiento_str": "Suscripción Activa (Permanente)",
            "motivo": "Suscripción activa y vigente"
        }

    # 3. Evaluar fecha de vencimiento
    try:
        dt_vence = datetime.fromisoformat(str(vence_iso))
        dt_now = datetime.now()
        
        diff = (dt_vence - dt_now).total_seconds()
        dias_restantes = max(0, int(diff // 86400) + 1)
        fecha_str = dt_vence.strftime("%d/%m/%Y a las %H:%M")

        if diff >= 0:
            return {
                "permitido": True,
                "estado": estado,
                "dias_restantes": dias_restantes,
                "fecha_vencimiento_str": fecha_str,
                "motivo": f"Vigente hasta el {fecha_str}"
            }
        else:
            return {
                "permitido": False,
                "estado": "vencido",
                "dias_restantes": 0,
                "fecha_vencimiento_str": fecha_str,
                "motivo": f"Tu periodo de {'prueba de 15 días' if estado == 'prueba' else 'suscripción'} ha finalizado el {fecha_str}."
            }
    except Exception as e:
        return {
            "permitido": True,
            "estado": "activo",
            "dias_restantes": 9999,
            "fecha_vencimiento_str": "Activo",
            "motivo": f"Vigente ({e})"
        }

# --- CATÁLOGO CORPORATIVO DE SECTORES Y AUTO-APROVISIONAMIENTO DE LÍDERES ---

RUTA_CATALOGO_SECTORES_JSON = 'catalogo_sectores.json'

def limpiar_nombre_sector_solo(nombre_sector):
    """
    Retorna únicamente el nombre del sector sin el nombre de la persona ni códigos adicionales.
    Ejemplo: 'SECTOR COLORES  KAREN' -> 'SECTOR COLORES'
             'SECTOR ARTESANÍA FERNANDA' -> 'SECTOR ARTESANÍA'
             'SECTOR MATICES CLERY' -> 'SECTOR MATICES'
             'Sector ABÁNICO Judy' -> 'SECTOR ABÁNICO'
    """
    if not nombre_sector:
        return ""
    partes = str(nombre_sector).strip().split()
    if len(partes) >= 2:
        if partes[0].lower() == 'sector':
            return f"SECTOR {partes[1].upper()}"
    return str(nombre_sector).strip().upper()

def cargar_catalogo_sectores():
    """
    Retorna el diccionario estructurado de sectores y líderes desde catalogo_sectores.json.
    Si no existe, intenta extraerlo desde Objetivos Arte.xlsx.
    """
    if os.path.exists(RUTA_CATALOGO_SECTORES_JSON):
        try:
            with open(RUTA_CATALOGO_SECTORES_JSON, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if data and isinstance(data, dict):
                    return data
        except Exception as e:
            safe_print(f"Nota al cargar {RUTA_CATALOGO_SECTORES_JSON}: {e}")

    if os.path.exists('Objetivos Arte.xlsx'):
        return extraer_catalogo_sectores_desde_arte('Objetivos Arte.xlsx')

    return {}

def extraer_catalogo_sectores_desde_arte(ruta_o_buffer='Objetivos Arte.xlsx'):
    """
    Analiza la hoja 'Desafíos LNN' de Objetivos Arte y extrae todos los sectores con sus códigos oficiales
    y la lista completa de sus líderes de negocio (con código de grupo y nombre oficial).
    """
    catalogo = {}
    try:
        xl = pd.ExcelFile(ruta_o_buffer)
        hoja_des = next((s for s in xl.sheet_names if 'desaf' in s.lower() and 'ln' in s.lower()), None)
        if not hoja_des:
            hoja_des = next((s for s in xl.sheet_names if 'desaf' in s.lower()), xl.sheet_names[0])
            
        df_raw = xl.parse(hoja_des)

        if any('unnamed' in str(c).lower() for c in df_raw.columns[:5]):
            for r_idx in range(min(5, len(df_raw))):
                row_vals = [str(x).lower() for x in df_raw.iloc[r_idx].values if pd.notna(x)]
                if any('sector' in v or 'grupo' in v or 'lider' in v or 'líder' in v for v in row_vals):
                    df_raw.columns = [str(col_name).strip() for col_name in df_raw.iloc[r_idx]]
                    df_raw = df_raw.iloc[r_idx + 1:].reset_index(drop=True)
                    break

        col_sec = next((c for c in df_raw.columns if str(c).lower() in ['sector', 'nombre setor', 'nombre sector']), None)
        col_cod = next((c for c in df_raw.columns if 'cod' in str(c).lower() and 'sector' in str(c).lower()), None)
        col_grp = next((c for c in df_raw.columns if any(k in str(c).lower() for k in ['cód. grupo', 'cod grupo', 'grupo'])), None)
        col_lc = next((c for c in df_raw.columns if 'cód. líder' in str(c).lower() or 'cod lider' in str(c).lower()), None)
        col_ln = next((c for c in df_raw.columns if 'nombre líder' in str(c).lower() or 'nombre lider' in str(c).lower()), None)

        if not col_sec:
            return catalogo

        for sec_name, grp in df_raw.groupby(col_sec):
            sec_clean = str(sec_name).strip()
            if not sec_clean or sec_clean.lower() in ['nan', '-', 'none', '']:
                continue
                
            cod_s = ""
            if col_cod and not grp[col_cod].dropna().empty:
                cod_s = str(grp[col_cod].dropna().iloc[0]).split('.')[0].strip()
                
            lideres = []
            for _, r in grp.iterrows():
                g = str(r.get(col_grp, '')).split('.')[0].strip() if col_grp else ''
                lc = str(r.get(col_lc, '')).split('.')[0].strip() if col_lc else ''
                ln = str(r.get(col_ln, '')).strip() if col_ln else ''
                
                if not ln or ln.lower() in ['nan', '-', 'none'] or 'fuera de grupo' in ln.lower():
                    continue
                    
                clean_name = ln.split(' - ', 1)[1].strip() if ' - ' in ln else ln
                lideres.append({
                    'codigo_grupo': g,
                    'codigo_consultora': lc if lc != '-' else '',
                    'nombre_lider': clean_name,
                    'nombre_original': ln
                })

            nombre_solo = limpiar_nombre_sector_solo(sec_clean)
            clave_sector = cod_s if cod_s else nombre_solo
            catalogo[clave_sector] = {
                'codigo_sector': cod_s,
                'nombre_sector': nombre_solo,
                'nombre_sector_original': sec_clean,
                'total_lideres': len(lideres),
                'lideres': lideres
            }

        with open(RUTA_CATALOGO_SECTORES_JSON, 'w', encoding='utf-8') as f:
            json.dump(catalogo, f, ensure_ascii=False, indent=2)

    except Exception as e:
        safe_print(f"Error al extraer catálogo de sectores desde Objetivos Arte: {e}")

    return catalogo

def auto_aprovisionar_lideres_sector(cod_sector, nombre_sector=""):
    """
    Crea o vincula automáticamente las cuentas de usuario de todas las Líderes de Negocio
    asociadas a un sector específico basándose en catalogo_sectores.json.
    """
    creadas = []
    sec_clean = str(cod_sector).strip()
    if not sec_clean:
        return creadas

    catalogo = cargar_catalogo_sectores()
    sec_info = catalogo.get(sec_clean)
    if not sec_info:
        for k, v in catalogo.items():
            if str(v.get('nombre_sector', '')).strip().lower() == str(nombre_sector).strip().lower():
                sec_info = v
                break

    if not sec_info or not sec_info.get('lideres'):
        return creadas

    usuarios = cargar_usuarios()
    cambios = False

    for lider in sec_info.get('lideres', []):
        g = lider.get('codigo_grupo', '').strip()
        nom = lider.get('nombre_lider', '').strip()
        if not g or g.lower() in ['nan', '-', 'none', '0']:
            continue

        # Validar y normalizar canónicamente el nombre de sector
        nom_sec_real = nombre_sector
        if sec_clean == "700000459":
            nom_sec_real = "MATICES CLERY"
        elif sec_clean == "700000466":
            nom_sec_real = "EMOCIONES DOLLY"
        elif not nom_sec_real:
            nom_sec_real = sec_info.get('nombre_sector', f'Sector {sec_clean}')

        username = f"lider{g}".lower()
        if username not in usuarios:
            usuarios[username] = {
                "nombre": nom,
                "password_hash": hashlib_sha256("lider123"),
                "rol": "lider",
                "codigo_grupo": g,
                "codigo_sector": sec_clean,
                "nombre_sector": nom_sec_real,
                "telefono": "",
                "debe_cambiar_password": False,
                "estado_suscripcion": "activo",
                "fecha_vencimiento": None
            }
            creadas.append({'username': username, 'nombre': nom, 'grupo': g})
            cambios = True
        else:
            user_u = usuarios[username]
            if not user_u.get('codigo_sector') or user_u.get('codigo_sector') != sec_clean:
                user_u['codigo_sector'] = sec_clean
                user_u['nombre_sector'] = nom_sec_real
                cambios = True
            elif user_u.get('nombre_sector') != nom_sec_real:
                user_u['nombre_sector'] = nom_sec_real
                cambios = True

    if cambios:
        guardar_usuarios(usuarios)
        sincronizar_usuarios_a_sqlite()

    return creadas

def registrar_nueva_gerente(nombre, correo, password, telefono, cod_sector, nombre_sector=""):
    """
    Registra a una nueva Gerente de forma autónoma con 15 días de prueba gratis.
    Valida que el correo no exista y que el código de sector no haya usado ya la prueba gratuita.
    Además, aprovisiona automáticamente las cuentas de todas las líderes de su sector.
    """
    from datetime import datetime, timedelta
    
    u_clean = str(correo).strip().lower()
    sec_clean = str(cod_sector).strip()
    nom_clean = str(nombre).strip()
    tel_clean = str(telefono).strip()
    sec_nom_clean = str(nombre_sector).strip() or f"Sector {sec_clean}"

    if not nom_clean:
        return False, "Por favor ingresa tu nombre completo.", None
    if not u_clean or '@' not in u_clean or '.' not in u_clean:
        return False, "Por favor ingresa un correo electrónico válido (ej. tu_correo@gmail.com).", None
    if not password or len(str(password).strip()) < 6:
        return False, "La contraseña debe tener mínimo 6 caracteres.", None
    if not sec_clean or len(sec_clean) < 3:
        return False, "Por favor ingresa un Código de Sector válido (ej. 700000459).", None

    usuarios = cargar_usuarios()
    if u_clean in usuarios:
        return False, f"El correo '{u_clean}' ya está registrado. Por favor inicia sesión o recupera tu contraseña.", None

    # Candado Anti-Fraude: Verificar si el sector ya disfrutó de su prueba gratis
    historico = cargar_historico_sectores()
    if sec_clean in historico:
        sec_reg = historico[sec_clean]
        if sec_reg.get("ha_consumido_prueba", False) and not sec_reg.get("ha_pagado", False):
            return False, (
                f"⚠️ **El Código de Sector {sec_clean} ({sec_reg.get('nombre_sector', '')}) ya utilizó su periodo de prueba gratuita de 15 días.**\n\n"
                f"Para activar la suscripción de este sector y habilitar a todo tu equipo de líderes, "
                f"comunícate con Soporte Administrativo al WhatsApp **3057939537**."
            ), None

    now = datetime.now()
    vencimiento = now + timedelta(days=15)
    
    nuevo_usuario = {
        "nombre": nom_clean,
        "password_hash": hashlib_sha256(password),
        "rol": "gerente",
        "codigo_grupo": None,
        "codigo_sector": sec_clean,
        "nombre_sector": sec_nom_clean,
        "telefono": tel_clean,
        "fecha_registro": now.isoformat(),
        "fecha_vencimiento": vencimiento.isoformat(),
        "estado_suscripcion": "prueba",
        "debe_cambiar_password": False
    }

    usuarios[u_clean] = nuevo_usuario
    guardar_usuarios(usuarios)
    sincronizar_usuarios_a_sqlite()

    # Actualizar histórico de sectores
    historico[sec_clean] = {
        "codigo_sector": sec_clean,
        "nombre_sector": sec_nom_clean,
        "primera_prueba_fecha": now.isoformat(),
        "ha_consumido_prueba": True,
        "ha_pagado": False,
        "estado": "prueba",
        "fecha_vencimiento": vencimiento.isoformat(),
        "correo_gerente": u_clean,
        "telefono_gerente": tel_clean
    }
    guardar_historico_sectores(historico)

    # Auto-aprovisionar instantáneamente las cuentas de las líderes de este sector
    lideres_creadas = auto_aprovisionar_lideres_sector(sec_clean, sec_nom_clean)
    msg_lideres = f" y {len(lideres_creadas)} cuentas de Líderes aprovisionadas automáticamente" if lideres_creadas else ""

    user_session_info = nuevo_usuario.copy()
    user_session_info["username"] = u_clean

    return True, f"¡Bienvenida {nom_clean}! Tu cuenta de Gerente y tu prueba gratis de 15 días han sido activadas exitosamente{msg_lideres}.", user_session_info

def actualizar_suscripcion_sector(cod_sector, nuevo_estado, dias_extension=0, es_pago=False):
    """
    Permite al Super Administrador:
    - Activar plan pagado (+30, +90, +365 días o permanente)
    - Dar prórroga de prueba (+5 días)
    - Suspender o desbloquear un sector
    Aplica el cambio en cascada a la Gerente y a todas sus Líderes en usuarios.json y sectores_historico.json.
    """
    from datetime import datetime, timedelta

    sec_clean = str(cod_sector).strip()
    historico = cargar_historico_sectores()
    usuarios = cargar_usuarios()

    now = datetime.now()

    if sec_clean not in historico:
        historico[sec_clean] = {
            "codigo_sector": sec_clean,
            "nombre_sector": f"Sector {sec_clean}",
            "primera_prueba_fecha": now.isoformat(),
            "ha_consumido_prueba": True,
            "ha_pagado": es_pago or (nuevo_estado == "activo"),
            "estado": nuevo_estado
        }

    vence_iso = None
    if dias_extension > 0:
        dt_base = now
        curr_vence = historico[sec_clean].get("fecha_vencimiento")
        if curr_vence:
            try:
                dt_curr = datetime.fromisoformat(curr_vence)
                if dt_curr > now:
                    dt_base = dt_curr
            except Exception:
                pass
        vence_iso = (dt_base + timedelta(days=dias_extension)).isoformat()
    elif dias_extension == -1:
        vence_iso = None
    else:
        vence_iso = historico[sec_clean].get("fecha_vencimiento")

    historico[sec_clean]["estado"] = nuevo_estado
    historico[sec_clean]["fecha_vencimiento"] = vence_iso
    if es_pago or nuevo_estado == "activo":
        historico[sec_clean]["ha_pagado"] = True
    guardar_historico_sectores(historico)

    cambiados = 0
    for u_k, u_v in usuarios.items():
        if str(u_v.get("codigo_sector") or "").strip() == sec_clean:
            u_v["estado_suscripcion"] = nuevo_estado
            u_v["fecha_vencimiento"] = vence_iso
            cambiados += 1

    guardar_usuarios(usuarios)
    sincronizar_usuarios_a_sqlite()
    return True, f"Sector {sec_clean} actualizado a '{nuevo_estado}'. {cambiados} cuentas asociadas actualizadas."

def obtener_resumen_suscripciones():
    """
    Retorna un DataFrame con todos los sectores registrados y el estado de sus suscripciones para el panel de Super Admin.
    """
    from datetime import datetime
    historico = cargar_historico_sectores()
    usuarios = cargar_usuarios()

    filas = []
    now = datetime.now()

    for sec_id, info in historico.items():
        nom_sec = info.get("nombre_sector", f"Sector {sec_id}")
        estado = info.get("estado", "activo")
        vence_iso = info.get("fecha_vencimiento")
        correo_g = info.get("correo_gerente", "")
        tel_g = info.get("telefono_gerente", "")
        ha_pagado = info.get("ha_pagado", False)

        nom_gerente = ""
        total_lideres = 0
        for u_id, u_data in usuarios.items():
            if str(u_data.get("codigo_sector") or "").strip() == str(sec_id).strip():
                if u_data.get("rol") == "gerente":
                    nom_gerente = u_data.get("nombre", "")
                    if not correo_g:
                        correo_g = u_id
                    if not tel_g:
                        tel_g = u_data.get("telefono", "")
                elif u_data.get("rol") == "lider":
                    total_lideres += 1

        dias_rest = "Indefinido"
        vence_str = "Permanente"
        estado_label = "🟢 Activo (Pagado)" if ha_pagado else "🟢 Activo"

        if vence_iso:
            try:
                dt_vence = datetime.fromisoformat(vence_iso)
                diff = (dt_vence - now).total_seconds()
                dias_num = max(0, int(diff // 86400) + 1)
                vence_str = dt_vence.strftime("%d/%m/%Y")
                
                if estado == "bloqueado":
                    estado_label = "⛔ Suspendido / Bloqueado"
                    dias_rest = "0 días"
                elif diff < 0:
                    estado_label = "🔴 Vencido (Requiere Pago)"
                    dias_rest = "0 días (Expirado)"
                elif estado == "prueba":
                    estado_label = f"⏳ En Prueba ({dias_num} días)"
                    dias_rest = f"{dias_num} días restantes"
                else:
                    estado_label = f"🟢 Activo ({dias_num} días)"
                    dias_rest = f"{dias_num} días restantes"
            except Exception:
                pass
        elif estado == "bloqueado":
            estado_label = "⛔ Suspendido / Bloqueado"
            dias_rest = "0 días"

        filas.append({
            "Código Sector": sec_id,
            "Nombre Sector": nom_sec,
            "Gerente Responsable": nom_gerente or correo_g or "N/D",
            "Contacto (WhatsApp)": tel_g or "N/D",
            "Líderes Activas": total_lideres,
            "Estado": estado_label,
            "Vence el": vence_str,
            "Tiempo Restante": dias_rest,
            "_raw_estado": estado,
            "_raw_sector": sec_id
        })

    return pd.DataFrame(filas)

def buscar_cuenta_usuario(identificador):
    """
    Busca una cuenta en usuarios.json o en el histórico de sectores por:
    1. Username exacto (ej. 'lider9334', 'gerente', 'dolly@gmail.com')
    2. Correo electrónico (o prefijo de correo antes de @ o dominio)
    3. Código de Grupo (ej. '9334' o 'lider9334')
    4. Código de Sector (ej. '700000459' o '700000466')
    5. Nombre de la persona (ej. 'Dolly', 'Clery', 'Yenny')
    Retorna (encontrado: bool, user_data: dict, mensaje: str)
    """
    if not identificador:
        return False, None, "Por favor ingresa tu usuario, correo o código."
        
    ident_clean = str(identificador).strip().lower()
    usuarios = cargar_usuarios()
    
    # 1. Búsqueda directa por username
    if ident_clean in usuarios:
        u_info = usuarios[ident_clean].copy()
        u_info['username'] = ident_clean
        return True, u_info, "Cuenta encontrada por usuario oficial."
        
    # 2. Búsqueda por grupo (ej. '9334' -> lider9334)
    grp_num = ident_clean.replace('lider', '').strip()
    if grp_num.isdigit():
        for u, data in usuarios.items():
            if str(data.get('codigo_grupo', '')).strip() == grp_num:
                u_info = data.copy()
                u_info['username'] = u
                return True, u_info, f"Cuenta encontrada por Código de Grupo {grp_num}."
                
    # 3. Búsqueda por código de sector (ej. '700000459', '700000466')
    if ident_clean.isdigit() and len(ident_clean) >= 6:
        for u, data in usuarios.items():
            if str(data.get('codigo_sector', '')).strip() == ident_clean and data.get('rol') == 'gerente':
                u_info = data.copy()
                u_info['username'] = u
                return True, u_info, f"Cuenta de Gerente encontrada por Código de Sector {ident_clean}."

    # 4. Búsqueda por correo o coincidencia de nombre (ej. 'dolly.parra@natura.net' o 'dolly')
    ident_part = ident_clean.split('@')[0].replace('.', '').replace('_', '')
    for u, data in usuarios.items():
        u_part = u.split('@')[0].replace('.', '').replace('_', '')
        nom_part = str(data.get('nombre', '')).lower().replace(' ', '')
        
        # Coincidencia por correo o nombre
        if (len(ident_part) >= 3 and (ident_part in u_part or u_part in ident_part)) or \
           (len(ident_part) >= 4 and (ident_part in nom_part or ident_clean.split('.')[0] in nom_part)):
            u_info = data.copy()
            u_info['username'] = u
            return True, u_info, f"Cuenta encontrada por coincidencia de datos ({data.get('nombre', u)})."
            
    return False, None, "No encontramos ninguna cuenta registrada con esos datos. Por favor verifica o contacta a Soporte."

def autenticar_usuario(username, password):
    """
    Valida credenciales. Retorna el diccionario del usuario si es correcto o None.
    Permite autenticarse por usuario oficial, correo o alias asociado.
    """
    u_clean = str(username).strip().lower()
    usuarios = cargar_usuarios()
    
    # 1. Intento directo por username exacto
    if u_clean in usuarios:
        user_info = usuarios[u_clean]
        p_hash = hashlib_sha256(password)
        if user_info.get("password_hash") == p_hash:
            user_copy = user_info.copy()
            user_copy["username"] = u_clean
            return user_copy
            
    # 2. Intento por resolución de alias / correo / código de sector / grupo
    ok_b, u_found, _ = buscar_cuenta_usuario(username)
    if ok_b and u_found:
        real_u = u_found['username']
        p_hash = hashlib_sha256(password)
        if u_found.get("password_hash") == p_hash:
            user_copy = u_found.copy()
            user_copy["username"] = real_u
            return user_copy
            
    return None

SESSION_SECRET_KEY = "natura_avon_meta_indicadores_secure_session_2026"

def generar_token_sesion(user_dict):
    """
    Genera un token firmado digitalmente con HMAC-SHA256 para persistir la sesión
    de forma segura ante recargas de página (F5 / refresh) en el navegador.
    """
    if not user_dict or not user_dict.get("username"):
        return ""
    try:
        payload = {
            "username": str(user_dict.get("username")).strip().lower(),
            "exp": int(time.time()) + (48 * 3600)  # Válido por 48 horas para máxima persistencia
        }
        payload_json = json.dumps(payload)
        payload_b64 = base64.urlsafe_b64encode(payload_json.encode('utf-8')).decode('utf-8')
        sig = hmac.new(SESSION_SECRET_KEY.encode('utf-8'), payload_b64.encode('utf-8'), hashlib.sha256).hexdigest()
        return f"{payload_b64}.{sig}"
    except Exception:
        return ""

def validar_token_sesion(token_str):
    """
    Valida la firma criptográfica y expiración del token.
    Si es válido, retorna el diccionario completo del usuario cargado de la base.
    """
    try:
        if not token_str or "." not in str(token_str):
            return None
        payload_b64, sig = str(token_str).split(".", 1)
        expected_sig = hmac.new(SESSION_SECRET_KEY.encode('utf-8'), payload_b64.encode('utf-8'), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(sig, expected_sig):
            return None
        payload = json.loads(base64.urlsafe_b64decode(payload_b64.encode('utf-8')).decode('utf-8'))
        if payload.get("exp", 0) < int(time.time()):
            return None
        username = str(payload.get("username", "")).strip().lower()
        if not username:
            return None
        usuarios = cargar_usuarios()
        if username in usuarios:
            user_copy = usuarios[username].copy()
            user_copy["username"] = username
            return user_copy
        return None
    except Exception:
        return None

def cambiar_password_usuario(username, nueva_password):
    """
    Actualiza la contraseña de un usuario y remueve la bandera de cambio obligatorio.
    """
    u_clean = str(username).strip().lower()
    usuarios = cargar_usuarios()
    if u_clean in usuarios:
        usuarios[u_clean]["password_hash"] = hashlib_sha256(nueva_password)
        usuarios[u_clean]["debe_cambiar_password"] = False
        if guardar_usuarios(usuarios):
            sincronizar_usuarios_a_sqlite()
            return True, "¡Contraseña actualizada exitosamente! Ya puedes navegar por el sistema."
    return False, "Error al guardar la nueva contraseña."

def restablecer_password_usuario(username, nueva_password="lider123", debe_cambiar=False):
    """
    Restablece la contraseña de un usuario a un valor por defecto o especificado
    y sincroniza usuarios.json con la base SQLite.
    """
    u_clean = str(username).strip().lower()
    usuarios = cargar_usuarios()
    if u_clean not in usuarios:
        return False, f"El usuario '{u_clean}' no existe en el sistema."
        
    usuarios[u_clean]["password_hash"] = hashlib_sha256(nueva_password)
    usuarios[u_clean]["debe_cambiar_password"] = debe_cambiar
    if guardar_usuarios(usuarios):
        sincronizar_usuarios_a_sqlite()
        return True, f"¡Contraseña del usuario '{u_clean}' restablecida exitosamente a '{nueva_password}'!"
    return False, "Error al guardar el archivo de usuarios."

def registrar_o_actualizar_usuario(username, nombre, password=None, rol="lider", codigo_grupo=None, codigo_sector=None, debe_cambiar_password=None, nombre_sector=None):
    """
    Permite al Gerente o Superadmin crear o actualizar un usuario.
    Si se proporciona una contraseña (usuario nuevo o reseteo), debe_cambiar_password se establece en True por defecto.
    """
    u_clean = str(username).strip().lower()
    if not u_clean:
        return False, "El nombre de usuario no puede estar vacío."
    
    usuarios = cargar_usuarios()
    usuario_previo = usuarios.get(u_clean, {})
    
    if password:
        p_hash = hashlib_sha256(password)
        req_cambio = True if debe_cambiar_password is None else debe_cambiar_password
    else:
        p_hash = usuario_previo.get("password_hash", hashlib_sha256("123456"))
        req_cambio = usuario_previo.get("debe_cambiar_password", False) if debe_cambiar_password is None else debe_cambiar_password

    usr_data = {
        "nombre": nombre,
        "password_hash": p_hash,
        "rol": rol,
        "codigo_grupo": str(codigo_grupo).strip() if codigo_grupo else None,
        "debe_cambiar_password": req_cambio
    }
    
    sec_id = None
    if codigo_sector:
        sec_id = str(codigo_sector).strip()
        usr_data["codigo_sector"] = sec_id
    elif u_clean in usuarios and "codigo_sector" in usuarios[u_clean]:
        sec_id = str(usuarios[u_clean]["codigo_sector"]).strip()
        usr_data["codigo_sector"] = sec_id

    # Asignar o heredar nombre del sector de forma canónica y autoritativa
    if sec_id == "700000459":
        usr_data["nombre_sector"] = "MATICES CLERY"
    elif sec_id == "700000466":
        usr_data["nombre_sector"] = "EMOCIONES DOLLY"
    elif nombre_sector:
        usr_data["nombre_sector"] = str(nombre_sector).strip()
    elif u_clean in usuarios and "nombre_sector" in usuarios[u_clean]:
        prev_nom = usuarios[u_clean]["nombre_sector"]
        if sec_id == "700000466" and "matices" in str(prev_nom).lower():
            usr_data["nombre_sector"] = "EMOCIONES DOLLY"
        elif sec_id == "700000459" and "emociones" in str(prev_nom).lower():
            usr_data["nombre_sector"] = "MATICES CLERY"
        else:
            usr_data["nombre_sector"] = prev_nom

    # Si pertenece a un sector, heredar estado de suscripción, fecha de vencimiento y nombre de sector si no lo tiene
    if sec_id:
        historico = cargar_historico_sectores()
        if sec_id in historico:
            usr_data["estado_suscripcion"] = historico[sec_id].get("estado", "activo")
            usr_data["fecha_vencimiento"] = historico[sec_id].get("fecha_vencimiento")
            if "nombre_sector" not in usr_data or not usr_data["nombre_sector"]:
                usr_data["nombre_sector"] = historico[sec_id].get("nombre_sector", f"Sector {sec_id}")
        elif nombre_sector:
            # Registrar nuevo sector en el histórico si no existía
            from datetime import datetime
            historico[sec_id] = {
                "codigo_sector": sec_id,
                "nombre_sector": str(nombre_sector).strip(),
                "primera_prueba_fecha": datetime.now().isoformat(),
                "ha_consumido_prueba": True,
                "ha_pagado": True,
                "estado": "activo"
            }
            guardar_historico_sectores(historico)

    usuarios[u_clean] = usr_data
    if guardar_usuarios(usuarios):
        sincronizar_usuarios_a_sqlite()
        return True, f"Usuario '{u_clean}' guardado exitosamente."
    return False, "Error al guardar el archivo de usuarios."

def eliminar_usuario_perfil(username, eliminar_sector_asociado=False):
    """
    Elimina la cuenta de un usuario de usuarios.json y sincroniza con SQLite.
    Protege la cuenta 'admin' para evitar auto-bloqueo.
    Si eliminar_sector_asociado es True, también limpia el sector de sectores_historico.json.
    """
    u_clean = str(username).strip().lower()
    if not u_clean:
        return False, "Nombre de usuario no válido."

    if u_clean == "admin":
        return False, "⚠️ No se puede eliminar la cuenta principal del Super Administrador ('admin')."

    usuarios = cargar_usuarios()
    if u_clean not in usuarios:
        return False, f"El usuario '{u_clean}' no existe en el sistema."

    u_data = usuarios[u_clean]
    sec_id = u_data.get("codigo_sector")
    rol = u_data.get("rol")

    del usuarios[u_clean]
    guardar_usuarios(usuarios)
    sincronizar_usuarios_a_sqlite()

    # Si se solicitó eliminar el registro del sector en el histórico
    if eliminar_sector_asociado and sec_id and rol == "gerente":
        historico = cargar_historico_sectores()
        if str(sec_id) in historico:
            del historico[str(sec_id)]
            guardar_historico_sectores(historico)

    return True, f"¡Perfil de usuario '{u_clean}' eliminado exitosamente!"

def validar_sector_archivo(origen_file, sector_esperado):
    """
    Inspecciona un archivo subido (o ruta o DataFrame) y verifica de manera flexible y robusta
    que el sector del archivo coincida con el sector del usuario en sesión (por código o por nombre).
    Soporta múltiples hojas de cálculo, detección automática de cabeceras desplazadas,
    y comparación por códigos (completos o cortos) y nombres/palabras clave del sector.
    Retorna (valido: bool, sector_encontrado: str, nombre_sector: str, mensaje: str).
    """
    if not sector_esperado:
        return False, None, None, "⚠️ Tu perfil de Gerente no tiene un código de sector configurado en el sistema. Por favor solicita la asignación de tu código de sector."
    
    sector_esp_str = str(sector_esperado).strip()
    
    # Obtener catálogo de sectores y usuarios para conocer todos los alias y nombres de este sector
    sectores_cat = cargar_catalogo_sectores()
    usuarios_todos = cargar_usuarios()
    historico_sec = cargar_historico_sectores()
    
    nombres_esperados = set()
    codigos_esperados = {sector_esp_str}
    
    # Extraer código corto si aplica (ej. 466 de 700000466)
    if len(sector_esp_str) > 3 and sector_esp_str.startswith("700000"):
        codigos_esperados.add(sector_esp_str[6:])
        if sector_esp_str[6:].isdigit():
            codigos_esperados.add(str(int(sector_esp_str[6:])))
    
    # Buscar nombres asociados al código de sector esperado
    if sector_esp_str in sectores_cat:
        info_cat = sectores_cat[sector_esp_str]
        nom_sec_cat = info_cat.get("nombre_sector", "")
        nom_orig_cat = info_cat.get("nombre_sector_original", "")
        if nom_sec_cat:
            nombres_esperados.add(nom_sec_cat.lower().strip())
        if nom_orig_cat:
            nombres_esperados.add(nom_orig_cat.lower().strip())
            
    if sector_esp_str in historico_sec:
        nom_h = historico_sec[sector_esp_str].get("nombre_sector", "")
        if nom_h:
            nombres_esperados.add(nom_h.lower().strip())

    for u_k, u_v in usuarios_todos.items():
        if str(u_v.get("codigo_sector", "")).strip() == sector_esp_str:
            nom_u_sec = u_v.get("nombre_sector", "")
            if nom_u_sec:
                nombres_esperados.add(nom_u_sec.lower().strip())
            nom_u_persona = u_v.get("nombre", "")
            if nom_u_persona:
                nombres_esperados.add(nom_u_persona.lower().strip())

    # Palabras clave individuales significativas (ej. 'dolly', 'emociones', 'matices', 'clery')
    palabras_clave = set()
    for ne in nombres_esperados:
        for w in ne.replace("sector", "").replace("gerencia", "").split():
            w_clean = w.strip().lower()
            if len(w_clean) >= 3:
                palabras_clave.add(w_clean)

    # Identificar hojas a inspeccionar
    dfs_a_inspeccionar = []
    
    try:
        if isinstance(origen_file, pd.DataFrame):
            dfs_a_inspeccionar.append(("DataFrame", origen_file.head(50)))
        elif hasattr(origen_file, 'read'):
            try:
                origen_file.seek(0)
            except Exception:
                pass
            xl = pd.ExcelFile(origen_file)
            hojas = xl.sheet_names
            hojas_prioritarias = [s for s in hojas if any(k in s.lower() for k in ['como vamos', 'base', 'metas', 'desaf', 'datos', 'hoja1', 'sheet1'])]
            if not hojas_prioritarias:
                hojas_prioritarias = hojas[:2]
            for h in hojas_prioritarias:
                try:
                    df_h = xl.parse(h, nrows=50)
                    dfs_a_inspeccionar.append((h, df_h))
                except Exception:
                    pass
            try:
                origen_file.seek(0)
            except Exception:
                pass
        elif isinstance(origen_file, str):
            if os.path.exists(origen_file):
                xl = pd.ExcelFile(origen_file)
                hojas = xl.sheet_names
                hojas_prioritarias = [s for s in hojas if any(k in s.lower() for k in ['como vamos', 'base', 'metas', 'desaf', 'datos', 'hoja1', 'sheet1'])]
                if not hojas_prioritarias:
                    hojas_prioritarias = hojas[:2]
                for h in hojas_prioritarias:
                    try:
                        df_h = xl.parse(h, nrows=50)
                        dfs_a_inspeccionar.append((h, df_h))
                    except Exception:
                        pass
    except Exception as e:
        return False, None, None, f"Error al leer la estructura del archivo para validación: {e}"

    if not dfs_a_inspeccionar:
        return False, None, None, "El archivo está vacío o no contiene hojas legibles."

    ultimo_sec_encontrado = None
    ultimo_nom_encontrado = None
    encontro_columna_sector = False

    for nombre_hoja, df_check in dfs_a_inspeccionar:
        if df_check is None or df_check.empty:
            continue

        # Detectar si la primera fila contiene 'Unnamed' y promover la cabecera real
        if any('unnamed' in str(c).lower() for c in df_check.columns[:5]):
            for r_idx in range(min(8, len(df_check))):
                row_vals = [str(x).lower() for x in df_check.iloc[r_idx].values if pd.notna(x)]
                if any('sector' in x or 'setor' in x or 'gerencia' in x or 'consultora' in x or 'grupo' in x for x in row_vals):
                    df_check.columns = [str(col_name).strip() for col_name in df_check.iloc[r_idx]]
                    df_check = df_check.iloc[r_idx + 1:].reset_index(drop=True)
                    break

        col_sec_cod = None
        col_sec_nom = None

        # Identificar columnas de código y de nombre de sector
        for c in df_check.columns:
            c_clean = str(c).replace('\ufffd', 'ó').strip()
            c_low = c_clean.lower()
            if ('cod' in c_low or 'cd' in c_low or 'cód' in c_low) and ('sector' in c_low or 'setor' in c_low):
                col_sec_cod = c
            elif ('nom' in c_low or 'nombre' in c_low) and ('sector' in c_low or 'setor' in c_low):
                col_sec_nom = c
            elif 'sector' in c_low or 'setor' in c_low:
                if not col_sec_cod and not col_sec_nom:
                    muestra = df_check[c].dropna().astype(str).str.strip()
                    if not muestra.empty:
                        primer_val = muestra.iloc[0]
                        if primer_val.replace('.0', '').isdigit():
                            col_sec_cod = c
                        else:
                            col_sec_nom = c

        if col_sec_cod or col_sec_nom:
            encontro_columna_sector = True

        # Validar por código de sector
        if col_sec_cod:
            valores_cod = df_check[col_sec_cod].dropna().astype(str).str.strip().tolist()
            for v in valores_cod:
                v_clean = v.split('.')[0].strip()
                if v_clean:
                    ultimo_sec_encontrado = v_clean
                    if v_clean in codigos_esperados:
                        return True, v_clean, str(ultimo_nom_encontrado or ""), "Validación de sector exitosa."

        # Validar por nombre de sector
        if col_sec_nom:
            valores_nom = df_check[col_sec_nom].dropna().astype(str).str.strip().tolist()
            for v in valores_nom:
                v_clean = v.strip()
                if v_clean:
                    ultimo_nom_encontrado = v_clean
                    v_low = v_clean.lower()
                    if any(ne in v_low for ne in nombres_esperados) or any(pk in v_low for pk in palabras_clave if len(pk) >= 4):
                        return True, str(ultimo_sec_encontrado or sector_esp_str), v_clean, "Validación de sector exitosa."

    # Si no se encontró ninguna columna de sector en ninguna hoja, permitir la carga
    if not encontro_columna_sector:
        return True, None, None, "No se encontró columna explícita de sector para validar."

    # Si se encontraron datos pero pertenecían a otro sector
    nom_sec_mostrar = f" ({ultimo_nom_encontrado})" if ultimo_nom_encontrado else ""
    cod_sec_mostrar = f"`{ultimo_sec_encontrado}`" if ultimo_sec_encontrado else ""
    
    # Nombre del sector del usuario
    nombre_usuario_sec = ""
    for ne in nombres_esperados:
        if len(ne) > len(nombre_usuario_sec):
            nombre_usuario_sec = ne.title()
    if not nombre_usuario_sec:
        nombre_usuario_sec = sector_esp_str

    msg = (
        f"❌ **Acceso Denegado - Validación de Sector**\n\n"
        f"El archivo subido contiene datos de otro Sector: **{cod_sec_mostrar}{nom_sec_mostrar}**.\n"
        f"Tu perfil de Gerente está configurado para el sector **{nombre_usuario_sec}** (`{sector_esp_str}`).\n\n"
        f"Se canceló la carga para proteger y evitar sobreescribir la información de tu sector.\n\n"
        f"📲 **¿Necesitas ayuda?** Por favor comunícate con el servicio de Soporte por WhatsApp al **3057939537**."
    )
    return False, ultimo_sec_encontrado, ultimo_nom_encontrado, msg

def validar_archivo_como_vamos(origen_file, sector_esperado):
    """
    Realiza una validación completa para la subida de 'Cómo Vamos':
    1. Verifica que pertenezca al sector del usuario en sesión.
    2. Verifica que el archivo contenga columnas reales de metas y facturación (no un Reporte de Niveles).
    Retorna (valido: bool, sector_encontrado: str, nombre_sector: str, mensaje: str).
    """
    valido_sec, sec_enc, nom_sec, msg_sec = validar_sector_archivo(origen_file, sector_esperado)
    if not valido_sec:
        return False, sec_enc, nom_sec, msg_sec

    # Inspeccionar columnas para verificar si trae datos de facturación / metas
    df_check = None
    try:
        if isinstance(origen_file, pd.DataFrame):
            df_check = origen_file.head(5)
        elif hasattr(origen_file, 'read'):
            try:
                origen_file.seek(0)
            except Exception:
                pass
            xl = pd.ExcelFile(origen_file)
            hoja_leer = xl.sheet_names[0]
            for s in xl.sheet_names:
                if 'como vamos' in s.lower() or 'base' in s.lower() or 'metas' in s.lower():
                    hoja_leer = s
                    break
            df_check = pd.read_excel(origen_file, sheet_name=hoja_leer, nrows=5)
            try:
                origen_file.seek(0)
            except Exception:
                pass
        elif isinstance(origen_file, str) and os.path.exists(origen_file):
            xl = pd.ExcelFile(origen_file)
            hoja_leer = xl.sheet_names[0]
            for s in xl.sheet_names:
                if 'como vamos' in s.lower() or 'base' in s.lower() or 'metas' in s.lower():
                    hoja_leer = s
                    break
            df_check = pd.read_excel(origen_file, sheet_name=hoja_leer, nrows=5)
    except Exception as e:
        return False, sec_enc, nom_sec, f"Error al verificar la estructura del archivo: {e}"

    if df_check is not None and not df_check.empty:
        # Detectar si primera fila es Unnamed
        if any('unnamed' in str(c).lower() for c in df_check.columns[:5]):
            for r_idx in range(min(5, len(df_check))):
                row_vals = [str(x).lower() for x in df_check.iloc[r_idx].values if pd.notna(x)]
                if any('facturac' in x or 'objetivo' in x or 'real' in x or 'saldo' in x for x in row_vals):
                    df_check.columns = [str(col_name).strip() for col_name in df_check.iloc[r_idx]]
                    break

        cols_str = ' '.join([str(c).lower() for c in df_check.columns])
        tiene_facturacion = any(k in cols_str for k in ['facturac', 'objetivo facturacion', 'real facturacion', 'cumplimiento facturacion', 'objetivo activas', 'real activas', 'saldo', 'ganancia', 'falta para'])
        
        # Si parece ser un Reporte de Niveles sin metas de facturación
        if not tiene_facturacion and any(k in cols_str for k in ['nivel', 'puntos', 'pts', 'color', 'camino', 'ascenso', 'cb']):
            nom_arch = getattr(origen_file, 'name', 'Reporte')
            msg = (
                f"⚠️ **Archivo Incorrecto para 'Cómo Vamos'**\n\n"
                f"El archivo que intentas subir (`{nom_arch}`) corresponde a un **Reporte de Niveles / Puntos**, "
                f"pero **no contiene las columnas de metas ni ventas del ciclo** (`Objetivo Facturación`, `Real Facturación`, etc.).\n\n"
                f"💡 **Solución**: Por favor solicita y sube el archivo oficial **'Cómo Vamos'** de Natura desde el portal para poder analizar los tacómetros y avances de facturación de tu sector."
            )
            return False, sec_enc, nom_sec, msg

    return True, sec_enc, nom_sec, "Validación de sector y estructura exitosa."

def generar_password_aleatoria(longitud=8):
    import random, string
    chars = string.ascii_letters + string.digits
    return 'Lider' + ''.join(random.choice(chars) for _ in range(longitud))

def auto_crear_usuarios_lideres_desde_bases(ruta_tableau='Base de Datos.xlsx', ruta_como_vamos='Base para el como vamos.xlsx'):
    """
    Detecta automáticamente los grupos de líderes y crea o actualiza sus cuentas de usuario en el sistema,
    vinculándolas al código de sector correspondiente.
    """
    creados = []
    if not os.path.exists(ruta_como_vamos):
        return creados

    try:
        df_metas = pd.read_excel(ruta_como_vamos, sheet_name=0)
        df_metas = normalizar_columnas(df_metas)
    except Exception:
        return creados

    col_grp = 'Código de grupo' if 'Código de grupo' in df_metas.columns else None
    col_nom = 'Nombre de consultora' if 'Nombre de consultora' in df_metas.columns else None

    if not col_grp:
        return creados

    df_tab = None
    if os.path.exists(ruta_tableau):
        try:
            df_tab = pd.read_excel(ruta_tableau, sheet_name=0)
        except Exception:
            pass

    # Extraer código de sector si existe en los archivos
    sec_detectado = None
    if df_tab is not None:
        col_s_tab = next((c for c in ['Cod. Sector', 'cod_sector', 'Codigo Sector', 'Cd Setor', 'Cd Sector', 'Sector'] if c in df_tab.columns), None)
        if col_s_tab and not df_tab[col_s_tab].dropna().empty:
            sec_detectado = str(df_tab[col_s_tab].dropna().iloc[0]).split('.')[0].strip()
    if not sec_detectado:
        for c_m in ['Cd Setor', 'Cd Sector', 'Cod. Sector', 'Codigo Sector', 'Cod Sector', 'Sector']:
            if c_m in df_metas.columns and not df_metas[c_m].dropna().empty:
                sec_detectado = str(df_metas[c_m].dropna().iloc[0]).split('.')[0].strip()
                break

    grupos_procesados = set()
    usuarios_existentes = cargar_usuarios()

    for _, row in df_metas.iterrows():
        g_raw = row.get(col_grp)
        if pd.isna(g_raw):
            continue
        g_str = str(int(limpiar_numero(g_raw))) if limpiar_numero(g_raw) > 0 else str(g_raw).strip()
        if not g_str or g_str in grupos_procesados or g_str.lower() in ['nan', 'none', '0']:
            continue

        nom_lider = str(row.get(col_nom, '')).strip()
        if not nom_lider or nom_lider.lower() in ['nan', 'none', 'null', '0', '']:
            continue

        grupos_procesados.add(g_str)

        correo_lider = None
        sec_lider = None
        nom_sec_lider = None

        # 1. Buscar sector y nombre relacional en la base SQLite (consultoras_tableau)
        try:
            conn_sec = obtener_conexion_db()
            cur_sec = conn_sec.cursor()
            cur_sec.execute("SELECT cod_sector, sector FROM consultoras_tableau WHERE grupo = ? LIMIT 1", (g_str,))
            r_sec = cur_sec.fetchone()
            if r_sec and r_sec[0]:
                sec_lider = str(r_sec[0]).split('.')[0].strip()
                nom_sec_lider = str(r_sec[1]).strip() if r_sec[1] else None
        except Exception:
            pass

        # 2. Si no se encontró en SQLite, buscar en la fila de df_metas
        if not sec_lider:
            for c_m in ['Cd Setor', 'Cd Sector', 'Cod. Sector', 'Codigo Sector', 'Cod Sector']:
                if c_m in df_metas.columns and pd.notna(row.get(c_m)):
                    val_c = str(row.get(c_m)).split('.')[0].strip()
                    if val_c and val_c.lower() not in ['nan', 'none', '0']:
                        sec_lider = val_c
                        break
            for c_nm in ['Nombre Setor', 'Nombre Sector', 'Sector']:
                if c_nm in df_metas.columns and pd.notna(row.get(c_nm)):
                    val_nm = str(row.get(c_nm)).strip()
                    if val_nm and val_nm.lower() not in ['nan', 'none']:
                        nom_sec_lider = val_nm
                        break

        # 3. Buscar en df_tab por grupo
        if df_tab is not None:
            mask_g = (df_tab['Grupo'].astype(str).str.strip() == g_str) if 'Grupo' in df_tab.columns else pd.Series(False, index=df_tab.index)
            if mask_g.any():
                df_g = df_tab[mask_g]
                if 'Correo' in df_g.columns and not df_g.dropna(subset=['Correo']).empty:
                    correo_lider = str(df_g['Correo'].dropna().iloc[0]).strip().lower()
                col_s_g = next((c for c in ['Cod. Sector', 'cod_sector'] if c in df_g.columns), None)
                if col_s_g and not df_g[col_s_g].dropna().empty:
                    sec_lider = str(df_g[col_s_g].dropna().iloc[0]).split('.')[0].strip()
                col_sn_g = next((c for c in ['Sector', 'sector', 'Nombre Sector'] if c in df_g.columns), None)
                if col_sn_g and not df_g[col_sn_g].dropna().empty:
                    nom_sec_lider = str(df_g[col_sn_g].dropna().iloc[0]).strip()

        if not sec_lider:
            sec_lider = sec_detectado

        # Normalizar nombre del sector canónico
        if sec_lider == "700000459":
            nom_sec_lider = "MATICES CLERY"
        elif sec_lider == "700000466":
            nom_sec_lider = "EMOCIONES DOLLY"

        username = correo_lider if (correo_lider and '@' in correo_lider) else f"lider{g_str}"
        ya_existe = (username in usuarios_existentes)

        if ya_existe:
            registrar_o_actualizar_usuario(
                username=username,
                nombre=nom_lider,
                password=None,
                rol="lider",
                codigo_grupo=g_str,
                codigo_sector=sec_lider,
                nombre_sector=nom_sec_lider
            )
        else:
            pass_gen = generar_password_aleatoria()
            exito, msg = registrar_o_actualizar_usuario(
                username=username,
                nombre=nom_lider,
                password=pass_gen,
                rol="lider",
                codigo_grupo=g_str,
                codigo_sector=sec_lider,
                nombre_sector=nom_sec_lider
            )
            if exito:
                creados.append({
                    "Código Grupo": g_str,
                    "Nombre Líder": nom_lider,
                    "Usuario (Login / Correo)": username,
                    "Contraseña Generada": pass_gen,
                    "Resultado": "✨ Nueva Cuenta Creada"
                })

    return creados

def obtener_mapa_lideres():
    """
    Retorna un diccionario { 'codigo_grupo': 'Nombre Líder' }
    recuperando nombres desde usuarios.json, la base de datos SQLite y 'Base para el como vamos.xlsx'.
    """
    mapa = {}

    # 1. Desde usuarios.json
    try:
        usuarios = cargar_usuarios()
        for u_id, u_data in usuarios.items():
            grp = str(u_data.get('codigo_grupo') or '').strip()
            nom = str(u_data.get('nombre') or '').strip()
            if grp and nom and grp.lower() not in ['none', 'nan', 'null', ''] and nom.lower() not in ['none', 'nan', 'null', '']:
                mapa[grp] = nom
    except Exception:
        pass

    # 2. Desde la tabla SQLite metas_como_vamos
    try:
        conn = obtener_conexion_db()
        cursor = conn.cursor()
        rows = cursor.execute("SELECT DISTINCT codigo_grupo, nombre_consultora FROM metas_como_vamos WHERE codigo_grupo IS NOT NULL AND nombre_consultora IS NOT NULL").fetchall()
        for r in rows:
            grp = str(r[0] or '').strip().split('.')[0]
            nom = str(r[1] or '').strip()
            if grp and nom and grp.lower() not in ['none', 'nan', 'null', ''] and nom.lower() not in ['none', 'nan', 'null', '']:
                mapa[grp] = nom
        conn.close()
    except Exception:
        pass

    # 3. Desde el archivo físico 'Base para el como vamos.xlsx' si existe
    try:
        if os.path.exists('Base para el como vamos.xlsx'):
            df_m = pd.read_excel('Base para el como vamos.xlsx', sheet_name=0)
            df_m = normalizar_columnas(df_m)
            col_grp = 'Código de grupo' if 'Código de grupo' in df_m.columns else None
            col_nom = 'Nombre de consultora' if 'Nombre de consultora' in df_m.columns else None
            if col_grp and col_nom:
                for _, row in df_m.iterrows():
                    g = str(row[col_grp] or '').strip().split('.')[0]
                    n = str(row[col_nom] or '').strip()
                    if g and n and g.lower() not in ['none', 'nan', 'null', ''] and n.lower() not in ['none', 'nan', 'null', '']:
                        mapa[g] = n
    except Exception:
        pass

    return mapa

# --- CONFIGURACIÓN DE PERMISOS GLOBALES DE CARGA ---
RUTA_CONFIG = ruta_persistente('configuracion.json')

DEFAULT_PERMISOS_PESTANAS = {
    "tab_tableau": {"nombre": "📊 Informe Tableau Cam", "gerente": True, "lider": True, "asesor": False},
    "tab_geral": {"nombre": "💳 Geral_Credito&Cobranza", "gerente": True, "lider": True, "asesor": False},
    "tab_resumen": {"nombre": "📊 Resumen & KPIs", "gerente": True, "lider": True, "asesor": False},
    "tab_ganancia": {"nombre": "🧮 Simuladores", "gerente": True, "lider": True, "asesor": False},
    "tab_diagnostico": {"nombre": "👑 Mis Líderes", "gerente": True, "lider": True, "asesor": True},
    "tab_metas": {"nombre": "🎯 Metas de Crecimiento (Procesador)", "gerente": True, "lider": True, "asesor": False},
    "tab_detalle": {"nombre": "📑 Generador de Informes", "gerente": True, "lider": True, "asesor": False}
}

def cargar_configuracion():
    """
    Carga la configuración global de la aplicación.
    Por defecto, incluye los permisos de visibilidad por pestaña con nombres actualizados.
    """
    config = {
        "permitir_carga_lideres": False,
        "permisos_pestanas": {k: v.copy() for k, v in DEFAULT_PERMISOS_PESTANAS.items()}
    }
    
    if os.path.exists(RUTA_CONFIG):
        try:
            with open(RUTA_CONFIG, 'r', encoding='utf-8') as f:
                loaded = json.load(f)
                if isinstance(loaded, dict):
                    if "permitir_carga_lideres" in loaded:
                        config["permitir_carga_lideres"] = loaded["permitir_carga_lideres"]
                    if "permisos_pestanas" in loaded and isinstance(loaded["permisos_pestanas"], dict):
                        for tab_key, tab_val in DEFAULT_PERMISOS_PESTANAS.items():
                            if tab_key in loaded["permisos_pestanas"]:
                                for r_key in ["gerente", "lider", "asesor"]:
                                    if r_key in loaded["permisos_pestanas"][tab_key]:
                                        config["permisos_pestanas"][tab_key][r_key] = bool(loaded["permisos_pestanas"][tab_key][r_key])
                            config["permisos_pestanas"][tab_key]["nombre"] = tab_val["nombre"]
        except Exception as e:
            print(f"Nota al cargar configuración: {e}")
            
    guardar_configuracion(config)
    return config

def guardar_configuracion(dict_config):
    """
    Guarda la configuración global en configuracion.json.
    """
    try:
        with open(RUTA_CONFIG, 'w', encoding='utf-8') as f:
            json.dump(dict_config, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"Error al guardar configuración: {e}")
        return False

# --- MOTOR DE BASE DE DATOS RELACIONAL SQLITE (base_matices.db) ---
import sqlite3

RUTA_DB_SQLITE = ruta_persistente('base_matices.db')

def obtener_conexion_db():
    conn = sqlite3.connect(RUTA_DB_SQLITE, timeout=30.0)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute("PRAGMA journal_mode=WAL;")
    except Exception:
        pass
    return conn

def inicializar_db_sqlite(conn=None):
    """
    Inicializa la base de datos relacional SQLite 'base_matices.db' y crea las tablas indexadas.
    Sincroniza automáticamente los datos de Excel y JSON si la base se crea por primera vez.
    """
    close_at_end = False
    if conn is None:
        conn = obtener_conexion_db()
        close_at_end = True

    cursor = conn.cursor()

    # 1. Tabla de Usuarios
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS usuarios (
        username TEXT PRIMARY KEY,
        nombre TEXT,
        password_hash TEXT,
        rol TEXT,
        codigo_grupo TEXT,
        codigo_sector TEXT
    )
    """)
    
    try:
        cursor.execute("ALTER TABLE usuarios ADD COLUMN codigo_sector TEXT")
    except Exception:
        pass

    # 2. Tabla de Configuración
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS configuracion (
        clave TEXT PRIMARY KEY,
        valor TEXT
    )
    """)

    # 3. Tabla de Consultoras Tableau
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS consultoras_tableau (
        codigo_cb TEXT PRIMARY KEY,
        nombre TEXT,
        tipo_documento TEXT,
        numero_documento TEXT,
        grupo TEXT,
        cod_sector TEXT,
        sector TEXT,
        estado_actividad TEXT,
        segmento_actual TEXT,
        saldo_vencido REAL,
        credito_disponible REAL,
        pedidos_pendientes INTEGER,
        pedidos_mora INTEGER,
        celular TEXT,
        correo TEXT,
        documento_gpp TEXT,
        cod_gerencia TEXT,
        gerencia TEXT,
        ciclo INTEGER,
        color TEXT,
        ascenso_cb TEXT,
        situacion TEXT,
        sit_comercial TEXT,
        ciclos_inactividad INTEGER,
        actividad_convergencia TEXT,
        inicios_completos TEXT,
        fact_vol REAL,
        fact_natura REAL,
        fact_avon REAL,
        fact_ce REAL,
        facturacion_total REAL,
        pts_natura INTEGER,
        pts_avon INTEGER,
        pts_total_vd INTEGER,
        pts_vol INTEGER,
        pts_acum INTEGER,
        pts_mant INTEGER,
        pts_asc INTEGER,
        tienda_online TEXT,
        tienda_sf TEXT,
        ciclo_primer_pedido INTEGER,
        fecha_nacimiento TEXT,
        mes_cumpleanos TEXT,
        deuda_total REAL,
        deuda_mora REAL,
        credito_total REAL,
        dpto_residencia TEXT,
        ciudad_residencia TEXT,
        barrio_residencia TEXT,
        direccion_residencia TEXT,
        complemento_residencia TEXT,
        referencia_residencia TEXT,
        dpto_entrega TEXT,
        ciudad_entrega TEXT,
        barrio_entrega TEXT,
        direccion_entrega TEXT,
        complemento_entrega TEXT,
        referencia_entrega TEXT,
        tiempo_casa INTEGER,
        origen_cb TEXT,
        notas_lider TEXT,
        indicador TEXT
    )
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_tableau_grupo ON consultoras_tableau (grupo)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_tableau_sector ON consultoras_tableau (sector)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_tableau_cod_sector ON consultoras_tableau (cod_sector)")

    # 4. Tabla de Metas "Cómo Vamos"
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS metas_como_vamos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo_cb TEXT,
        nombre_consultora TEXT,
        nombre_gerencia TEXT,
        nombre_sector TEXT,
        codigo_grupo TEXT,
        color TEXT,
        obj_facturacion REAL,
        real_facturacion REAL,
        cump_facturacion REAL,
        obj_activas REAL,
        real_activas REAL,
        cump_activas REAL,
        saldo REAL,
        disponibles INTEGER,
        inicios INTEGER,
        reinicios INTEGER,
        recuperos INTEGER,
        ganancia_estimada REAL
    )
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_metas_grupo ON metas_como_vamos (codigo_grupo)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_metas_sector ON metas_como_vamos (nombre_sector)")

    # 5. Tabla de Cartera Geral (Crédito & Cobranza)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS cartera_geral (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        titulo TEXT,
        cuota INTEGER,
        numero_pedido TEXT,
        ciclo_captacion TEXT,
        ciclo_indicador TEXT,
        grupo TEXT,
        numero_factura TEXT,
        fecha_pedido TEXT,
        fecha_vencimiento_original TEXT,
        fecha_vencimiento TEXT,
        valor_titulo REAL,
        saldo_principal REAL,
        saldo_financiero REAL,
        saldo_total REAL,
        saldo_actualizado REAL,
        situacion TEXT,
        dias_retraso REAL,
        fase_cobro TEXT,
        cod_sector TEXT,
        sector TEXT,
        codigo_cb TEXT,
        nombre TEXT,
        direccion TEXT,
        telefono_movil TEXT,
        telefono_movil_2 TEXT,
        correo TEXT,
        plan_recibimiento TEXT,
        origen_empresa TEXT,
        fecha_carga TEXT
    )
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_geral_grupo ON cartera_geral (grupo)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_geral_cod_sector ON cartera_geral (cod_sector)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_geral_situacion ON cartera_geral (situacion)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_geral_fecha_venc ON cartera_geral (fecha_vencimiento)")

    # 6. Tabla de Auditoría, Logs & Usabilidad
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS auditoria_eventos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha_hora TEXT,
        fecha TEXT,
        hora TEXT,
        username TEXT,
        nombre TEXT,
        rol TEXT,
        codigo_sector TEXT,
        nombre_sector TEXT,
        codigo_grupo TEXT,
        categoria TEXT,
        accion TEXT,
        detalle TEXT,
        dispositivo TEXT
    )
    """)
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_auditoria_fecha ON auditoria_eventos (fecha)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_auditoria_username ON auditoria_eventos (username)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_auditoria_sector ON auditoria_eventos (codigo_sector)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_auditoria_categoria ON auditoria_eventos (categoria)")

    conn.commit()

    # Sincronización inicial
    sincronizar_usuarios_a_sqlite(conn)
    sincronizar_configuracion_a_sqlite(conn)
    sincronizar_auditoria_a_sqlite(conn)

    if close_at_end:
        conn.close()

def sincronizar_usuarios_a_sqlite(conn=None):
    close_at_end = False
    if conn is None:
        conn = obtener_conexion_db()
        close_at_end = True
    
    usuarios = cargar_usuarios()
    cursor = conn.cursor()
    
    try:
        cursor.execute("ALTER TABLE usuarios ADD COLUMN codigo_sector TEXT")
    except Exception:
        pass

    cursor.execute("DELETE FROM usuarios")

    for uname, uinfo in usuarios.items():
        cursor.execute("""
        INSERT INTO usuarios (username, nombre, password_hash, rol, codigo_grupo, codigo_sector)
        VALUES (?, ?, ?, ?, ?, ?)
        """, (uname, uinfo.get("nombre"), uinfo.get("password_hash"), uinfo.get("rol"), uinfo.get("codigo_grupo"), uinfo.get("codigo_sector")))
    conn.commit()
    if close_at_end:
        conn.close()

def sincronizar_configuracion_a_sqlite(conn=None):
    close_at_end = False
    if conn is None:
        conn = obtener_conexion_db()
        close_at_end = True
    
    cfg = cargar_configuracion()
    cursor = conn.cursor()
    for k, v in cfg.items():
        cursor.execute("""
        INSERT OR REPLACE INTO configuracion (clave, valor)
        VALUES (?, ?)
        """, (str(k), json.dumps(v)))
    conn.commit()
    if close_at_end:
        conn.close()

def sincronizar_auditoria_a_sqlite(conn=None):
    """
    Sincroniza los eventos desde auditoria_logs.json a la tabla auditoria_eventos en SQLite si la tabla está vacía.
    """
    close_at_end = False
    if conn is None:
        conn = obtener_conexion_db()
        close_at_end = True

    try:
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM auditoria_eventos")
        count = cursor.fetchone()[0]
        if count == 0:
            rutas = [RUTA_AUDITORIA_JSON, 'auditoria_logs.json']
            logs = []
            for r in rutas:
                if r and os.path.exists(r):
                    try:
                        with open(r, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                            if isinstance(data, list):
                                logs = data
                                break
                    except Exception:
                        pass
            if logs:
                for ev in logs:
                    cursor.execute("""
                    INSERT INTO auditoria_eventos (
                        fecha_hora, fecha, hora, username, nombre, rol,
                        codigo_sector, nombre_sector, codigo_grupo,
                        categoria, accion, detalle, dispositivo
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        ev.get("fecha_hora"), ev.get("fecha"), ev.get("hora"),
                        ev.get("username"), ev.get("nombre"), ev.get("rol"),
                        ev.get("codigo_sector"), ev.get("nombre_sector"), ev.get("codigo_grupo"),
                        ev.get("categoria"), ev.get("accion"), ev.get("detalle"), ev.get("dispositivo", "PC")
                    ))
                conn.commit()
    except Exception as e:
        safe_print(f"Nota al sincronizar auditoria a SQLite: {e}")
    finally:
        if close_at_end:
            conn.close()

def registrar_evento_auditoria(user_info, categoria, accion, detalle="", dispositivo="PC"):
    """
    Registra un evento de auditoría y usabilidad en SQLite y archivo persistente redundante.
    Tolerante a fallos: no bloquea ni detiene el flujo de la aplicación.
    """
    from datetime import datetime
    try:
        now = datetime.now()
        fecha_hora = now.isoformat()
        fecha = now.strftime("%Y-%m-%d")
        hora = now.strftime("%H:%M:%S")

        username = ""
        nombre = ""
        rol = ""
        cod_sector = ""
        nom_sector = ""
        cod_grupo = ""

        if isinstance(user_info, dict):
            username = str(user_info.get("username") or "").strip()
            nombre = str(user_info.get("nombre") or username).strip()
            rol = str(user_info.get("rol") or "").strip()
            cod_sector = str(user_info.get("codigo_sector") or "").strip()
            nom_sector = str(user_info.get("nombre_sector") or "").strip()
            cod_grupo = str(user_info.get("codigo_grupo") or "").strip()
        elif isinstance(user_info, str):
            username = user_info.strip()
            nombre = username
        
        if not nom_sector and cod_sector:
            nom_sector = f"Sector {cod_sector}"

        evento = {
            "fecha_hora": fecha_hora,
            "fecha": fecha,
            "hora": hora,
            "username": username,
            "nombre": nombre,
            "rol": rol,
            "codigo_sector": cod_sector,
            "nombre_sector": nom_sector,
            "codigo_grupo": cod_grupo,
            "categoria": str(categoria),
            "accion": str(accion),
            "detalle": str(detalle),
            "dispositivo": str(dispositivo)
        }

        # 1. Guardar en SQLite
        try:
            conn = obtener_conexion_db()
            cursor = conn.cursor()
            cursor.execute("""
            CREATE TABLE IF NOT EXISTS auditoria_eventos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha_hora TEXT,
                fecha TEXT,
                hora TEXT,
                username TEXT,
                nombre TEXT,
                rol TEXT,
                codigo_sector TEXT,
                nombre_sector TEXT,
                codigo_grupo TEXT,
                categoria TEXT,
                accion TEXT,
                detalle TEXT,
                dispositivo TEXT
            )
            """)
            cursor.execute("""
            INSERT INTO auditoria_eventos (
                fecha_hora, fecha, hora, username, nombre, rol,
                codigo_sector, nombre_sector, codigo_grupo,
                categoria, accion, detalle, dispositivo
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                fecha_hora, fecha, hora, username, nombre, rol,
                cod_sector, nom_sector, cod_grupo,
                str(categoria), str(accion), str(detalle), str(dispositivo)
            ))
            conn.commit()
            conn.close()
        except Exception as e_sql:
            safe_print(f"Nota SQLite auditoria: {e_sql}")

        # 2. Guardar en JSON persistente redundante
        try:
            logs = []
            rutas = [RUTA_AUDITORIA_JSON, 'auditoria_logs.json']
            for r in rutas:
                if r and os.path.exists(r):
                    try:
                        with open(r, 'r', encoding='utf-8') as f:
                            loaded = json.load(f)
                            if isinstance(loaded, list):
                                logs = loaded
                                break
                    except Exception:
                        pass
            
            logs.append(evento)
            if len(logs) > 5000:
                logs = logs[-5000:]
                
            rutas_guardar = set(filter(None, [RUTA_AUDITORIA_JSON, 'auditoria_logs.json']))
            if DIR_PERSISTENTE and os.path.isdir(DIR_PERSISTENTE):
                rutas_guardar.add(os.path.join(DIR_PERSISTENTE, 'auditoria_logs.json'))

            for r in rutas_guardar:
                p_dir = os.path.dirname(r)
                if p_dir:
                    os.makedirs(p_dir, exist_ok=True)
                with open(r, 'w', encoding='utf-8') as f:
                    json.dump(logs, f, ensure_ascii=False, indent=2)
        except Exception as e_json:
            safe_print(f"Nota JSON auditoria: {e_json}")

        return True
    except Exception as e:
        safe_print(f"Error general en registrar_evento_auditoria: {e}")
        return False

def consultar_auditoria_df(filtro_fecha_inicio=None, filtro_fecha_fin=None, filtro_usuario=None, filtro_rol=None, filtro_sector=None, filtro_categoria=None, limite=500):
    """
    Retorna un DataFrame con los registros de auditoría filtrados.
    """
    try:
        conn = obtener_conexion_db()
        cursor = conn.cursor()
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS auditoria_eventos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha_hora TEXT,
            fecha TEXT,
            hora TEXT,
            username TEXT,
            nombre TEXT,
            rol TEXT,
            codigo_sector TEXT,
            nombre_sector TEXT,
            codigo_grupo TEXT,
            categoria TEXT,
            accion TEXT,
            detalle TEXT,
            dispositivo TEXT
        )
        """)
        conn.commit()

        query = "SELECT id, fecha_hora, fecha, hora, username, nombre, rol, codigo_sector, nombre_sector, codigo_grupo, categoria, accion, detalle, dispositivo FROM auditoria_eventos WHERE 1=1"
        params = []

        if filtro_fecha_inicio:
            query += " AND fecha >= ?"
            params.append(str(filtro_fecha_inicio))
        if filtro_fecha_fin:
            query += " AND fecha <= ?"
            params.append(str(filtro_fecha_fin))
        if filtro_usuario and str(filtro_usuario).strip() and str(filtro_usuario).lower() != "todos":
            query += " AND (username = ? OR nombre LIKE ?)"
            params.extend([str(filtro_usuario).strip(), f"%{str(filtro_usuario).strip()}%"])
        if filtro_rol and str(filtro_rol).strip() and str(filtro_rol).lower() != "todos":
            query += " AND rol = ?"
            params.append(str(filtro_rol).strip())
        if filtro_sector and str(filtro_sector).strip() and str(filtro_sector).lower() != "todos":
            query += " AND (codigo_sector = ? OR nombre_sector LIKE ?)"
            params.extend([str(filtro_sector).strip(), f"%{str(filtro_sector).strip()}%"])
        if filtro_categoria and str(filtro_categoria).strip() and str(filtro_categoria).lower() != "todas":
            query += " AND categoria = ?"
            params.append(str(filtro_categoria).strip())

        query += " ORDER BY id DESC LIMIT ?"
        params.append(int(limite))

        df = pd.read_sql_query(query, conn, params=params)
        conn.close()
        return df
    except Exception as e:
        safe_print(f"Error al consultar auditoria SQLite: {e}")
        try:
            rutas = [RUTA_AUDITORIA_JSON, 'auditoria_logs.json']
            for r in rutas:
                if r and os.path.exists(r):
                    with open(r, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        if isinstance(data, list) and data:
                            df = pd.DataFrame(data)
                            if 'id' not in df.columns:
                                df['id'] = range(1, len(df) + 1)
                            return df.tail(limite).iloc[::-1].reset_index(drop=True)
        except Exception:
            pass
        return pd.DataFrame()

def obtener_metricas_usabilidad(dias_atras=30):
    """
    Calcula métricas agregadas de adopción, usabilidad y actividad para el dashboard.
    """
    df = consultar_auditoria_df(limite=5000)
    if df.empty:
        return {
            "total_eventos": 0,
            "total_logins": 0,
            "usuarios_activos_hoy": 0,
            "usuarios_activos_semana": 0,
            "ranking_usuarios": pd.DataFrame(),
            "ranking_sectores": pd.DataFrame(),
            "actividad_por_categoria": pd.DataFrame(),
            "actividad_diaria": pd.DataFrame(),
            "uso_dispositivos": pd.DataFrame(),
            "sectores_alerta": []
        }

    from datetime import datetime, timedelta
    now = datetime.now()
    hoy_str = now.strftime("%Y-%m-%d")
    hace_7_dias = (now - timedelta(days=7)).strftime("%Y-%m-%d")

    # Asegurar columnas requeridas
    for col_req in ['id', 'fecha_hora', 'fecha', 'username', 'nombre', 'rol', 'codigo_sector', 'nombre_sector', 'accion', 'categoria', 'dispositivo']:
        if col_req not in df.columns:
            df[col_req] = ""

    total_eventos = len(df)
    df_logins = df[df['accion'].astype(str).str.contains('Inicio de Sesión', case=False, na=False)]
    total_logins = len(df_logins)

    # Usuarios activos hoy
    df_hoy = df[df['fecha'] == hoy_str]
    usuarios_activos_hoy = df_hoy['username'].nunique() if not df_hoy.empty else 0

    # Usuarios activos últimos 7 días
    df_7d = df[df['fecha'] >= hace_7_dias]
    usuarios_activos_semana = df_7d['username'].nunique() if not df_7d.empty else 0

    # Ranking de usuarios más activos
    grp_user = df.groupby(['username', 'nombre', 'rol', 'codigo_sector']).agg(
        total_acciones=('fecha_hora', 'count'),
        total_logins=('accion', lambda x: sum('Inicio de Sesión' in str(v) for v in x)),
        ultimo_acceso=('fecha_hora', 'max')
    ).reset_index().sort_values(by='total_acciones', ascending=False)

    # Ranking de sectores más activos
    df_con_sec = df[df['codigo_sector'].astype(str).str.strip().str.len() > 2]
    if not df_con_sec.empty:
        grp_sec = df_con_sec.groupby(['codigo_sector', 'nombre_sector']).agg(
            total_acciones=('fecha_hora', 'count'),
            usuarios_unicos=('username', 'nunique'),
            total_logins=('accion', lambda x: sum('Inicio de Sesión' in str(v) for v in x)),
            ultimo_evento=('fecha_hora', 'max')
        ).reset_index().sort_values(by='total_acciones', ascending=False)
    else:
        grp_sec = pd.DataFrame()

    # Actividad por categoría
    act_cat = df.groupby('categoria').size().reset_index(name='total').sort_values(by='total', ascending=False)

    # Actividad diaria (últimos 14 días)
    hace_14_dias = (now - timedelta(days=14)).strftime("%Y-%m-%d")
    df_14d = df[df['fecha'] >= hace_14_dias]
    if not df_14d.empty:
        act_dia = df_14d.groupby('fecha').agg(
            total_eventos=('fecha_hora', 'count'),
            logins=('accion', lambda x: sum('Inicio de Sesión' in str(v) for v in x)),
            usuarios_unicos=('username', 'nunique')
        ).reset_index().sort_values(by='fecha')
    else:
        act_dia = pd.DataFrame()

    # Dispositivos
    disp = df.groupby('dispositivo').size().reset_index(name='total').sort_values(by='total', ascending=False)

    # Detección de sectores registrados con inactividad (para soporte / alertas)
    historico = cargar_historico_sectores()
    sectores_alerta = []
    for cod_s, info_s in historico.items():
        nom_s = info_s.get('nombre_sector', f'Sector {cod_s}')
        correo_g = info_s.get('correo_gerente', '')
        estado_s = info_s.get('estado', 'prueba')
        
        df_sec_ev = df[df['codigo_sector'] == str(cod_s)]
        if df_sec_ev.empty:
            sectores_alerta.append({
                "codigo_sector": cod_s,
                "nombre_sector": nom_s,
                "estado": estado_s,
                "correo_gerente": correo_g,
                "dias_inactivo": "Sin actividad registrada",
                "nivel_alerta": "🔴 Alta (Sin ingresos)"
            })
        else:
            ult_fecha_str = str(df_sec_ev['fecha'].max())
            try:
                dt_ult = datetime.strptime(ult_fecha_str, "%Y-%m-%d")
                dias_diff = (now - dt_ult).days
                if dias_diff >= 3:
                    sectores_alerta.append({
                        "codigo_sector": cod_s,
                        "nombre_sector": nom_s,
                        "estado": estado_s,
                        "correo_gerente": correo_g,
                        "dias_inactivo": f"{dias_diff} días",
                        "nivel_alerta": "🟡 Media (Inactivo > 3 días)" if dias_diff < 7 else "🔴 Alta (Inactivo >= 7 días)"
                    })
            except Exception:
                pass

    return {
        "total_eventos": total_eventos,
        "total_logins": total_logins,
        "usuarios_activos_hoy": usuarios_activos_hoy,
        "usuarios_activos_semana": usuarios_activos_semana,
        "ranking_usuarios": grp_user,
        "ranking_sectores": grp_sec,
        "actividad_por_categoria": act_cat,
        "actividad_diaria": act_dia,
        "uso_dispositivos": disp,
        "sectores_alerta": sectores_alerta
    }

def sincronizar_excel_tableau_a_sqlite(ruta_excel='Base de Datos.xlsx', conn=None):
    """
    Convierte y vuelca el archivo Excel de Tableau hacia la tabla consultoras_tableau en SQLite.
    """
    df = procesar_base_tableau_manager(ruta_excel)
    if df is None or df.empty:
        return False
    
    close_at_end = False
    if conn is None:
        conn = obtener_conexion_db()
        close_at_end = True
    
    cursor = conn.cursor()
    try:
        cursor.execute("ALTER TABLE consultoras_tableau ADD COLUMN indicador TEXT")
    except Exception:
        pass
    
    # Extraer el código de sector del dataframe subido para borrar ÚNICAMENTE los registros de ese sector
    col_sec_check = None
    for c in df.columns:
        c_low = str(c).lower().replace('ó', 'o')
        if ('cod' in c_low or 'cd' in c_low) and ('sector' in c_low or 'setor' in c_low):
            col_sec_check = c
            break
        elif 'sector' in c_low or 'setor' in c_low:
            col_sec_check = c
            
    if col_sec_check and not df[col_sec_check].dropna().empty:
        col_sec_found_val = str(df[col_sec_check].dropna().iloc[0]).strip().split('.')[0]
        cursor.execute("DELETE FROM consultoras_tableau WHERE cod_sector = ? OR sector LIKE ?", (col_sec_found_val, f"%{col_sec_found_val}%"))
    else:
        cursor.execute("DELETE FROM consultoras_tableau")
    
    for _, row in df.iterrows():
        cb = limpiar_codigo_cb_estandar(row.get('Codigo CB') if 'Codigo CB' in df.columns else row.get('Código CB', ''))
        if not cb or cb.lower() == 'nan':
            continue
        
        nom = str(row.get('Asesora / Consultora') if 'Asesora / Consultora' in df.columns else row.get('Nombre', ''))
        col = str(row.get('Nivel / Color') if 'Nivel / Color' in df.columns else row.get('Color', ''))
        
        cursor.execute("""
        INSERT OR REPLACE INTO consultoras_tableau (
            codigo_cb, nombre, documento_gpp, cod_gerencia, gerencia, cod_sector, sector, grupo, ciclo, color,
            ascenso_cb, situacion, sit_comercial, ciclos_inactividad, actividad_convergencia, inicios_completos,
            fact_vol, fact_natura, fact_avon, fact_ce, facturacion_total,
            pts_natura, pts_avon, pts_total_vd, pts_vol, pts_acum, pts_mant, pts_asc,
            tienda_online, tienda_sf, ciclo_primer_pedido, fecha_nacimiento, mes_cumpleanos,
            deuda_total, deuda_mora, credito_total, credito_disponible, pedidos_pendientes, pedidos_mora,
            celular, correo,
            dpto_residencia, ciudad_residencia, barrio_residencia, direccion_residencia, complemento_residencia, referencia_residencia,
            dpto_entrega, ciudad_entrega, barrio_entrega, direccion_entrega, complemento_entrega, referencia_entrega,
            tiempo_casa, origen_cb, notas_lider, indicador
        ) VALUES (
            ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?, ?,
            ?, ?,
            ?, ?, ?, ?, ?, ?,
            ?, ?, ?, ?, ?, ?,
            ?, ?, ?, ?
        )
        """, (
            cb,
            nom,
            str(row.get('DocumentoGPP', '')),
            str(row.get('Cod. Gerencia', '')),
            str(row.get('Gerencia', '')),
            str(row.get('Cod. Sector', '')),
            str(row.get('Sector', '')),
            str(int(limpiar_numero(row.get('Grupo', 0)))) if limpiar_numero(row.get('Grupo', 0)) > 0 else str(row.get('Grupo', '')),
            int(limpiar_numero(row.get('Ciclo', 0))),
            col,
            str(row.get('Ascenso CB', '')),
            str(row.get('Situación', '')),
            str(row.get('Sit. Comercial', '')),
            int(limpiar_numero(row.get('Ciclos Inactividad', 0))),
            str(row.get('Actividad Convergencia', '')),
            str(row.get('Inicios Completos', '')),
            float(limpiar_numero(row.get('Fact. VOL', 0.0))),
            float(limpiar_numero(row.get('Fact. Natura', 0.0))),
            float(limpiar_numero(row.get('Fact. AVON', 0.0))),
            float(limpiar_numero(row.get('Fact. C&E', 0.0))),
            float(limpiar_numero(row.get('Fact. Total', 0.0))),
            int(limpiar_numero(row.get('Pts Natura', 0))),
            int(limpiar_numero(row.get('Pts AVON', 0))),
            int(limpiar_numero(row.get('Pts Total VD', 0))),
            int(limpiar_numero(row.get('Pts VOL', 0))),
            int(limpiar_numero(row.get('Pts Acum', row.get('Pts Acumulados', 0)))),
            int(limpiar_numero(row.get('Pts Mant', row.get('Pts Para Mantener', row.get('Pts para Mantener', 0))))),
            int(limpiar_numero(row.get('Pts Asc', row.get('Pts para Ascender', row.get('Pts Para Ascender', 0))))),
            str(row.get('Tienda Online', '')),
            str(row.get('TiendaSF', '')),
            int(limpiar_numero(row.get('Ciclo Primer Pedido', 0))),
            str(row.get('Fecha De Nacimiento', '')),
            str(row.get('Mes Cumpleaños', '')),
            float(limpiar_numero(row.get('Deuda Total', 0.0))),
            float(limpiar_numero(row.get('Deuda Mora', 0.0))),
            float(limpiar_numero(row.get('Credito Total', 0.0))),
            float(limpiar_numero(row.get('Credito Disponible', 0.0))),
            int(limpiar_numero(row.get('Ped. Pendientes', 0))),
            int(limpiar_numero(row.get('Ped. Mora', 0))),
            str(row.get('celular', '')),
            str(row.get('Correo', '')),
            str(row.get('Dpto - Residencia', '')),
            str(row.get('Ciudad - Residencia', '')),
            str(row.get('Barrio - Residencia', '')),
            str(row.get('Dirección - Residencia', '')),
            str(row.get('Complemento - Residencia', '')),
            str(row.get('Referencia - Residencia', '')),
            str(row.get('Dpto - Entrega', '')),
            str(row.get('Ciudad - Entrega', '')),
            str(row.get('Barrio - Entrega', '')),
            str(row.get('Dirección - Entrega', '')),
            str(row.get('Complemento - Entrega', '')),
            str(row.get('Referencia - Entrega', '')),
            int(limpiar_numero(row.get('Tiempo de Casa (Ciclo)', 0))),
            str(row.get('Origen CB', '')),
            str(row.get('Notas / Comentarios Líder', '')),
            str(row.get('Indicador', ''))
        ))
    
    conn.commit()
    if close_at_end:
        conn.close()
    return True

def sincronizar_excel_metas_a_sqlite(df_metas, conn=None):
    """
    Convierte y vuelca el DataFrame de metas hacia la tabla metas_como_vamos en SQLite.
    """
    if df_metas is None or df_metas.empty:
        return False
    
    close_at_end = False
    if conn is None:
        conn = obtener_conexion_db()
        close_at_end = True
    
    col_sec = 'Nombre Setor' if 'Nombre Setor' in df_metas.columns else 'Sector'
    if not df_metas[col_sec].dropna().empty:
        sec_val = str(df_metas[col_sec].dropna().iloc[0]).strip()
        cursor.execute("DELETE FROM metas_como_vamos WHERE nombre_sector = ? OR nombre_sector LIKE ?", (sec_val, f"%{sec_val}%"))
    else:
        cursor.execute("DELETE FROM metas_como_vamos")
    
    col_cb = 'Código de consultora' if 'Código de consultora' in df_metas.columns else 'Cd Consultora'
    col_nom = 'Nombre de consultora' if 'Nombre de consultora' in df_metas.columns else 'Nombre Consultora'
    col_ger = 'Nombre Gerencia' if 'Nombre Gerencia' in df_metas.columns else 'Gerencia'
    col_grp = 'Código de grupo' if 'Código de grupo' in df_metas.columns else 'Cód. Grupo'
    
    for _, row in df_metas.iterrows():
        cb = str(row.get(col_cb, '')).strip()
        cursor.execute("""
        INSERT INTO metas_como_vamos (
            codigo_cb, nombre_consultora, nombre_gerencia, nombre_sector, codigo_grupo, color,
            obj_facturacion, real_facturacion, cump_facturacion, obj_activas, real_activas, cump_activas,
            saldo, disponibles, inicios, reinicios, recuperos, ganancia_estimada
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            cb,
            str(row.get(col_nom, '')),
            str(row.get(col_ger, '')),
            str(row.get(col_sec, '')),
            str(int(limpiar_numero(row.get(col_grp, 0)))) if limpiar_numero(row.get(col_grp, 0)) > 0 else str(row.get(col_grp, '')),
            str(row.get('Color', '')),
            float(limpiar_numero(row.get('Objetivo Facturación', 0.0))),
            float(limpiar_numero(row.get('Real Facturación', 0.0))),
            float(limpiar_numero(row.get('Cumplimiento Facturación', 0.0))),
            float(limpiar_numero(row.get('Objetivo Activas', 0.0))),
            float(limpiar_numero(row.get('Real Activas', 0.0))),
            float(limpiar_numero(row.get('Cumplimiento Activas', 0.0))),
            float(limpiar_numero(row.get('Saldo', 0.0))),
            int(limpiar_numero(row.get('Disponibles', 0))),
            int(limpiar_numero(row.get('Inicios', 0))),
            int(limpiar_numero(row.get('Reinicios', 0))),
            int(limpiar_numero(row.get('Recuperos', 0))),
            float(limpiar_numero(row.get('Ganancia estimada', 0.0)))
        ))
        
    conn.commit()
    if close_at_end:
        conn.close()
    return True

def consultar_tableau_sql(grupo=None, sector=None):
    """
    Ejecuta consulta SQL ultra-rápida indexada sobre consultoras_tableau en base_matices.db,
    filtrando por grupo de líder o sector de gerencia.
    """
    conn = obtener_conexion_db()
    try:
        conn.cursor().execute("ALTER TABLE consultoras_tableau ADD COLUMN indicador TEXT")
    except Exception:
        pass
    query = """
    SELECT 
        codigo_cb AS 'Código CB',
        nombre AS 'Asesora / Consultora',
        documento_gpp AS 'DocumentoGPP',
        cod_gerencia AS 'Cod. Gerencia',
        gerencia AS 'Gerencia',
        cod_sector AS 'Cod. Sector',
        sector AS 'Sector',
        grupo AS 'Grupo',
        ciclo AS 'Ciclo',
        color AS 'Nivel / Color',
        ascenso_cb AS 'Ascenso CB',
        situacion AS 'Situación',
        sit_comercial AS 'Sit. Comercial',
        ciclos_inactividad AS 'Ciclos Inactividad',
        actividad_convergencia AS 'Actividad Convergencia',
        inicios_completos AS 'Inicios Completos',
        fact_vol AS 'Fact. VOL',
        fact_natura AS 'Fact. Natura',
        fact_avon AS 'Fact. AVON',
        fact_ce AS 'Fact. C&E',
        facturacion_total AS 'Fact. Total',
        pts_natura AS 'Pts Natura',
        pts_avon AS 'Pts AVON',
        pts_total_vd AS 'Pts Total VD',
        pts_vol AS 'Pts VOL',
        pts_acum AS 'Pts Acum',
        pts_mant AS 'Pts Mant',
        pts_asc AS 'Pts Asc',
        tienda_online AS 'Tienda Online',
        tienda_sf AS 'TiendaSF',
        ciclo_primer_pedido AS 'Ciclo Primer Pedido',
        fecha_nacimiento AS 'Fecha De Nacimiento',
        mes_cumpleanos AS 'Mes Cumpleaños',
        deuda_total AS 'Deuda Total',
        deuda_mora AS 'Deuda Mora',
        credito_total AS 'Credito Total',
        credito_disponible AS 'Credito Disponible',
        pedidos_pendientes AS 'Ped. Pendientes',
        pedidos_mora AS 'Ped. Mora',
        celular AS 'celular',
        correo AS 'Correo',
        dpto_residencia AS 'Dpto - Residencia',
        ciudad_residencia AS 'Ciudad - Residencia',
        barrio_residencia AS 'Barrio - Residencia',
        direccion_residencia AS 'Dirección - Residencia',
        complemento_residencia AS 'Complemento - Residencia',
        referencia_residencia AS 'Referencia - Residencia',
        dpto_entrega AS 'Dpto - Entrega',
        ciudad_entrega AS 'Ciudad - Entrega',
        barrio_entrega AS 'Barrio - Entrega',
        direccion_entrega AS 'Dirección - Entrega',
        complemento_entrega AS 'Complemento - Entrega',
        referencia_entrega AS 'Referencia - Entrega',
        tiempo_casa AS 'Tiempo de Casa (Ciclo)',
        origen_cb AS 'Origen CB',
        notas_lider AS 'Notas / Comentarios Líder',
        indicador AS 'Indicador'
    FROM consultoras_tableau
    """
    where_clauses = []
    params = []
    
    if grupo:
        where_clauses.append("(grupo LIKE ? OR sector LIKE ?)")
        params.extend([f"%{grupo}%", f"%{grupo}%"])
        
    if sector and str(sector).strip():
        sec_str = str(sector).strip()
        where_clauses.append("(cod_sector = ? OR sector LIKE ?)")
        params.extend([sec_str, f"%{sec_str}%"])
        
    if where_clauses:
        query += " WHERE " + " AND ".join(where_clauses)
    
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()

    # Sincronizar dinámicamente con comentarios_lideres.json para asegurar consistencia total
    if not df.empty and 'Código CB' in df.columns:
        comentarios = cargar_comentarios_lideres()
        if comentarios:
            df['__cb_clean_sync__'] = df['Código CB'].apply(limpiar_codigo_cb_estandar)
            for idx, row in df.iterrows():
                cb_clean_val = row['__cb_clean_sync__']
                if cb_clean_val in comentarios:
                    nota_guardada = str(comentarios[cb_clean_val]).strip()
                    if nota_guardada:
                        df.at[idx, 'Notas / Comentarios Líder'] = nota_guardada
            df = df.drop(columns=['__cb_clean_sync__'], errors='ignore')

    # Añadir alias de columnas para máxima compatibilidad con las pestañas de app.py
    if 'Código CB' in df.columns and 'Codigo CB' not in df.columns:
        df['Codigo CB'] = df['Código CB']
    if 'Asesora / Consultora' in df.columns and 'Nombre' not in df.columns:
        df['Nombre'] = df['Asesora / Consultora']
    if 'Nivel / Color' in df.columns and 'Color' not in df.columns:
        df['Color'] = df['Nivel / Color']
    if 'Notas / Comentarios Líder' in df.columns and 'Comentarios_Lider' not in df.columns:
        df['Comentarios_Lider'] = df['Notas / Comentarios Líder']

    return df

def extraer_telefonos_colombia(val_raw):
    """
    Limpia el teléfono móvil eliminando ceros a la izquierda hasta encontrar
    el primer dígito '3' (formato estándar móvil Colombia de 10 dígitos: 3XXXXXXXXX).
    Si contiene más de 10 dígitos o números concatenados, extrae el segundo en 'movil2'.
    Retorna (movil1: str, movil2: str).
    """
    if val_raw is None:
        return "", ""
        
    s = str(val_raw).strip()
    if '.' in s:
        s = s.split('.')[0]
        
    digitos = re.sub(r'\D', '', s)
    if not digitos:
        return "", ""
        
    if digitos.startswith('57') and len(digitos) in [12, 24] and len(digitos) >= 3 and digitos[2] == '3':
        digitos = digitos[2:]

    pos_3 = digitos.find('3')
    if pos_3 != -1:
        digitos_utiles = digitos[pos_3:]
    else:
        digitos_utiles = digitos.lstrip('0')
        
    matches = re.findall(r'3\d{9}', digitos_utiles)
    
    movil1 = ""
    movil2 = ""
    
    if matches:
        movil1 = matches[0]
        if len(matches) > 1:
            movil2 = matches[1] if matches[1] != movil1 else ""
    else:
        if digitos_utiles.startswith('3'):
            movil1 = digitos_utiles[:10]
            resto = digitos_utiles[10:]
            p2 = resto.find('3')
            if p2 != -1 and len(resto[p2:]) >= 10:
                c2 = resto[p2:p2+10]
                movil2 = c2 if c2 != movil1 else ""
        elif len(digitos_utiles) >= 10:
            movil1 = digitos_utiles[:10]
            if len(digitos_utiles) >= 20:
                c2 = digitos_utiles[10:20]
                movil2 = c2 if c2 != movil1 else ""
        else:
            movil1 = digitos_utiles

    if movil1 == movil2:
        movil2 = ""

    return movil1, movil2

# --- MOTOR DE CRÉDITO & COBRANZA (Geral.xlsx) ---

def sincronizar_excel_geral_a_sqlite(origen_file="Geral.xlsx", sector_esperado=None, conn=None):
    """
    Lee el archivo Geral.xlsx (o buffer subido), valida sus columnas esenciales,
    y guarda o reemplaza los registros en la tabla relacional 'cartera_geral' de SQLite.
    Retorna (exito: bool, num_registros: int, mensaje: str).
    """
    import unicodedata
    
    if origen_file is None:
        return False, 0, "No se proporcionó ningún archivo para procesar."
        
    try:
        if isinstance(origen_file, str):
            if not os.path.exists(origen_file):
                return False, 0, f"El archivo '{origen_file}' no existe en el disco."
            df_raw = pd.read_excel(origen_file, sheet_name=0)
        else:
            df_raw = pd.read_excel(origen_file, sheet_name=0)
    except Exception as e_read:
        return False, 0, f"Error al abrir el archivo Excel: {e_read}"

    if df_raw is None or df_raw.empty:
        return False, 0, "El archivo subido no contiene registros válidos."

    # Normalizar nombres de columnas
    def _norm_txt(t):
        if not t:
            return ""
        nfkd = unicodedata.normalize('NFKD', str(t))
        return "".join([c for c in nfkd if not unicodedata.combining(c)]).lower().strip()

    clean_cols = {}
    for col in df_raw.columns:
        norm = _norm_txt(col)
        norm = "".join(c if c.isalnum() or c.isspace() else " " for c in norm)
        
        if 'titulo' in norm and 'valor' not in norm:
            clean_cols[col] = 'titulo'
        elif 'cuota' in norm:
            clean_cols[col] = 'cuota'
        elif 'pedido' in norm and ('numero' in norm or 'num' in norm):
            clean_cols[col] = 'numero_pedido'
        elif 'factura' in norm and ('numero' in norm or 'num' in norm):
            clean_cols[col] = 'numero_factura'
        elif 'captacion' in norm:
            clean_cols[col] = 'ciclo_captacion'
        elif 'indicador' in norm:
            clean_cols[col] = 'ciclo_indicador'
        elif 'grupo' in norm:
            clean_cols[col] = 'grupo'
        elif 'fechapedido' in norm or ('fecha' in norm and 'pedido' in norm):
            clean_cols[col] = 'fecha_pedido'
        elif 'vencimiento' in norm and 'orig' in norm:
            clean_cols[col] = 'fecha_vencimiento_original'
        elif 'vencimiento' in norm and 'orig' not in norm:
            clean_cols[col] = 'fecha_vencimiento'
        elif 'valor' in norm and 'titulo' in norm:
            clean_cols[col] = 'valor_titulo'
        elif 'saldo' in norm and 'principal' in norm:
            clean_cols[col] = 'saldo_principal'
        elif 'saldo' in norm and 'financiero' in norm:
            clean_cols[col] = 'saldo_financiero'
        elif 'saldo' in norm and 'total' in norm:
            clean_cols[col] = 'saldo_total'
        elif 'saldo' in norm and 'actualizado' in norm:
            clean_cols[col] = 'saldo_actualizado'
        elif 'situacion' in norm:
            clean_cols[col] = 'situacion'
        elif 'retraso' in norm:
            clean_cols[col] = 'dias_retraso'
        elif 'fase' in norm:
            clean_cols[col] = 'fase_cobro'
        elif 'estructura' in norm and 'padre' not in norm and ('cod' in norm or 'cd' in norm):
            clean_cols[col] = 'cod_sector'
        elif 'estructura' in norm and 'padre' not in norm and 'cod' not in norm:
            clean_cols[col] = 'sector'
        elif 'persona' in norm and ('codigo' in norm or 'cod' in norm):
            clean_cols[col] = 'codigo_cb'
        elif 'nombre' in norm:
            clean_cols[col] = 'nombre'
        elif 'direccion' in norm:
            clean_cols[col] = 'direccion'
        elif 'movil' in norm or 'celular' in norm:
            clean_cols[col] = 'telefono_movil'
        elif 'residencial' in norm or 'telefono' in norm:
            if 'telefono_movil' not in clean_cols.values():
                clean_cols[col] = 'telefono_movil'
        elif 'correo' in norm or 'email' in norm:
            clean_cols[col] = 'correo'
        elif 'plan' in norm:
            clean_cols[col] = 'plan_recibimiento'
        elif 'origen' in norm:
            clean_cols[col] = 'origen_empresa'

    df = df_raw.rename(columns=clean_cols)
    
    # Validar columnas mínimas requeridas
    columnas_minimas = ['titulo', 'numero_factura', 'fecha_vencimiento', 'saldo_total', 'situacion', 'nombre']
    faltantes = [c for c in columnas_minimas if c not in df.columns]
    if faltantes:
        msg_guia = (
            f"El archivo no contiene las columnas requeridas: `{', '.join(faltantes)}`.\n\n"
            f"📌 **Guía paso a paso para descargar el archivo correcto desde Geral:**\n\n"
            f"1️⃣ Ingresa a **Geral** ➔ **Crédito & Cobranza**\n"
            f"2️⃣ Selecciona **Consultar Deuda**\n"
            f"3️⃣ En **Ciclo de Captación**, selecciona los ciclos a consultar\n"
            f"4️⃣ Haz clic en el botón **Consultar**\n"
            f"5️⃣ Presiona **Exportar Listado** ➔ Selecciona **Excel Inmediata**"
        )
        return False, 0, msg_guia

    # Validar sector si se especificó
    if sector_esperado and 'cod_sector' in df.columns:
        sec_esp_str = str(sector_esperado).strip()
        secs_encontrados = [str(x).strip().split('.')[0] for x in df['cod_sector'].dropna().unique() if str(x).strip()]
        
        # Códigos válidos aceptados (código completo o código corto)
        codigos_validos = {sec_esp_str}
        if len(sec_esp_str) > 3 and sec_esp_str.startswith("700000"):
            codigos_validos.add(sec_esp_str[6:])
            if sec_esp_str[6:].isdigit():
                codigos_validos.add(str(int(sec_esp_str[6:])))
                
        if secs_encontrados and not any(s in codigos_validos for s in secs_encontrados):
            return False, 0, f"❌ El archivo subido pertenece al Sector '{secs_encontrados[0]}', el cual no coincide con tu Sector asignado ({sec_esp_str})."

    close_at_end = False
    if conn is None:
        conn = obtener_conexion_db()
        close_at_end = True
        
    cursor = conn.cursor()
    try:
        cursor.execute("ALTER TABLE cartera_geral ADD COLUMN telefono_movil_2 TEXT")
        conn.commit()
    except Exception:
        pass
    
    # Si viene con sector específico, borrar solo los registros de ese sector
    if 'cod_sector' in df.columns and not df['cod_sector'].dropna().empty:
        secs_a_borrar = [str(x).strip().split('.')[0] for x in df['cod_sector'].dropna().unique() if str(x).strip()]
        for cod_sec_val in secs_a_borrar:
            cursor.execute("DELETE FROM cartera_geral WHERE cod_sector = ? OR cod_sector = ? OR sector LIKE ?", (cod_sec_val, f"700000{cod_sec_val}", f"%{cod_sec_val}%"))
    else:
        cursor.execute("DELETE FROM cartera_geral")
        
    fecha_carga_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Mapeo de grupos y sectores reales desde consultoras_tableau por codigo_cb
    mapa_tableau_geral = {}
    try:
        cursor.execute("SELECT codigo_cb, grupo, cod_sector, sector, nombre, celular FROM consultoras_tableau")
        for r_cb, r_grp, r_csec, r_sec, r_nom, r_cel in cursor.fetchall():
            cb_k = limpiar_codigo_cb_estandar(r_cb)
            if cb_k:
                mapa_tableau_geral[cb_k] = {
                    'grupo': str(r_grp).strip().split('.')[0] if r_grp else '',
                    'cod_sector': str(r_csec).strip().split('.')[0] if r_csec else '',
                    'sector': str(r_sec).strip() if r_sec else '',
                    'nombre': str(r_nom).strip() if r_nom else '',
                    'celular': str(r_cel).strip() if r_cel else ''
                }
    except Exception as e_map:
        safe_print(f"Nota: no se pudo extraer mapa previo de consultoras_tableau para Geral: {e_map}")
    
    registros_insertar = []
    for _, row in df.iterrows():
        tit = str(row.get('titulo', '')).strip().split('.')[0]
        if not tit or tit.lower() in ['nan', 'none', '']:
            continue
            
        f_venc = str(row.get('fecha_vencimiento', ''))[:10] if pd.notna(row.get('fecha_vencimiento')) else ''
        f_orig = str(row.get('fecha_vencimiento_original', ''))[:10] if pd.notna(row.get('fecha_vencimiento_original')) else ''
        f_ped = str(row.get('fecha_pedido', ''))[:10] if pd.notna(row.get('fecha_pedido')) else ''
        
        # Limpiar teléfono con extraer_telefonos_colombia
        tel_raw = str(row.get('telefono_movil', '')).strip()
        m1, m2 = extraer_telefonos_colombia(tel_raw)
        
        # Limpiar números
        val_tit = float(limpiar_numero(row.get('valor_titulo', 0.0)))
        s_prin = float(limpiar_numero(row.get('saldo_principal', 0.0)))
        s_fin = float(limpiar_numero(row.get('saldo_financiero', 0.0)))
        s_tot = float(limpiar_numero(row.get('saldo_total', 0.0)))
        if s_tot == 0.0 and s_prin > 0.0:
            s_tot = s_prin + s_fin
        s_act = float(limpiar_numero(row.get('saldo_actualizado', s_tot)))
        
        cb_raw = str(row.get('codigo_cb', '')).strip().split('.')[0]
        cb_clean = limpiar_codigo_cb_estandar(cb_raw)

        # Determinar grupo y sector reales cruzando con Tableau
        grp = str(row.get('grupo', '')).strip().split('.')[0]
        cod_sec = str(row.get('cod_sector', '')).strip().split('.')[0]
        sec_nom = str(row.get('sector', '')).strip()
        nom_cb = str(row.get('nombre', '')).strip()

        if cb_clean in mapa_tableau_geral:
            info_t = mapa_tableau_geral[cb_clean]
            if info_t.get('grupo'):
                grp = info_t['grupo']
            if info_t.get('cod_sector'):
                cod_sec = info_t['cod_sector']
            if info_t.get('sector'):
                sec_nom = info_t['sector']
            if info_t.get('nombre') and not nom_cb:
                nom_cb = info_t['nombre']
            if info_t.get('celular') and not m1:
                m1, m2 = extraer_telefonos_colombia(info_t['celular'])
        
        registros_insertar.append((
            tit,
            int(limpiar_numero(row.get('cuota', 1))),
            str(row.get('numero_pedido', '')).strip().split('.')[0],
            str(row.get('ciclo_captacion', '')).strip(),
            str(row.get('ciclo_indicador', '')).strip(),
            grp,
            str(row.get('numero_factura', '')).strip().split('.')[0],
            f_ped,
            f_orig,
            f_venc,
            val_tit,
            s_prin,
            s_fin,
            s_tot,
            s_act,
            str(row.get('situacion', 'Pendiente')).strip(),
            float(limpiar_numero(row.get('dias_retraso', 0.0))),
            str(row.get('fase_cobro', '')).strip(),
            cod_sec,
            sec_nom,
            cb_clean if cb_clean else cb_raw,
            nom_cb,
            str(row.get('direccion', '')).strip(),
            m1,
            m2,
            str(row.get('correo', '')).strip(),
            str(row.get('plan_recibimiento', '')).strip(),
            str(row.get('origen_empresa', 'Natura')).strip(),
            fecha_carga_actual
        ))
        
    cursor.executemany("""
    INSERT INTO cartera_geral (
        titulo, cuota, numero_pedido, ciclo_captacion, ciclo_indicador, grupo, numero_factura,
        fecha_pedido, fecha_vencimiento_original, fecha_vencimiento, valor_titulo, saldo_principal,
        saldo_financiero, saldo_total, saldo_actualizado, situacion, dias_retraso, fase_cobro,
        cod_sector, sector, codigo_cb, nombre, direccion, telefono_movil, telefono_movil_2, correo,
        plan_recibimiento, origen_empresa, fecha_carga
    ) VALUES (
        ?, ?, ?, ?, ?, ?, ?,
        ?, ?, ?, ?, ?,
        ?, ?, ?, ?, ?, ?,
        ?, ?, ?, ?, ?, ?, ?, ?,
        ?, ?, ?
    )
    """, registros_insertar)
    
    conn.commit()
    if close_at_end:
        conn.close()
        
    return True, len(registros_insertar), f"¡Se sincronizaron exitosamente {len(registros_insertar)} títulos comerciales en la base de Crédito & Cobranza con cruce de grupo de líder!"

def consultar_geral_sql(grupo=None, sector=None, situacion=None):
    """
    Consulta la base relacional de cartera_geral en SQLite con cruce dinámico hacia consultoras_tableau
    para garantizar que cada consultora refleje su verdadero grupo de líder (y no el código de sector 5515/5522).
    """
    try:
        inicializar_db_sqlite()
    except Exception:
        pass

    conn = obtener_conexion_db()
    
    # Asegurar que la tabla exista
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='cartera_geral'")
        if not cursor.fetchone():
            inicializar_db_sqlite()
    except Exception:
        pass

    query = """
    SELECT 
        cg.id, cg.titulo, cg.cuota, cg.numero_pedido, cg.numero_factura,
        cg.ciclo_captacion, cg.ciclo_indicador,
        COALESCE(NULLIF(ct.grupo, ''), NULLIF(cg.grupo, ''), 'Sin Grupo') AS grupo,
        cg.fecha_pedido, cg.fecha_vencimiento_original, cg.fecha_vencimiento,
        cg.valor_titulo, cg.saldo_principal, cg.saldo_financiero, cg.saldo_total,
        cg.saldo_actualizado, cg.situacion, cg.dias_retraso, cg.fase_cobro,
        COALESCE(NULLIF(ct.cod_sector, ''), cg.cod_sector) AS cod_sector,
        COALESCE(NULLIF(ct.sector, ''), cg.sector) AS sector,
        cg.codigo_cb,
        COALESCE(NULLIF(ct.nombre, ''), cg.nombre) AS nombre,
        cg.direccion,
        COALESCE(NULLIF(ct.celular, ''), cg.telefono_movil) AS telefono_movil,
        cg.telefono_movil_2, cg.correo, cg.plan_recibimiento, cg.origen_empresa, cg.fecha_carga
    FROM cartera_geral cg
    LEFT JOIN consultoras_tableau ct ON CAST(cg.codigo_cb AS TEXT) = CAST(ct.codigo_cb AS TEXT)
    """
    where_clauses = []
    params = []
    
    if situacion:
        where_clauses.append("cg.situacion = ?")
        params.append(str(situacion).strip())
        
    if grupo:
        grp_str = str(grupo).strip()
        where_clauses.append("(COALESCE(NULLIF(ct.grupo, ''), cg.grupo) = ? OR COALESCE(NULLIF(ct.grupo, ''), cg.grupo) LIKE ?)")
        params.extend([grp_str, f"%{grp_str}%"])
        
    if sector and str(sector).strip():
        sec_str = str(sector).strip()
        where_clauses.append("(COALESCE(NULLIF(ct.cod_sector, ''), cg.cod_sector) = ? OR COALESCE(NULLIF(ct.sector, ''), cg.sector) LIKE ?)")
        params.extend([sec_str, f"%{sec_str}%"])
        
    if where_clauses:
        query += " WHERE " + " AND ".join(where_clauses)
        
    query += " ORDER BY cg.fecha_vencimiento ASC, cg.saldo_total DESC"
    
    try:
        df = pd.read_sql_query(query, conn, params=params)
    except Exception as e:
        inicializar_db_sqlite()
        try:
            df = pd.read_sql_query(query, conn, params=params)
        except Exception:
            df = pd.DataFrame()
    finally:
        conn.close()
        
    # Si la consulta viene vacía y existe Geral.xlsx en disco, auto-sincronizar
    if (df is None or df.empty) and os.path.exists("Geral.xlsx"):
        try:
            sincronizar_excel_geral_a_sqlite("Geral.xlsx")
            conn2 = obtener_conexion_db()
            df = pd.read_sql_query(query, conn2, params=params)
            conn2.close()
        except Exception:
            pass
            
    return df

def procesar_analisis_geral_cobranza(df_geral, fecha_base=None):
    """
    Procesa el DataFrame de cartera_geral calculando proyecciones temporales,
    semáforos de mora y vencimiento preventivo, sumatorias financieras y datos para el heatmap.
    """
    if df_geral is None or df_geral.empty:
        return {
            'df_completo': pd.DataFrame(),
            'df_pendientes': pd.DataFrame(),
            'df_vence_manana': pd.DataFrame(),
            'df_pasado_manana': pd.DataFrame(),
            'df_en_mora': pd.DataFrame(),
            'df_proximos_7d': pd.DataFrame(),
            'kpis': {
                'total_cartera': 0.0,
                'total_facturas_pendientes': 0,
                'total_mora': 0.0,
                'facturas_mora': 0,
                'total_manana': 0.0,
                'facturas_manana': 0,
                'total_7d': 0.0,
                'facturas_7d': 0,
                'consultoras_unicas': 0
            },
            'heatmap_data': pd.DataFrame()
        }
        
    df = df_geral.copy()
    hoy = fecha_base if fecha_base is not None else date.today()
    hoy_ts = pd.to_datetime(hoy)

    # Limpieza estandarizada de teléfonos móviles (10 dígitos Colombia empezando por 3 y segundo móvil)
    if 'telefono_movil' in df.columns:
        tels_tuple = df['telefono_movil'].apply(extraer_telefonos_colombia)
        df['telefono_movil'] = [t[0] for t in tels_tuple]
        if 'telefono_movil_2' not in df.columns:
            df['telefono_movil_2'] = [t[1] for t in tels_tuple]
        else:
            df['telefono_movil_2'] = df['telefono_movil_2'].fillna('').astype(str).replace('None', '').replace('nan', '')
            # Si telefono_movil_2 está vacío y la tupla trajo un segundo móvil, asignarlo
            df['telefono_movil_2'] = [t[1] if (not m2 or m2 == '') else m2 for t, m2 in zip(tels_tuple, df['telefono_movil_2'])]

    # Conversión de fecha de vencimiento
    df['fecha_venc_dt'] = pd.to_datetime(df['fecha_vencimiento'], errors='coerce')
    df['dias_para_vencer'] = (df['fecha_venc_dt'] - hoy_ts).dt.days
    
    # Clasificación de tramos
    def _clasificar(d, sit, saldo):
        if str(sit).strip().lower() == 'pagado' or saldo <= 0:
            return "Pagado / Sin Deuda"
        if pd.isna(d):
            return "Sin Fecha"
        d = int(d)
        if d < 0:
            return "Vencida / En Mora"
        elif d == 0:
            return "Vence Hoy"
        elif d == 1:
            return "Vence Mañana"
        elif d in [2, 3]:
            return "Pasado Mañana (+2 a +3 d)"
        elif 4 <= d <= 7:
            return "Próximos 4 a 7 días"
        elif 8 <= d <= 15:
            return "Próximos 8 a 15 días"
        else:
            return "Más de 15 días"

    df['tramo_vencimiento'] = df.apply(lambda r: _clasificar(r['dias_para_vencer'], r.get('situacion'), r.get('saldo_total', 0)), axis=1)
    
    # Cartera viva y pendiente (excluye pagados y saldo $0)
    df_pendientes = df[(df['situacion'] == 'Pendiente') & (df['saldo_total'] > 0)].copy()
    
    # Sub-segmentos
    df_manana = df_pendientes[df_pendientes['dias_para_vencer'] == 1].copy()
    df_pasado = df_pendientes[df_pendientes['dias_para_vencer'].isin([2, 3])].copy()
    df_mora = df_pendientes[df_pendientes['dias_para_vencer'] < 0].copy()
    df_7d = df_pendientes[(df_pendientes['dias_para_vencer'] >= 0) & (df_pendientes['dias_para_vencer'] <= 7)].copy()
    
    # KPIs
    kpis = {
        'total_cartera': float(df_pendientes['saldo_total'].sum()),
        'total_facturas_pendientes': len(df_pendientes),
        'total_mora': float(df_mora['saldo_total'].sum()),
        'facturas_mora': len(df_mora),
        'total_manana': float(df_manana['saldo_total'].sum()),
        'facturas_manana': len(df_manana),
        'total_7d': float(df_7d['saldo_total'].sum()),
        'facturas_7d': len(df_7d),
        'consultoras_unicas': df_pendientes['nombre'].nunique() if 'nombre' in df_pendientes.columns else len(df_pendientes)
    }
    
    # Datos para el Heatmap / Calendario térmico
    df_heat = df_pendientes.dropna(subset=['fecha_venc_dt']).copy()
    if not df_heat.empty:
        heatmap_data = df_heat.groupby('fecha_vencimiento').agg(
            Saldo_Total=('saldo_total', 'sum'),
            Total_Facturas=('titulo', 'count'),
            Consultoras=('nombre', 'nunique')
        ).reset_index().sort_values('fecha_vencimiento')
    else:
        heatmap_data = pd.DataFrame()
        
    return {
        'df_completo': df,
        'df_pendientes': df_pendientes,
        'df_vence_manana': df_manana,
        'df_pasado_manana': df_pasado,
        'df_en_mora': df_mora,
        'df_proximos_7d': df_7d,
        'kpis': kpis,
        'heatmap_data': heatmap_data
    }

def generar_mensaje_whatsapp_cobranza(row, tipo='manana', nombre_remitente='Tu Líder'):
    """
    Genera el texto formateado para el mensaje de WhatsApp de recordatorio de cobro.
    """
    nombre = str(row.get('nombre', 'Consultora')).split()[0].title()
    factura = str(row.get('numero_factura', '')).split('.')[0]
    pedido = str(row.get('numero_pedido', '')).split('.')[0]
    f_venc = str(row.get('fecha_vencimiento', ''))[:10]
    saldo_tot = float(row.get('saldo_total', 0.0))
    saldo_prin = float(row.get('saldo_principal', saldo_tot))
    saldo_fin = float(row.get('saldo_financiero', 0.0))
    dias_ret = int(row.get('dias_retraso', 0)) if pd.notna(row.get('dias_retraso')) else 0
    saldo_fmt = f"${saldo_tot:,.0f} COP".replace(",", ".")
    
    if tipo == 'manana':
        msg = (
            f"🌸 ¡Hola {nombre}! Te saluda {nombre_remitente} de Natura & Avon.\n\n"
            f"Queremos recordarte con mucho cariño que el día de *mañana ({f_venc})* vence tu factura *N° {factura}* (Pedido #{pedido}) por un valor de *{saldo_fmt}*.\n\n"
            f"💡 *Recuerda:* Realizar tu pago a tiempo te permite mantener tu crédito activo y seguir recibiendo tus pedidos sin retrasos. ✨\n\n"
            f"Puedes cancelar fácilmente por Nequi, Daviplata o PSE. ¡Que tengas un excelente día! 💕"
        )
    elif tipo == 'hoy':
        msg = (
            f"🚨 ¡Hola {nombre}! Te recordamos que *HOY ({f_venc})* es la fecha límite para el pago de tu factura *N° {factura}* por un valor de *{saldo_fmt}*.\n\n"
            f"Evita recargos financieros y bloqueos en tus próximos pedidos pagando hoy antes de las 9:00 PM. ✨\n\n"
            f"Cualquier duda con tu pago, con gusto te apoyo. — {nombre_remitente} 📲"
        )
    elif tipo == 'mora':
        msg = (
            f"⚠️ Estimada {nombre}, te informamos que tu factura *N° {factura}* presenta *{dias_ret} días de vencida* con un saldo pendiente de *{saldo_fmt}*.\n\n"
            f"📌 *Detalle de la obligación:*\n"
            f"• Saldo Capital: ${saldo_prin:,.0f} COP\n"
            f"• Saldo Financiero: ${saldo_fin:,.0f} COP\n"
            f"• Saldo Total a Pagar: *{saldo_fmt}*\n\n"
            f"Por favor reporta tu comprobante de pago a la brevedad para normalizar tu estado de cuenta. Estamos para apoyarte. — {nombre_remitente} 📲"
        )
    else: # preventivo general
        msg = (
            f"🌸 ¡Hola {nombre}! Te recordamos que tienes una factura programada para vencer el *{f_venc}* (Factura N° {factura}) por valor de *{saldo_fmt}*.\n\n"
            f"¡Muchos éxitos en tu negocio! ✨ — {nombre_remitente}"
        )
    return msg

def eliminar_datos_por_grupo_o_usuario(codigo_grupo, eliminar_cuenta=False):
    """
    Elimina todos los registros asociados a un código de grupo en las tablas SQLite
    (consultoras_tableau, metas_como_vamos, comentarios_lideres), en los archivos Excel locales
    (Base para el como vamos.xlsx) y opcionalmente su usuario en usuarios.json.
    """
    conn = obtener_conexion_db()
    cursor = conn.cursor()
    grp_str = str(codigo_grupo).strip()

    borrados = {}

    # 1. Borrar en consultoras_tableau
    try:
        cursor.execute("DELETE FROM consultoras_tableau WHERE TRIM(grupo) = ?", (grp_str,))
        borrados['consultoras_tableau'] = cursor.rowcount
    except Exception:
        borrados['consultoras_tableau'] = 0

    # 2. Borrar en metas_como_vamos
    try:
        cursor.execute("DELETE FROM metas_como_vamos WHERE TRIM(codigo_grupo) = ?", (grp_str,))
        borrados['metas_como_vamos'] = cursor.rowcount
    except Exception:
        borrados['metas_como_vamos'] = 0

    # 3. Borrar en comentarios_lideres
    try:
        cursor.execute("DELETE FROM comentarios_lideres WHERE TRIM(grupo) = ? OR TRIM(username) = ?", (grp_str, grp_str))
        borrados['comentarios_lideres'] = cursor.rowcount
    except Exception:
        borrados['comentarios_lideres'] = 0

    # 4. Borrar en usuarios (SQLite y usuarios.json)
    if eliminar_cuenta:
        try:
            cursor.execute("DELETE FROM usuarios WHERE TRIM(codigo_grupo) = ? OR TRIM(username) = ?", (grp_str, grp_str))
            borrados['usuarios_sqlite'] = cursor.rowcount
        except Exception:
            borrados['usuarios_sqlite'] = 0
        
        try:
            usuarios = cargar_usuarios()
            a_eliminar = [u for u, info in usuarios.items() if str(info.get('codigo_grupo')).strip() == grp_str or str(u).strip() == grp_str]
            for u in a_eliminar:
                del usuarios[u]
            guardar_usuarios(usuarios)
            borrados['usuarios_json'] = len(a_eliminar)
        except Exception:
            borrados['usuarios_json'] = 0

    conn.commit()
    conn.close()

    # 5. Filtrar y actualizar el archivo Excel físico 'Base para el como vamos.xlsx'
    try:
        if os.path.exists("Base para el como vamos.xlsx"):
            xl = pd.ExcelFile("Base para el como vamos.xlsx")
            with pd.ExcelWriter("Base para el como vamos.xlsx", engine="openpyxl") as writer:
                for sheet_name in xl.sheet_names:
                    df_s = pd.read_excel("Base para el como vamos.xlsx", sheet_name=sheet_name)
                    col_g = None
                    for c in df_s.columns:
                        if 'grupo' in str(c).lower() or 'codigo_grupo' in str(c).lower():
                            col_g = c
                            break
                    if col_g:
                        df_s = df_s[df_s[col_g].astype(str).str.strip() != grp_str]
                    df_s.to_excel(writer, sheet_name=sheet_name, index=False)
            borrados['excel_como_vamos'] = "Actualizado"
    except Exception as e_xl:
        borrados['excel_como_vamos'] = f"Nota: {e_xl}"

    return borrados

def vaciar_base_datos_completa(vaciar_usuarios=False, eliminar_archivos_excel=True):
    """
    Elimina todos los datos cargados de consultoras_tableau, metas_como_vamos, comentarios_lideres
    y los archivos Excel locales (Base para el como vamos.xlsx, Base de Datos.xlsx, etc.).
    Si vaciar_usuarios=True, resetea la lista de usuarios conservando solo las cuentas principales (superadmin/gerente).
    """
    conn = obtener_conexion_db()
    cursor = conn.cursor()

    res = {}
    try:
        cursor.execute("DELETE FROM consultoras_tableau")
        res['consultoras_tableau'] = cursor.rowcount
    except Exception:
        res['consultoras_tableau'] = 0

    try:
        cursor.execute("DELETE FROM metas_como_vamos")
        res['metas_como_vamos'] = cursor.rowcount
    except Exception:
        res['metas_como_vamos'] = 0

    try:
        cursor.execute("DELETE FROM comentarios_lideres")
        res['comentarios_lideres'] = cursor.rowcount
    except Exception:
        res['comentarios_lideres'] = 0

    if vaciar_usuarios:
        try:
            usuarios = cargar_usuarios()
            usuarios_filtrados = {u: info for u, info in usuarios.items() if info.get('rol') in ['superadmin', 'gerente']}
            guardar_usuarios(usuarios_filtrados)
            cursor.execute("DELETE FROM usuarios WHERE rol = 'lider'")
            res['usuarios_lideres_borrados'] = cursor.rowcount
        except Exception:
            res['usuarios_lideres_borrados'] = 0

    conn.commit()
    conn.close()

    # Eliminar archivos Excel locales para que no queden datos viejos en 'Cómo Vamos'
    if eliminar_archivos_excel:
        archivos_a_limpiar = [
            "Base para el como vamos.xlsx",
            "Base de Datos.xlsx",
            "Resultado_Metas_Procesadas.xlsx",
            "mi_grupo.xls",
            "activas.xlsx"
        ]
        borrados_archivos = []
        for arch in archivos_a_limpiar:
            if os.path.exists(arch):
                try:
                    os.remove(arch)
                    borrados_archivos.append(arch)
                except Exception:
                    pass
        res['archivos_excel_eliminados'] = borrados_archivos

    return res

# ---------------------------------------------------------
# MÓDULO DE CUMPLEAÑOS Y RECONOCIMIENTO DE ASESORAS / LÍDERES
# ---------------------------------------------------------
MESES_ESPANOL = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
    5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
    9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
}

PLANTILLA_CUMPLEANOS_DEFAULT = (
    "¡Hola {primer_nombre}! 🎂✨ Te deseo un muy Feliz Cumpleaños 🎉🎁. "
    "Que hoy pases un día maravilloso lleno de bendiciones, salud y grandes alegrías. "
    "¡Gracias por ser parte fundamental de nuestro equipo en Natura & Avon! "
    "Te mando un fuerte abrazo con mucho cariño. 💕 — Tu Líder {nombre_lider}"
)

def parse_dia_mes_fecha(fecha_val):
    """
    Extrae (día, mes) de cadenas de texto en diversos formatos:
    '15/10/1985', '1/5/1990', '1990-05-01', '1990/05/01 00:00:00', etc.
    """
    if pd.isna(fecha_val) or not str(fecha_val).strip():
        return None, None
    s = str(fecha_val).strip()
    if ' ' in s:
        s = s.split(' ')[0]
    for sep in ['/', '-']:
        if sep in s:
            parts = s.split(sep)
            if len(parts) >= 2:
                try:
                    if len(parts[0]) <= 2 and len(parts[1]) <= 2:
                        return int(parts[0]), int(parts[1])
                    elif len(parts[0]) == 4:
                        return int(parts[2]), int(parts[1])
                except Exception:
                    pass
    return None, None

def obtener_cumpleanos_equipo(df_tableau, user_nombre="Líder", plantilla_wa=None, fecha_referencia=None):
    """
    Filtra y clasifica las asesoras según su fecha de cumpleaños:
    - 'hoy': Asesoras que cumplen años exactamente hoy.
    - 'semana': Asesoras que cumplen en los próximos 7 días.
    - 'mes': Todas las asesoras que cumplen en el mes en curso.
    Genera el mensaje personalizado y enlace de WhatsApp directo para cada una.
    """
    if df_tableau is None or df_tableau.empty:
        return {'hoy': [], 'semana': [], 'mes': [], 'total_mes': 0, 'nombre_mes': ''}
    
    hoy = fecha_referencia if fecha_referencia else date.today()
    dia_actual = hoy.day
    mes_actual = hoy.month
    ano_actual = hoy.year
    nombre_mes = MESES_ESPANOL.get(mes_actual, 'Mes Actual')
    
    plantilla = plantilla_wa if (plantilla_wa and str(plantilla_wa).strip()) else PLANTILLA_CUMPLEANOS_DEFAULT
    
    col_fecha = next((c for c in ['Fecha De Nacimiento', 'fecha_nacimiento', 'Fecha Nacimiento'] if c in df_tableau.columns), None)
    col_nombre = next((c for c in ['Asesora / Consultora', 'nombre', 'Nombre de consultora', 'Nombre'] if c in df_tableau.columns), 'Nombre')
    col_cel = next((c for c in ['celular', 'Celular', 'Telefono', 'Teléfono'] if c in df_tableau.columns), 'celular')
    col_cb = next((c for c in ['Código CB', 'codigo_cb', 'Codigo CB'] if c in df_tableau.columns), 'Código CB')
    col_grupo = next((c for c in ['Grupo', 'grupo', 'Código de grupo'] if c in df_tableau.columns), 'Grupo')
    col_nivel = next((c for c in ['Nivel / Color', 'color', 'Color', 'Nivel'] if c in df_tableau.columns), 'Nivel / Color')
    col_sit = next((c for c in ['Sit. Comercial', 'sit_comercial', 'Situación'] if c in df_tableau.columns), 'Sit. Comercial')
    col_ped = next((c for c in ['Ped. Pendientes', 'pedidos_pendientes', 'Pedidos Pendientes'] if c in df_tableau.columns), None)
    col_deuda = next((c for c in ['Deuda Mora', 'deuda_mora', 'Deuda Total'] if c in df_tableau.columns), None)
    
    if not col_fecha:
        return {'hoy': [], 'semana': [], 'mes': [], 'total_mes': 0, 'nombre_mes': nombre_mes}
        
    registros_hoy = []
    registros_semana = []
    registros_mes = []
    
    for _, row in df_tableau.iterrows():
        f_val = row.get(col_fecha, '')
        d, m = parse_dia_mes_fecha(f_val)
        if not d or not m:
            continue
            
        nom = str(row.get(col_nombre, '')).strip()
        primer_n = nom.split()[0].title() if nom else "Consultora"
        nom_formateado = nom.title()
        cel = str(row.get(col_cel, '')).replace('.0', '').strip()
        cel_clean = "".join(ch for ch in cel if ch.isdigit())
        nivel_val = str(row.get(col_nivel, 'Consultora')).strip()
        sit_val = str(row.get(col_sit, '')).strip()
        ped_val = int(row.get(col_ped, 0)) if col_ped and pd.notna(row.get(col_ped)) else 0
        deuda_val = float(row.get(col_deuda, 0.0)) if col_deuda and pd.notna(row.get(col_deuda)) else 0.0
        
        # Formatear mensaje WhatsApp
        msg_wa = (
            plantilla
            .replace("{primer_nombre}", primer_n)
            .replace("{nombre}", nom_formateado)
            .replace("{nivel}", nivel_val)
            .replace("{nombre_lider}", str(user_nombre))
            .replace("{lider}", str(user_nombre))
        )
        link_wa = f"https://api.whatsapp.com/send?phone=57{cel_clean}&text={urllib.parse.quote(msg_wa)}" if len(cel_clean) >= 10 else ""
        
        item = {
            'codigo_cb': str(row.get(col_cb, '')),
            'nombre': nom_formateado,
            'primer_nombre': primer_n,
            'grupo': str(row.get(col_grupo, '')),
            'celular': cel_clean,
            'nivel': nivel_val if nivel_val else "Consultora",
            'sit_comercial': sit_val,
            'pedidos_pendientes': ped_val,
            'deuda_mora': deuda_val,
            'dia': d,
            'mes': m,
            'fecha_nacimiento_raw': str(f_val),
            'nombre_mes': MESES_ESPANOL.get(m, str(m)),
            'link_wa': link_wa,
            'msg_wa': msg_wa
        }
        
        # Cumpleaños Hoy
        if d == dia_actual and m == mes_actual:
            item['dias_falta'] = 0
            item['etiqueta_tiempo'] = "🎉 ¡HOY!"
            registros_hoy.append(item)
            
        # Cumpleaños en los próximos 7 días
        try:
            target_year = ano_actual if m >= mes_actual else ano_actual + 1
            fecha_cumple = date(target_year, m, d)
            diff = (fecha_cumple - hoy).days
            if 0 < diff <= 7:
                item['dias_falta'] = diff
                item['etiqueta_tiempo'] = "Mañana 🎁" if diff == 1 else f"En {diff} días ({d} {MESES_ESPANOL.get(m, '')[:3]})"
                registros_semana.append(item)
        except Exception:
            pass
            
        # Cumpleaños del Mes en curso
        if m == mes_actual:
            item['etiqueta_mes'] = f"Día {d} ({MESES_ESPANOL.get(m, '')})"
            registros_mes.append(item)
            
    registros_semana.sort(key=lambda x: x['dias_falta'])
    registros_mes.sort(key=lambda x: x['dia'])
    
    return {
        'hoy': registros_hoy,
        'semana': registros_semana,
        'mes': registros_mes,
        'total_mes': len(registros_mes),
        'nombre_mes': nombre_mes
    }

# Ejecutamos la función si se invoca el script directamente
if __name__ == "__main__":
    df_resultado = calcular_metas_ciclo()
    if df_resultado is not None:
        diagnostico = generar_analisis_como_vamos(df_resultado)
        print("=== DIAGNÓSTICO AUTOMÁTICO DE 'CÓMO VAMOS' ===")
        print(f"Facturación Real vs Obj : ${diagnostico['facturacion']['real']:,.0f} / ${diagnostico['facturacion']['objetivo']:,.0f} ({diagnostico['facturacion']['cumplimiento_pct']:.1f}%)")
        print(f"Productividad Promedio  : ${diagnostico['facturacion']['productividad_promedio']:,.0f} COP / activa")
        print(f"Tasa Conversión Disponib: {diagnostico['disponibles']['tasa_conversion_pct']:.1f}% ({diagnostico['activas']['real']:.0f} de {diagnostico['disponibles']['total_disponibles']:.0f})")
        print(f"Total Saldo Pendiente   : {diagnostico['saldos']['total_saldo']:.0f} pedidos/cartera ({diagnostico['saldos']['lideres_afectados']} líderes con saldo)")


